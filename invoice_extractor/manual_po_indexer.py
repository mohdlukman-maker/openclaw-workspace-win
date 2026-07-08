import argparse
import json
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SOURCE_ROOT = Path(r"D:\Purchase Oder")
DEFAULT_DB_PATH = BASE_DIR / "data" / "procurement_index_prototype.sqlite"
DEFAULT_REPORT_PATH = BASE_DIR / "data" / "procurement_index_prototype_report.json"
FULL_DB_PATH = BASE_DIR / "data" / "procurement_index.sqlite"
FULL_REPORT_PATH = BASE_DIR / "data" / "procurement_index_report.json"
FILENAME_RE = re.compile(
    r"^BFE PO (?P<category>\S+) (?P<month_code>\d{4}) (?P<po_number>\S+) (?P<supplier>.+)$",
    re.IGNORECASE,
)
PHONE_RE = re.compile(r"(?:\+?6?0|0)\d[\d\-\s]{6,}\d")
FIELD_LABELS = {
    "amount",
    "approved by",
    "date",
    "date require",
    "date request",
    "delivery order no",
    "description",
    "invoice",
    "invoice no",
    "item",
    "person to contact",
    "pr number",
    "project code",
    "project name",
    "quantity",
    "reference no",
    "requested by",
    "request by",
    "total",
    "unit price",
    "verified by",
}
SUPPLIER_ALIASES = {
    "MENG SOON HUAT": "MENG SOON HUAT ELECTRICAL SDN BHD",
}


@dataclass(frozen=True)
class SourceFile:
    path: Path
    sheet_names: tuple[str, ...]
    template_variant: str


def normalized_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalized_item_key(value: Any) -> str:
    text = normalized_text(value).upper()
    text = text.replace("²", "2")
    text = text.replace("×", "X")
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\s*/\s*", "/", text)
    text = re.sub(r"\b(\d+)\s*KG\s*/\s*BAG\b", r"\1KG/BAG", text)
    text = re.sub(r"\b(\d+)\s*BOX\b", r"\1 BOX", text)
    text = re.sub(r"\bX\s*(\d+)\s*KG\b", r"X \1KG", text)
    text = re.sub(r"\bMM\s*2\b", "MM2", text)
    return re.sub(r"\s+", " ", text).strip()


def normalized_supplier_key(value: Any) -> str:
    supplier = normalized_text(value).upper()
    return SUPPLIER_ALIASES.get(supplier, supplier)


def normalize_sheet_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def parse_filename(path: Path) -> dict[str, str]:
    match = FILENAME_RE.match(path.stem)
    if not match:
        return {"category": "", "month_code": "", "po_number": "", "supplier": ""}
    return {key: normalized_text(value) for key, value in match.groupdict().items()}


def workbook_sources(root: Path) -> list[SourceFile]:
    sources: list[SourceFile] = []
    for path in sorted(root.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_names = tuple(workbook.sheetnames)
        workbook.close()
        sources.append(SourceFile(path=path, sheet_names=sheet_names, template_variant=" | ".join(sheet_names)))
    return sources


def choose_sample_sources(root: Path) -> list[SourceFile]:
    sources = workbook_sources(root)
    selected: list[SourceFile] = []
    selected_paths: set[Path] = set()

    def add_matching(limit: int, predicate: Any) -> None:
        for source in sources:
            if len([item for item in selected if predicate(item)]) >= limit:
                return
            if source.path in selected_paths or not predicate(source):
                continue
            selected.append(source)
            selected_paths.add(source.path)

    add_matching(5, lambda source: " TECH " in source.path.name and "TUJU GALAKSI" in source.path.name)
    add_matching(3, lambda source: " ELEC " in source.path.name and "MENG SOON" in source.path.name)

    common_variants = {
        "PURCHASEORDER | MATERIAL REQUISITION",
        "PURCHASE ORDER | MATERIAL REQUISITION",
        "Purchase Order | Material Requisition | Phone No",
    }
    seen_variants = {source.template_variant for source in selected}
    for source in sources:
        if len(selected) >= 10:
            break
        if source.path in selected_paths:
            continue
        if source.template_variant in common_variants or source.template_variant in seen_variants:
            continue
        selected.append(source)
        selected_paths.add(source.path)
        seen_variants.add(source.template_variant)
    return selected


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS import_batches (
            id INTEGER PRIMARY KEY,
            batch_name TEXT NOT NULL,
            source_root TEXT NOT NULL,
            started_at TEXT NOT NULL,
            finished_at TEXT,
            status TEXT NOT NULL,
            notes TEXT
        );

        CREATE TABLE IF NOT EXISTS source_documents (
            id INTEGER PRIMARY KEY,
            batch_id INTEGER NOT NULL,
            source_path TEXT NOT NULL UNIQUE,
            file_name TEXT NOT NULL,
            file_stem TEXT NOT NULL,
            file_ext TEXT NOT NULL,
            file_size INTEGER NOT NULL,
            modified_at TEXT NOT NULL,
            year_folder TEXT,
            month_code TEXT,
            category TEXT,
            supplier_name_from_filename TEXT,
            po_number_from_filename TEXT,
            has_pdf_pair INTEGER NOT NULL,
            template_variant TEXT NOT NULL,
            parse_status TEXT NOT NULL,
            review_status TEXT NOT NULL,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY,
            canonical_name TEXT NOT NULL UNIQUE,
            raw_name TEXT NOT NULL,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY,
            name TEXT,
            phone TEXT,
            raw_text TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS purchase_orders (
            id INTEGER PRIMARY KEY,
            source_document_id INTEGER NOT NULL,
            po_reference TEXT,
            po_number TEXT,
            po_category TEXT,
            po_date TEXT,
            invoice_no TEXT,
            delivery_order_no TEXT,
            supplier_id INTEGER,
            project_site TEXT,
            person_to_contact_id INTEGER,
            total_amount REAL,
            source_sheet TEXT,
            confidence REAL NOT NULL,
            review_status TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS material_requisitions (
            id INTEGER PRIMARY KEY,
            source_document_id INTEGER NOT NULL,
            purchase_order_id INTEGER,
            mr_reference TEXT,
            project_code TEXT,
            project_name TEXT,
            date_request TEXT,
            requested_by_contact_id INTEGER,
            source_sheet TEXT,
            confidence REAL NOT NULL,
            review_status TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY,
            canonical_name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS item_aliases (
            id INTEGER PRIMARY KEY,
            item_id INTEGER NOT NULL,
            alias_text TEXT NOT NULL UNIQUE,
            confidence REAL NOT NULL,
            approved INTEGER NOT NULL
        );

        CREATE TABLE IF NOT EXISTS line_items (
            id INTEGER PRIMARY KEY,
            purchase_order_id INTEGER,
            material_requisition_id INTEGER,
            row_number INTEGER,
            raw_description TEXT NOT NULL,
            normalized_item_id INTEGER,
            quantity_raw TEXT,
            quantity_value REAL,
            quantity_unit TEXT,
            unit_price REAL,
            amount REAL,
            source_sheet TEXT,
            source_row INTEGER,
            confidence REAL NOT NULL,
            review_status TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS review_issues (
            id INTEGER PRIMARY KEY,
            source_document_id INTEGER,
            entity_type TEXT,
            entity_id INTEGER,
            field_name TEXT,
            raw_value TEXT,
            issue_type TEXT NOT NULL,
            severity TEXT NOT NULL,
            resolved INTEGER NOT NULL DEFAULT 0,
            notes TEXT
        );
        """
    )


def nonempty_rows(worksheet: Any) -> list[tuple[int, list[Any]]]:
    rows: list[tuple[int, list[Any]]] = []
    for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [value for value in row if value not in (None, "")]
        if values:
            rows.append((index, values))
    return rows


def value_after_label(rows: list[tuple[int, list[Any]]], *labels: str) -> str:
    wanted = tuple(label.lower() for label in labels)
    for _, values in rows:
        texts = [normalized_text(value) for value in values]
        for position, text in enumerate(texts):
            label_text = text.rstrip(":").strip().lower()
            if any(label in label_text for label in wanted):
                for later in texts[position + 1 :]:
                    if later and later != ":":
                        later_label = later.rstrip(":").strip().lower()
                        if later_label in FIELD_LABELS:
                            continue
                        return later
    return ""


def date_after_label(rows: list[tuple[int, list[Any]]], *labels: str) -> str:
    wanted = tuple(label.lower() for label in labels)
    for _, values in rows:
        texts = [normalized_text(value) for value in values]
        for position, value in enumerate(values):
            label_text = texts[position].rstrip(":").strip().lower()
            if not any(label in label_text for label in wanted):
                continue
            for later in values[position + 1 :]:
                if isinstance(later, datetime):
                    return later.date().isoformat()
                text = normalized_text(later)
                if text and text != ":":
                    return text
    return ""


def numeric_value(value: Any) -> float | None:
    if isinstance(value, int | float):
        return float(value)
    text = normalized_text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def split_quantity(value: Any) -> tuple[str, float | None, str]:
    raw = normalized_text(value)
    if not raw:
        return "", None, ""
    match = re.match(r"(?P<number>-?\d+(?:\.\d+)?)\s*(?P<unit>.*)", raw)
    if not match:
        return raw, None, ""
    return raw, float(match.group("number")), match.group("unit").strip()


def parse_contact(raw_text: str) -> tuple[str, str]:
    text = normalized_text(raw_text)
    match = PHONE_RE.search(text)
    phone = normalized_text(match.group(0)) if match else ""
    name = normalized_text(text.replace(phone, "")) if phone else text
    return name, phone


def get_or_create_supplier(connection: sqlite3.Connection, raw_name: str) -> int | None:
    supplier = normalized_supplier_key(raw_name)
    if not supplier:
        return None
    now = datetime.now(timezone.utc).isoformat()
    connection.execute(
        "INSERT OR IGNORE INTO suppliers (canonical_name, raw_name, created_at) VALUES (?, ?, ?)",
        (supplier, raw_name, now),
    )
    return int(connection.execute("SELECT id FROM suppliers WHERE canonical_name = ?", (supplier,)).fetchone()[0])


def get_or_create_contact(connection: sqlite3.Connection, raw_text: str) -> int | None:
    text = normalized_text(raw_text)
    if not text:
        return None
    name, phone = parse_contact(text)
    now = datetime.now(timezone.utc).isoformat()
    connection.execute(
        "INSERT OR IGNORE INTO contacts (name, phone, raw_text, created_at) VALUES (?, ?, ?, ?)",
        (name, phone, text, now),
    )
    return int(connection.execute("SELECT id FROM contacts WHERE raw_text = ?", (text,)).fetchone()[0])


def get_or_create_item(connection: sqlite3.Connection, raw_description: str) -> int:
    canonical = normalized_item_key(raw_description)
    now = datetime.now(timezone.utc).isoformat()
    connection.execute(
        "INSERT OR IGNORE INTO items (canonical_name, created_at) VALUES (?, ?)",
        (canonical, now),
    )
    item_id = int(connection.execute("SELECT id FROM items WHERE canonical_name = ?", (canonical,)).fetchone()[0])
    connection.execute(
        "INSERT OR IGNORE INTO item_aliases (item_id, alias_text, confidence, approved) VALUES (?, ?, ?, ?)",
        (item_id, raw_description, 0.8, 0),
    )
    connection.execute(
        "INSERT OR IGNORE INTO item_aliases (item_id, alias_text, confidence, approved) VALUES (?, ?, ?, ?)",
        (item_id, canonical, 0.95, 1),
    )
    return item_id


def choose_po_sheet(workbook: Any, po_number: str) -> Any:
    normalized_number = re.sub(r"\D", "", po_number)
    for sheet_name in workbook.sheetnames:
        if normalized_number and normalized_number in normalize_sheet_name(sheet_name):
            return workbook[sheet_name]
    for sheet_name in workbook.sheetnames:
        normalized = normalize_sheet_name(sheet_name)
        if "purchaseorder" in normalized or normalized in {"po", "poelectrical"} or normalized.startswith("po"):
            return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def choose_mr_sheet(workbook: Any) -> Any | None:
    for sheet_name in workbook.sheetnames:
        if "materialrequisition" in normalize_sheet_name(sheet_name):
            return workbook[sheet_name]
    return None


def extract_line_items(rows: list[tuple[int, list[Any]]]) -> list[dict[str, Any]]:
    header_index = None
    for position, (_, values) in enumerate(rows):
        header = " ".join(normalized_text(value).lower() for value in values)
        if "description" in header and ("qty" in header or "quantity" in header):
            header_index = position
            break
    if header_index is None:
        return []

    items: list[dict[str, Any]] = []
    for source_row, values in rows[header_index + 1 :]:
        if not values:
            continue
        first = values[0]
        if not isinstance(first, int | float):
            text = " ".join(normalized_text(value).lower() for value in values)
            if "total" in text:
                break
            continue
        if len(values) < 4:
            continue
        description = normalized_text(values[1])
        if not description:
            continue
        quantity_raw, quantity_value, quantity_unit = split_quantity(values[2] if len(values) > 2 else "")
        unit_price = numeric_value(values[3] if len(values) > 3 else None)
        amount = numeric_value(values[4] if len(values) > 4 else None)
        items.append(
            {
                "row_number": int(first),
                "source_row": source_row,
                "description": description,
                "quantity_raw": quantity_raw,
                "quantity_value": quantity_value,
                "quantity_unit": quantity_unit,
                "unit_price": unit_price,
                "amount": amount,
            }
        )
    return items


def total_from_rows(rows: list[tuple[int, list[Any]]]) -> float | None:
    for _, values in rows:
        text = " ".join(normalized_text(value).lower() for value in values)
        if "total" not in text:
            continue
        numbers = [numeric_value(value) for value in values]
        numbers = [value for value in numbers if value is not None]
        if numbers:
            return numbers[-1]
    return None


def insert_review_issue(
    connection: sqlite3.Connection,
    source_document_id: int,
    field_name: str,
    issue_type: str,
    severity: str,
    notes: str,
) -> None:
    connection.execute(
        """
        INSERT INTO review_issues
            (source_document_id, entity_type, field_name, raw_value, issue_type, severity, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (source_document_id, "source_document", field_name, "", issue_type, severity, notes),
    )


def index_source(connection: sqlite3.Connection, batch_id: int, source: SourceFile, root: Path) -> dict[str, Any]:
    metadata = parse_filename(source.path)
    stat = source.path.stat()
    now = datetime.now(timezone.utc).isoformat()
    pdf_pair = source.path.with_suffix(".pdf")
    year_folder = source.path.parent.name
    source_document_id = int(
        connection.execute(
            """
            INSERT INTO source_documents
                (batch_id, source_path, file_name, file_stem, file_ext, file_size, modified_at,
                 year_folder, month_code, category, supplier_name_from_filename,
                 po_number_from_filename, has_pdf_pair, template_variant, parse_status,
                 review_status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                str(source.path),
                source.path.name,
                source.path.stem,
                source.path.suffix.lower(),
                stat.st_size,
                datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
                year_folder,
                metadata["month_code"],
                metadata["category"],
                metadata["supplier"],
                metadata["po_number"],
                int(pdf_pair.exists()),
                source.template_variant,
                "parsed",
                "needs_review",
                now,
            ),
        ).lastrowid
    )

    if not pdf_pair.exists():
        insert_review_issue(connection, source_document_id, "pdf_pair", "missing_pdf_pair", "medium", "No exact PDF pair found.")

    workbook = load_workbook(source.path, read_only=True, data_only=True)
    po_sheet = choose_po_sheet(workbook, metadata["po_number"])
    po_rows = nonempty_rows(po_sheet)
    mr_sheet = choose_mr_sheet(workbook)
    mr_rows = nonempty_rows(mr_sheet) if mr_sheet else []

    supplier_id = get_or_create_supplier(connection, metadata["supplier"])
    contact_text = value_after_label(po_rows, "person to contact", "contact person")
    contact_id = get_or_create_contact(connection, contact_text)
    po_reference = value_after_label(po_rows, "pr number")
    purchase_order_id = int(
        connection.execute(
            """
            INSERT INTO purchase_orders
                (source_document_id, po_reference, po_number, po_category, po_date, invoice_no,
                 delivery_order_no, supplier_id, project_site, person_to_contact_id, total_amount,
                 source_sheet, confidence, review_status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source_document_id,
                po_reference,
                metadata["po_number"],
                metadata["category"],
                date_after_label(po_rows, "date"),
                value_after_label(po_rows, "invoice no", "invoice"),
                value_after_label(po_rows, "delivery order no"),
                supplier_id,
                value_after_label(po_rows, "project site"),
                contact_id,
                total_from_rows(po_rows),
                po_sheet.title,
                0.75,
                "needs_review",
            ),
        ).lastrowid
    )

    if not po_reference:
        insert_review_issue(connection, source_document_id, "po_reference", "missing_value", "low", "Could not find PR Number.")

    mr_id = None
    if mr_sheet:
        requested_by = value_after_label(mr_rows, "requested by", "request by")
        requested_by_id = get_or_create_contact(connection, requested_by)
        mr_id = int(
            connection.execute(
                """
                INSERT INTO material_requisitions
                    (source_document_id, purchase_order_id, mr_reference, project_code, project_name,
                     date_request, requested_by_contact_id, source_sheet, confidence, review_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    source_document_id,
                    purchase_order_id,
                    value_after_label(mr_rows, "reference no"),
                    value_after_label(mr_rows, "project code"),
                    value_after_label(mr_rows, "project name"),
                    date_after_label(mr_rows, "date request"),
                    requested_by_id,
                    mr_sheet.title,
                    0.7,
                    "needs_review",
                ),
            ).lastrowid
        )
    else:
        insert_review_issue(connection, source_document_id, "material_requisition", "missing_sheet", "medium", "No MR sheet found.")

    line_items = extract_line_items(po_rows)
    if not line_items and mr_rows:
        line_items = extract_line_items(mr_rows)
    if not line_items:
        insert_review_issue(connection, source_document_id, "line_items", "missing_items", "high", "No item rows were extracted.")

    for item in line_items:
        item_id = get_or_create_item(connection, item["description"])
        connection.execute(
            """
            INSERT INTO line_items
                (purchase_order_id, material_requisition_id, row_number, raw_description,
                 normalized_item_id, quantity_raw, quantity_value, quantity_unit, unit_price,
                 amount, source_sheet, source_row, confidence, review_status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                purchase_order_id,
                mr_id,
                item["row_number"],
                item["description"],
                item_id,
                item["quantity_raw"],
                item["quantity_value"],
                item["quantity_unit"],
                item["unit_price"],
                item["amount"],
                po_sheet.title,
                item["source_row"],
                0.75,
                "needs_review",
            ),
        )

    workbook.close()
    return {"file": str(source.path), "items": len(line_items), "po_sheet": po_sheet.title, "mr_sheet": mr_sheet.title if mr_sheet else None}


def summarize_database(connection: sqlite3.Connection) -> dict[str, Any]:
    def scalar(query: str) -> int:
        return int(connection.execute(query).fetchone()[0])

    top_items = [
        {"description": row[0], "quantity": row[1], "spend": row[2]}
        for row in connection.execute(
            """
            SELECT raw_description, SUM(COALESCE(quantity_value, 0)), SUM(COALESCE(amount, 0))
            FROM line_items
            GROUP BY raw_description
            ORDER BY SUM(COALESCE(quantity_value, 0)) DESC
            LIMIT 10
            """
        )
    ]
    expensive_items = [
        {"description": row[0], "unit_price": row[1], "source": row[2]}
        for row in connection.execute(
            """
            SELECT li.raw_description, li.unit_price, sd.file_name
            FROM line_items li
            JOIN purchase_orders po ON po.id = li.purchase_order_id
            JOIN source_documents sd ON sd.id = po.source_document_id
            WHERE li.unit_price IS NOT NULL
            ORDER BY li.unit_price DESC
            LIMIT 10
            """
        )
    ]
    return {
        "source_documents": scalar("SELECT COUNT(*) FROM source_documents"),
        "purchase_orders": scalar("SELECT COUNT(*) FROM purchase_orders"),
        "material_requisitions": scalar("SELECT COUNT(*) FROM material_requisitions"),
        "line_items": scalar("SELECT COUNT(*) FROM line_items"),
        "suppliers": scalar("SELECT COUNT(*) FROM suppliers"),
        "contacts": scalar("SELECT COUNT(*) FROM contacts"),
        "review_issues": scalar("SELECT COUNT(*) FROM review_issues"),
        "top_items_by_quantity": top_items,
        "most_expensive_items": expensive_items,
    }


def run_import(source_root: Path, db_path: Path, report_path: Path, reset: bool, import_all: bool) -> dict[str, Any]:
    if not source_root.exists():
        raise FileNotFoundError(f"Source root not found: {source_root}")
    if reset and db_path.exists():
        db_path.unlink()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)

    sources = workbook_sources(source_root) if import_all else choose_sample_sources(source_root)
    connection = sqlite3.connect(db_path)
    try:
        create_schema(connection)
        started_at = datetime.now(timezone.utc).isoformat()
        batch_id = int(
            connection.execute(
                "INSERT INTO import_batches (batch_name, source_root, started_at, status, notes) VALUES (?, ?, ?, ?, ?)",
                ("full_import" if import_all else "phase2_prototype", str(source_root), started_at, "running", f"File count: {len(sources)}"),
            ).lastrowid
        )
        processed = [index_source(connection, batch_id, source, source_root) for source in sources]
        finished_at = datetime.now(timezone.utc).isoformat()
        connection.execute(
            "UPDATE import_batches SET finished_at = ?, status = ? WHERE id = ?",
            (finished_at, "complete", batch_id),
        )
        summary = summarize_database(connection)
        report = {
            "db_path": str(db_path),
            "source_root": str(source_root),
            "sample_files": processed,
            "summary": summary,
        }
        report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
        connection.commit()
        return report
    finally:
        connection.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prototype importer for manual P.O/MR Excel files.")
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_PATH)
    parser.add_argument("--reset", action="store_true")
    parser.add_argument("--all", action="store_true", help="Import all non-temp Excel files instead of the prototype sample.")
    parser.add_argument("--full-defaults", action="store_true", help="Use the standard full database and report paths.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.full_defaults:
        args.db = FULL_DB_PATH
        args.report = FULL_REPORT_PATH
        args.all = True
    report = run_import(args.source_root, args.db, args.report, args.reset, args.all)
    print(json.dumps(report["summary"], indent=2))
    print(f"Database: {report['db_path']}")
    print(f"Report: {args.report}")


if __name__ == "__main__":
    main()

import os
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from pathlib import Path

from PIL import Image

import pending_store
import procurement
import retention
import suppliers
from openpyxl import load_workbook
from invoice_bot import (
    DOCUMENT_TYPE_INVOICE,
    MR_SHEET_NAME,
    TEMPLATE_SHEET_NAME,
    UNKNOWN_DOCUMENT_FORMAT_MESSAGE,
    apply_manual_po_running_number,
    classify_document,
    compare_and_merge_documents,
    ensure_po_output_stem,
    file_sha256,
    extraction_review_warnings,
    is_openai_auth_error,
    is_openai_credit_error,
    load_cached_extraction,
    manual_po_running_number_is_available,
    normalize_ai_date,
    normalize_ai_document_number,
    validate_ai_extraction,
    parse_po_running_number,
    parse_ocr_contact_person,
    parse_ocr_document_total,
    parse_ocr_line_items,
    parse_ocr_tax_invoice,
    save_cached_extraction,
    save_material_requisition_workbook,
    save_template_workbook,
    text_looks_like_tuju_invoice,
)


class ProcurementTests(unittest.TestCase):
    def test_procurement_bundle_uses_supplier_invoice_and_po_folders(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            po_pdf = root / "po.pdf"
            mr_pdf = root / "mr.pdf"
            po_pdf.write_bytes(b"%PDF-1.4\n")
            mr_pdf.write_bytes(b"%PDF-1.4\n")

            invoice_image = root / "invoice.jpg"
            do_image = root / "do.jpg"
            Image.new("RGB", (10, 10), "white").save(invoice_image)
            Image.new("RGB", (10, 10), "white").save(do_image)

            folder, created, issues = procurement.create_procurement_bundle(
                root / "PROCUREMENT",
                "TUJU GALAXY",
                "TG-K08849",
                "TEST BFE PO TUJU JUNE 0004",
                po_pdf,
                mr_pdf,
                invoice_image,
                do_image,
            )

            self.assertFalse(issues)
            self.assertEqual(folder, root / "PROCUREMENT" / "TUJU GALAXY" / "TG-K08849" / "TEST BFE PO TUJU JUNE 0004")
            self.assertEqual([path.name for path in created], ["1. MR.pdf", "2. PO.pdf", "3. Invoice.pdf", "4. D.O.pdf"])
            for path in created:
                self.assertTrue(path.exists())


class PendingStoreTests(unittest.TestCase):
    def test_pending_save_load_clear_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            pending = {"mode": "pair", "invoice_id": "I1", "documents": {"invoice": {"tax_invoice": "TG-K08849"}}}
            pending_store.save_pending(root, "5037627395", pending)

            self.assertEqual(pending_store.load_pending(root, "5037627395"), pending)
            pending_store.clear_pending(root, "5037627395")
            self.assertIsNone(pending_store.load_pending(root, "5037627395"))


class ExtractionCacheTests(unittest.TestCase):
    def test_extraction_cache_round_trip_uses_source_hash(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            image_path = root / "invoice.jpg"
            image_path.write_bytes(b"same image bytes")
            source_hash = file_sha256(image_path)

            old_cache_dir = os.environ.get("EXTRACTION_CACHE_DIR")
            # The bot cache dir is a module constant, so use a unique hash and remove it after.
            save_cached_extraction(source_hash, {"tax_invoice": "TG-K08941", "line_items": []})
            try:
                cached = load_cached_extraction(source_hash)
            finally:
                Path("data/extraction_cache", f"{source_hash}.json").unlink(missing_ok=True)
                if old_cache_dir is None:
                    os.environ.pop("EXTRACTION_CACHE_DIR", None)

            self.assertIsNotNone(cached)
            self.assertEqual(cached["tax_invoice"], "TG-K08941")
            self.assertEqual(cached["source_image_hash"], source_hash)


class SupplierTests(unittest.TestCase):
    def test_detect_supplier_from_invoice_prefix(self) -> None:
        supplier = suppliers.detect_supplier_name(
            {"tax_invoice": "TG-K08849"},
            "UNKNOWN SUPPLIER",
            {"TUJU GALAXY": ["tuju galaxy", "tg-"]},
        )
        self.assertEqual(supplier, "TUJU GALAXY")

    def test_explicit_supplier_wins(self) -> None:
        self.assertEqual(
            suppliers.detect_supplier_name({"supplier_name": "ABC SDN BHD", "tax_invoice": "TG-K08849"}, "TUJU GALAXY"),
            "ABC SDN BHD",
        )


class PoRunningNumberTests(unittest.TestCase):
    def test_parse_po_running_number_accepts_plain_and_padded_numbers(self) -> None:
        self.assertEqual(parse_po_running_number("7"), 7)
        self.assertEqual(parse_po_running_number("0007"), 7)
        self.assertIsNone(parse_po_running_number("0"))
        self.assertIsNone(parse_po_running_number("7 extra"))

    def test_apply_manual_po_running_number_sets_month_type_and_stem(self) -> None:
        data = {"record_type": "record"}

        stem = apply_manual_po_running_number(
            data,
            7,
            datetime(2026, 7, 7, tzinfo=timezone.utc),
            "record",
        )

        self.assertEqual(stem, "BFE PO TUJU JULY 0007")
        self.assertEqual(data["po_month_key"], "2026-07")
        self.assertEqual(data["po_month_name"], "JULY")
        self.assertEqual(data["po_running_number"], "0007")
        self.assertEqual(data["po_output_stem"], "BFE PO TUJU JULY 0007")

    def test_manual_po_running_number_rejects_used_register_number(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            register = root / "invoice_register.csv"
            register.write_text(
                "po_month_key,record_type,po_running_number,workbook_file\n"
                "2026-07,record,0007,\n",
                encoding="utf-8",
            )
            old_register = os.environ.get("INVOICE_REGISTER_PATH")
            old_workbook_dir = os.environ.get("INVOICE_WORKBOOK_DIR")
            os.environ["INVOICE_REGISTER_PATH"] = str(register)
            os.environ["INVOICE_WORKBOOK_DIR"] = str(root / "invoices")
            try:
                self.assertFalse(
                    manual_po_running_number_is_available(
                        7,
                        datetime(2026, 7, 7, tzinfo=timezone.utc),
                        "record",
                    )
                )
                self.assertTrue(
                    manual_po_running_number_is_available(
                        8,
                        datetime(2026, 7, 7, tzinfo=timezone.utc),
                        "record",
                    )
                )
            finally:
                if old_register is None:
                    os.environ.pop("INVOICE_REGISTER_PATH", None)
                else:
                    os.environ["INVOICE_REGISTER_PATH"] = old_register
                if old_workbook_dir is None:
                    os.environ.pop("INVOICE_WORKBOOK_DIR", None)
                else:
                    os.environ["INVOICE_WORKBOOK_DIR"] = old_workbook_dir


class RetentionTests(unittest.TestCase):
    def test_archive_old_files_moves_only_old_target_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            target = root / "ocr"
            target.mkdir()
            old_file = target / "old.txt"
            new_file = target / "new.txt"
            old_file.write_text("old", encoding="utf-8")
            new_file.write_text("new", encoding="utf-8")

            old_time = (datetime.now(timezone.utc) - timedelta(days=40)).timestamp()
            os.utime(old_file, (old_time, old_time))

            result = retention.archive_old_files([target], root / "cleanup_archive", older_than_days=30)

            self.assertEqual(result.scanned, 1)
            self.assertEqual(result.archived, 1)
            self.assertFalse(old_file.exists())
            self.assertTrue(new_file.exists())
            self.assertTrue((result.archive_dir / "ocr" / "old.txt").exists())


class PairMergeTests(unittest.TestCase):
    def test_merge_uses_do_contact_and_invoice_prices(self) -> None:
        delivery_order = {
            "document_type": "delivery_order",
            "tax_invoice": "TG-K08849",
            "invoice_date": "2026-06-29",
            "contact_person": "Farah 011-54302725",
            "line_items": [{"item_no": "1", "description": "Widget", "quantity": 2, "quantity_unit": "pcs"}],
        }
        invoice = {
            "document_type": "invoice",
            "tax_invoice": "TG-K08849",
            "invoice_date": "2026-06-30",
            "contact_person": "Invoice Contact",
            "line_items": [{"item_no": "1", "description": "Widget", "quantity": 2, "unit_price": 5, "line_total": 10}],
        }

        merged, warnings = compare_and_merge_documents(delivery_order, invoice)

        self.assertFalse(warnings)
        self.assertEqual(merged["contact_person"], "Farah 011-54302725")
        self.assertEqual(merged["delivery_order_contact_person"], "Farah 011-54302725")
        self.assertEqual(merged["invoice_document_date"], "2026-06-30")
        self.assertEqual(merged["invoice_date"], "2026-06-30")
        self.assertEqual(merged["delivery_order_date"], "2026-06-29")
        self.assertEqual(merged["supplier_name"], "TUJU GALAXY")
        self.assertEqual(merged["line_items"][0]["unit_price"], 5)
        self.assertEqual(merged["line_items"][0]["line_total"], 10)

    def test_merge_can_use_invoice_items_when_pair_rows_differ(self) -> None:
        delivery_order = {
            "document_type": "delivery_order",
            "tax_invoice": "TG-K08935",
            "invoice_date": "2026-07-01",
            "contact_person": "Zukiram 019-9848812",
            "line_items": [{"item_no": "2", "description": 'Kayu 1" x 2" x 8', "quantity": 1, "quantity_unit": "ton"}],
        }
        invoice = {
            "document_type": "invoice",
            "tax_invoice": "TG-K08935",
            "invoice_date": "2026-07-01",
            "line_items": [
                {"item_no": "1", "description": "WBP Plywood", "quantity": 62, "quantity_unit": "pcs", "unit_price": 30, "line_total": 1860},
                {"item_no": "2", "description": 'Kayu 1" x 2" x 8', "quantity": 1, "quantity_unit": "ton", "unit_price": 1750, "line_total": 1750},
            ],
        }

        merged, warnings = compare_and_merge_documents(delivery_order, invoice, DOCUMENT_TYPE_INVOICE)

        self.assertTrue(any("extra item row" in warning for warning in warnings))
        self.assertEqual(merged["item_source"], DOCUMENT_TYPE_INVOICE)
        self.assertEqual(len(merged["line_items"]), 2)


class DateAndStemTests(unittest.TestCase):
    def test_po_month_name_and_key_from_iso_string(self) -> None:
        from invoice_bot import po_month_key, po_month_name

        self.assertEqual(po_month_name("2026-06-20"), "JUNE")
        self.assertEqual(po_month_key("2026-06-20"), "2026-06")

    def test_ensure_po_output_stem_uses_invoice_date(self) -> None:
        from invoice_bot import ensure_po_output_stem

        data = {
            "invoice_date": "2026-06-20",
            "tax_invoice": "TG-K08849",
        }
        # Even if received_at is in August, the stem should use June from invoice_date
        received_august = datetime(2026, 8, 28, 12, 0, 0, tzinfo=timezone.utc)
        stem = ensure_po_output_stem(data, received_at=received_august)

        self.assertIn("JUNE", stem)
        self.assertEqual(data.get("po_month_name"), "JUNE")
        self.assertEqual(data.get("po_month_key"), "2026-06")

    def test_save_material_requisition_workbook_openpyxl(self) -> None:
        from openpyxl import Workbook, load_workbook
        from invoice_bot import MR_SHEET_NAME, save_material_requisition_workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            template_path = root / "mr_template.xlsx"
            target_path = root / "test_mr.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = MR_SHEET_NAME
            wb.save(template_path)
            wb.close()

            data = {
                "invoice_date": "2026-06-20",
                "delivery_order_contact_person": "Farah 011-54302725",
                "line_items": [
                    {"item_no": "1", "description": "Steel Band", "quantity": "4 roll", "unit_price": 13.8, "line_total": 55.2},
                ],
            }

            count = save_material_requisition_workbook(target_path, template_path, data, "BFE PO TUJU JUNE 0001")

            self.assertEqual(count, 1)
            self.assertTrue(target_path.exists())

            saved_wb = load_workbook(target_path)
            saved_ws = saved_wb[MR_SHEET_NAME]
            self.assertEqual(saved_ws["N10"].value, "BFE PO TUJU JUNE 0001")
            self.assertEqual(saved_ws["D15"].value, "Farah 011-54302725")
            self.assertEqual(saved_ws["B19"].value, "1")
            self.assertEqual(saved_ws["C19"].value, "Steel Band")
            self.assertEqual(saved_ws["K19"].value, "4 roll")
            self.assertEqual(saved_ws["L19"].value, 13.8)
            self.assertEqual(saved_ws["N19"].value, 55.2)
            saved_wb.close()

    def test_walihin_quotation_and_tech_stem_generation(self) -> None:
        data = {
            "document_type": "quotation",
            "supplier_name": "WALIHIN PETROLEUM SDN. BHD. (432843-D)",
            "supplier_address": "Lot 7,GSL 3104,Hakka Avenue Estate, 5th Miles Penrissen Road,93250 Kuching,Sarawak.",
            "supplier_phone": "082-575987 / 016-8865086",
            "supplier_email": "walihinpetroleum@yahoo.com",
            "tax_invoice": "QUOTATION",
            "invoice_date": "2026-08-10",
            "contact_person": "Feddy Sim 016-8868203",
            "submitter_name": "Mrs. Azyan Nasuha",
            "line_items": [
                {"item_no": 1, "description": "Diesel", "quantity": 1600, "quantity_unit": "Lts", "unit_price": 4.96, "line_total": 7936.0},
                {"item_no": 2, "description": "Transport Charge", "quantity": 1, "quantity_unit": "", "unit_price": 120.0, "line_total": 120.0},
            ],
        }
        stem = ensure_po_output_stem(data)
        self.assertIn("BFE PO TECH 0826", stem)
        self.assertIn("WALIHIN PETROLEUM SDN BHD", stem)
        self.assertEqual(data.get("pr_number"), "BFE/PO/TECH/AN/0826-0001")

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            po_path = temp_path / f"{stem}.xlsx"
            template_path = Path(__file__).resolve().parent.parent / "templates" / "purchase_order_template.xlsx"
            if not template_path.exists():
                template_path = Path(__file__).resolve().parent.parent / "data" / "templates" / "purchase_order_template.xlsx"

            count = save_template_workbook(po_path, template_path, data)
            self.assertEqual(count, 2)
            saved_wb = load_workbook(po_path)
            saved_ws = saved_wb[TEMPLATE_SHEET_NAME]
            self.assertEqual(saved_ws["J15"].value, "BFE/PO/TECH/AN/0826-0001")
            self.assertEqual(saved_ws["B15"].value, "WALIHIN PETROLEUM SDN BHD")
            self.assertEqual(saved_ws["B16"].value, "Lot 7, GSL 3104, Hakka Avenue Estate")
            self.assertEqual(saved_ws["B17"].value, "5th Miles Penrissen Road, 93250 Kuching, Sarawak.")
            self.assertEqual(saved_ws["B18"].value, "TEL/FAX : 082-575987 / Office H/P : 016-8865086")
            self.assertEqual(saved_ws["B19"].value, "Email : walihinpetroleum@yahoo.com")
            self.assertEqual(saved_ws["B20"].value, "No Acc : 561118064592 (Maybank)")
            self.assertEqual(saved_ws["C26"].value, "Diesel")
            self.assertEqual(saved_ws["G26"].value, "1600 Lts")
            self.assertEqual(saved_ws["H26"].value, 4.96)
            self.assertEqual(saved_ws["J26"].value, 7936.0)
            saved_wb.close()


class LocalOCRParsingTests(unittest.TestCase):
    def test_document_profile_matches_tuju_delivery_order(self) -> None:
        profile_id, status, score, runner_up = classify_document(
            "TUJU GALAKSI SDN BHD Delivery Order No. TG-K08849"
        )
        self.assertEqual(profile_id, "tuju_galaxy_delivery_order")
        self.assertEqual(status, "matched")
        self.assertGreaterEqual(score, 6)

    def test_document_profile_matches_tuju_invoice(self) -> None:
        profile_id, status, score, runner_up = classify_document(
            "Blackfox Engineering Sdn Bhd TAXINVOICE :TG-K08849"
        )
        self.assertEqual(profile_id, "tuju_galaxy_invoice")
        self.assertEqual(status, "matched")
        self.assertGreaterEqual(score, 6)

    def test_document_profile_rejects_unknown_format(self) -> None:
        profile_id, status, score, runner_up = classify_document(
            "Random receipt without known TUJU markers"
        )
        self.assertEqual(status, "below_threshold")
        self.assertIn("New document format detected", UNKNOWN_DOCUMENT_FORMAT_MESSAGE)

    def test_tuju_detection_tolerates_ocr_zero_letter_confusion(self) -> None:
        self.assertTrue(text_looks_like_tuju_invoice("Blackfox TAKIWVOKE 1TG-KO8935 OurD/ONo TG-KO8935"))

    def test_ai_date_uses_visible_malaysian_date_from_notes(self) -> None:
        normalized, warning = normalize_ai_date("2026-01-07", "Date shown is 01.07.2026")

        self.assertEqual(normalized, "2026-07-01")
        self.assertIsNone(warning)

    def test_ai_document_number_rejects_invalid_or_phone_values(self) -> None:
        self.assertEqual(normalize_ai_document_number("TG-KO8941")[0], "TG-K08941")
        self.assertIsNone(normalize_ai_document_number("TIC-X08941")[0])
        self.assertIsNone(normalize_ai_document_number("016-8873726")[0])

    def test_validate_ai_extraction_adds_warnings_for_bad_fields(self) -> None:
        data = {
            "tax_invoice": "TIC-X08941",
            "invoice_date": "not visible",
            "notes": "unclear header",
            "line_items": [],
        }

        validate_ai_extraction(data)

        self.assertIsNone(data["tax_invoice"])
        self.assertIsNone(data["invoice_date"])
        self.assertTrue(data["validation_warnings"])

    def test_delivery_order_number_normalizes_ocr_o_zero_confusion(self) -> None:
        text = "Delivery OrderNo. :TG-KO8849.—"

        self.assertEqual(parse_ocr_tax_invoice(text), "TG-K08849")

    def test_delivery_order_number_normalizes_ocr_prefix_confusion(self) -> None:
        text = "Delivery Order No. :1G-KO8935"

        self.assertEqual(parse_ocr_tax_invoice(text), "TG-K08935")

    def test_delivery_order_number_does_not_cross_into_billing_address(self) -> None:
        text = "Delivery Order\nBilling Address:\nBlackfox Engineering Sdn Bhd\nTG-K08935"

        self.assertEqual(parse_ocr_tax_invoice(text), "TG-K08935")

    def test_delivery_order_rows_handle_quotes_and_units(self) -> None:
        text = "\n".join(
            [
                "No. Product Description Quantity",
                "1 3/4\"(17mm) x 10m stee! band 4 ‘roll",
                "2 PVC plug 6mm x 40mm 10\\card",
                "3 Sleeve anchor 3/8\" x 3', 100pcs/box 1 box",
            ]
        )

        items = parse_ocr_line_items(text)

        self.assertEqual(len(items), 3)
        self.assertEqual(items[0]["description"], '3/4"(17mm) x 10m steel band')
        self.assertEqual(items[0]["quantity"], 4)
        self.assertEqual(items[0]["quantity_unit"], "roll")

    def test_delivery_order_rows_handle_noisy_quantity_punctuation(self) -> None:
        text = "\n".join(
            [
                "1 WBP Plywood 12mmx 4' x 8! (Black glue), 85pcs/bdl 62. PCs",
                "2 Kayu 1\" x2\" x8! 1: ton:",
            ]
        )

        items = parse_ocr_line_items(text)

        self.assertEqual(len(items), 2)
        self.assertEqual(items[0]["quantity_unit"], "pcs")
        self.assertEqual(items[1]["quantity"], 1)
        self.assertEqual(items[1]["quantity_unit"], "ton")

    def test_contact_cleaning_removes_fax_noise(self) -> None:
        text = "Contact Person:\nI: Fax: Farah 011-54302725"

        self.assertEqual(parse_ocr_contact_person(text), "Farah 011-54302725")

    def test_invoice_rows_handle_noisy_prices(self) -> None:
        text = "\n".join(
            [
                "No. Product Description Quantity Unit Price Amount",
                '1 =3/4"(17mm) x 10m steel band 4 roll 13.80. 55.20;',
                "2 PVCplugé6mmx40mm 10 card 1:60 16:00",
                "3. Sleeve anchor 3/8\" x 3', 100pcs/box 1 box 97.50 97.50)",
            ]
        )

        items = parse_ocr_line_items(text)

        self.assertEqual(len(items), 3)
        self.assertEqual(items[1]["description"], "PVC plug 6mm x 40mm")
        self.assertEqual(items[1]["unit_price"], 1.6)
        self.assertEqual(items[1]["line_total"], 16)

    def test_invoice_total_parses_noisy_total_payable(self) -> None:
        text = "Nets _Total Payable :| RM___168.70| ~"

        self.assertEqual(parse_ocr_document_total(text), 168.7)

    def test_delivery_order_review_does_not_require_prices(self) -> None:
        warnings = extraction_review_warnings(
            {
                "document_type": "delivery_order",
                "tax_invoice": "TG-K08849",
                "invoice_date": "2026-06-20",
                "confidence": 0.9,
                "line_items": [
                    {"item_no": "1", "description": "Widget", "quantity": 2, "quantity_unit": "pcs"},
                ],
            }
        )

        self.assertFalse(any("unit price" in warning or "amount" in warning for warning in warnings))

    def test_invoice_review_warns_on_total_mismatch(self) -> None:
        warnings = extraction_review_warnings(
            {
                "document_type": "invoice",
                "tax_invoice": "TG-K08849",
                "invoice_date": "2026-06-20",
                "confidence": 0.6,
                "document_total": 20,
                "line_items": [
                    {"item_no": "1", "description": "Widget", "quantity": 2, "unit_price": 5, "line_total": 10},
                    {"item_no": "2", "description": "Widget 2", "quantity": 1, "unit_price": 5, "line_total": 5},
                ],
            }
        )

        self.assertTrue(any("Low OCR confidence" in warning for warning in warnings))
        self.assertTrue(any("Invoice total check differs" in warning for warning in warnings))


class OpenAIErrorClassificationTests(unittest.TestCase):
    @staticmethod
    def _api_status_error(status_code: int, message: str = "error"):
        import httpx
        from openai import APIStatusError

        request = httpx.Request("POST", "https://openrouter.ai/api/v1/chat/completions")
        response = httpx.Response(status_code, request=request, json={"error": {"message": message}})
        return APIStatusError(message, response=response, body=None)

    def test_credit_error_detects_402(self) -> None:
        exc = self._api_status_error(402, "insufficient credits")
        self.assertTrue(is_openai_credit_error(exc))
        self.assertFalse(is_openai_auth_error(exc))

    def test_credit_error_ignores_other_status_codes(self) -> None:
        exc = self._api_status_error(429, "rate limited")
        self.assertFalse(is_openai_credit_error(exc))
        self.assertFalse(is_openai_auth_error(exc))

class AIProviderTests(unittest.TestCase):
    def test_configured_ai_provider_detects_gemini(self) -> None:
        from invoice_bot import ai_fallback_enabled, configured_ai_provider, gemini_api_key, gemini_model_name

        old_gemini = os.environ.get("GEMINI_API_KEY")
        old_provider = os.environ.get("AI_PROVIDER")
        old_fallback = os.environ.get("AI_FALLBACK_ENABLED")
        try:
            os.environ["GEMINI_API_KEY"] = "AIzaSyFakeKey123"
            os.environ.pop("AI_PROVIDER", None)
            os.environ.pop("AI_FALLBACK_ENABLED", None)
            self.assertEqual(gemini_api_key(), "AIzaSyFakeKey123")
            self.assertEqual(configured_ai_provider(), "gemini")
            self.assertEqual(gemini_model_name(), "gemini-3.7-flash")
            self.assertTrue(ai_fallback_enabled())

            os.environ["AI_PROVIDER"] = "openai"
            self.assertEqual(configured_ai_provider(), "openai")

            os.environ["AI_PROVIDER"] = "gemini"
            self.assertEqual(configured_ai_provider(), "gemini")
        finally:
            if old_gemini is not None:
                os.environ["GEMINI_API_KEY"] = old_gemini
            else:
                os.environ.pop("GEMINI_API_KEY", None)
            if old_provider is not None:
                os.environ["AI_PROVIDER"] = old_provider
            else:
                os.environ.pop("AI_PROVIDER", None)
            if old_fallback is not None:
                os.environ["AI_FALLBACK_ENABLED"] = old_fallback
            else:
                os.environ.pop("AI_FALLBACK_ENABLED", None)


class ServiceOrderTests(unittest.TestCase):
    def test_is_service_order_detection(self) -> None:
        from invoice_bot import is_service_order

        data_so1 = {"order_type": "service_order"}
        self.assertTrue(is_service_order(data_so1))

        data_so2 = {"service_description": "Services one unit Heli forklift at HPJ"}
        self.assertTrue(is_service_order(data_so2))

        data_so3 = {
            "line_items": [
                {"description": "Transmission filter", "quantity": 1, "unit_price": 23.4},
                {"description": "Labour charges for maintenance work", "quantity": 1, "unit_price": 230},
            ]
        }
        self.assertTrue(is_service_order(data_so3))

        data_po = {
            "line_items": [
                {"description": "Wash coarse sand", "quantity": 115.71, "unit_price": 59.0},
            ]
        }
        self.assertFalse(is_service_order(data_po))

    def test_ensure_so_output_stem_and_so_number(self) -> None:
        from invoice_bot import ensure_po_output_stem

        data = {
            "order_type": "service_order",
            "invoice_date": "2026-07-24",
            "supplier_name": "TUJU GALAKSI SDN BHD",
            "submitter_name": "Azyan Nasuha",
            "line_items": [
                {"description": "Services one unit Heli forklift", "quantity": 1, "unit_price": 230},
            ],
        }
        stem = ensure_po_output_stem(data)
        self.assertIn("BFE SO TECH 0726", stem)
        self.assertIn("TUJU GALAKSI SDN BHD", stem)
        self.assertIn("BFE/SO/TUJU/AN/0726/", data.get("so_number", ""))

    def test_save_service_order_workbook_openpyxl(self) -> None:
        import openpyxl
        from invoice_bot import (
            DEFAULT_SERVICE_ORDER_TEMPLATE_PATH,
            SO_SHEET_NAME,
            save_service_order_workbook,
        )

        if not DEFAULT_SERVICE_ORDER_TEMPLATE_PATH.exists():
            self.skipTest("SO template not found")

        data = {
            "tax_invoice": "TG-K09124",
            "invoice_date": "2026-07-24",
            "submitter_name": "Azyan Nasuha",
            "service_description": "Services one unit Heli forklift at HPJ",
            "so_number": "BFE/SO/TUJU/AN/0726/012",
            "supplier_name": "TUJU GALAKSI SDN BHD",
            "order_type": "service_order",
            "line_items": [
                {"item_no": 1, "description": "Transmission filter", "quantity": 1.0, "quantity_unit": "pc", "unit_price": 23.4, "line_total": 23.4},
                {"item_no": 2, "description": "Oil filter", "quantity": 1.0, "quantity_unit": "pc", "unit_price": 23.4, "line_total": 23.4},
                {"item_no": 3, "description": "Fuel filter", "quantity": 1.0, "quantity_unit": "pc", "unit_price": 36.5, "line_total": 36.5, "warranty": "-"},
                {"item_no": 4, "description": "Engine Oil", "quantity": 8.0, "quantity_unit": "ltr", "unit_price": 23.5, "line_total": 188.0},
                {"item_no": 5, "description": "Greasing", "quantity": 1.0, "quantity_unit": "set", "unit_price": 39.0, "line_total": 39.0},
                {"item_no": 6, "description": "Labour charges for maintenance work", "quantity": 1.0, "quantity_unit": "unit", "unit_price": 230.0, "line_total": 230.0},
            ],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            out_file = Path(tmpdir) / "BFE SO TECH 0726 012 TUJU GALAKSI SDN BHD.xlsx"
            saved_count = save_service_order_workbook(out_file, DEFAULT_SERVICE_ORDER_TEMPLATE_PATH, data)
            self.assertEqual(saved_count, 6)
            self.assertTrue(out_file.exists())

            wb = openpyxl.load_workbook(out_file, data_only=False)
            self.assertIn(SO_SHEET_NAME, wb.sheetnames)
            ws = wb[SO_SHEET_NAME]
            self.assertEqual(ws["I15"].value, "BFE/SO/TUJU/AN/0726/012")
            self.assertEqual(ws["I17"].value, "TG-K09124")
            self.assertEqual(ws["I18"].value, "=I17")
            self.assertEqual(ws["B27"].value, "Services one unit Heli forklift at HPJ")
            self.assertEqual(ws["A28"].value, 1)
            self.assertEqual(ws["B28"].value, "Transmission filter")
            self.assertEqual(ws["F28"].value, "1 pcs")
            self.assertEqual(ws["G28"].value, 23.4)
            self.assertEqual(ws["H28"].value, 23.4)
            self.assertEqual(ws["H51"].value, "=SUM(H27:H50)")
            self.assertIn("Azyan Nasuha", str(ws["G62"].value))


if __name__ == "__main__":
    unittest.main()


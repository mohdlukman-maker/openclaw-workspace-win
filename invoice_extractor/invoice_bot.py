import argparse
import asyncio
import base64
import csv
import hashlib
import json
import logging
from logging.handlers import RotatingFileHandler
import os
import re
import shutil
import subprocess
import time
from dataclasses import dataclass
from datetime import date as datetime_date, datetime, timedelta, timezone
import math
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openai import APIStatusError, AsyncOpenAI, AuthenticationError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from PIL import Image, ImageEnhance, ImageFilter, ImageOps, ImageStat
from telegram import BotCommand, InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.error import TelegramError, TimedOut
from telegram.ext import Application, CallbackQueryHandler, CommandHandler, ContextTypes, MessageHandler, filters

import document_profiles
import image_processor
import ocr_enhanced
import pending_store
import profile_management
import registration_flow
import retention
import suppliers
from procurement_query import answer_question, format_table
from procurement import create_procurement_bundle as create_procurement_packet


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
IMAGE_DIR = DATA_DIR / "images"
ENHANCED_DIR = DATA_DIR / "enhanced"
ENHANCED_OUTPUTS_DIR = BASE_DIR / "enhanced_outputs"
EXTRACTION_DIR = DATA_DIR / "extractions"
OCR_DIR = DATA_DIR / "ocr"
EXTRACTION_CACHE_DIR = DATA_DIR / "extraction_cache"
DEFAULT_INVOICE_WORKBOOK_DIR = DATA_DIR / "invoices"
PENDING_REVIEW_KEY = "pending_invoice_review"
EDIT_PO_RUNNING_NUMBER_KEY = "edit_po_running_number"
SAVE_WITH_CUSTOM_NAME_KEY = "save_with_custom_name"
DOCUMENT_TYPE_DELIVERY_ORDER = "delivery_order"
DOCUMENT_TYPE_INVOICE = "invoice"
DOCUMENT_TYPE_QUOTATION = "quotation"
DOCUMENT_TYPE_CASH_BILL = "cash_bill"
AUTO_SAVE_RETRY_SECONDS = 10
AUTO_SAVE_MAX_ATTEMPTS = 90
DEFAULT_LOCAL_OCR_MIN_CONFIDENCE = 55.0
DEFAULT_LOCAL_OCR_REVIEW_CONFIDENCE = 75.0
DEFAULT_LOCAL_OCR_MIN_ITEMS = 1
DEFAULT_IMAGE_QUALITY_MIN_SHORT_SIDE = 900
DEFAULT_IMAGE_QUALITY_MIN_CONTRAST = 25.0
DEFAULT_IMAGE_QUALITY_MIN_SHARPNESS = 80.0
DEFAULT_IMAGE_QUALITY_MIN_BRIGHTNESS = 45.0
DEFAULT_IMAGE_QUALITY_MAX_BRIGHTNESS = 225.0
DEFAULT_INVOICE_TEMPLATE_PATH = DATA_DIR / "templates" / "purchase_order_template.xlsx"
DEFAULT_MATERIAL_REQUISITION_TEMPLATE_PATH = DATA_DIR / "templates" / "material_requisition_template.xlsx"
DEFAULT_INVOICE_REGISTER_PATH = DATA_DIR / "invoice_register.csv"
DEFAULT_PROCUREMENT_DIR = DATA_DIR / "PROCUREMENT"
DEFAULT_PROCUREMENT_SUPPLIER_NAME = "TUJU GALAXY"
DEFAULT_CLEANUP_ARCHIVE_DIR = DATA_DIR / "cleanup_archive"
DEFAULT_CLEANUP_RETENTION_DAYS = 30
PO_FILENAME_PREFIX = "BFE PO TUJU"
TEST_PO_FILENAME_PREFIX = f"TEST {PO_FILENAME_PREFIX}"
DEFAULT_TEST_CHAT_IDS = {"5037627395"}
TUJU_PROFILE_ENABLED = True
TEMPLATE_SHEET_NAME = "PURCHASEORDER"
TEMPLATE_FIRST_ITEM_ROW = 26
TEMPLATE_LAST_ITEM_ROW = 54
MR_SHEET_NAME = "MATERIAL REQUISITION"
MR_FIRST_ITEM_ROW = 19
MR_LAST_ITEM_ROW = 40
DEFAULT_EXPORT_PDF = True
LEGACY_TUJU_PIPELINE_DEFAULT = False


# ── Profile extraction helpers ────────────────────────────────────────
PROFILE_PROFILES_DIR = BASE_DIR / "data" / "document_profiles"


def profile_by_id(profile_id: str) -> document_profiles.DocumentProfile | None:
    """Look up a profile by its id from the loaded profiles."""
    for p in get_profiles():
        if p.id == profile_id:
            return p
    return None


def create_profile_crops(image_path: Path, profile: document_profiles.DocumentProfile) -> dict[str, Path]:
    """Create cropped images from profile field crop_hints.

    Returns a dict of {field_name: cropped_path} for fields that have crop_hints.
    Falls back to an empty dict if no fields have hints.
    """
    ENHANCED_DIR.mkdir(parents=True, exist_ok=True)
    crops: dict[str, Path] = {}
    if not profile.fields:
        return crops
    with Image.open(image_path) as image:
        for field in profile.fields:
            if field.crop_hint is None:
                continue
            x1, y1, x2, y2 = field.crop_hint
            path = ENHANCED_DIR / f"{image_path.stem}_{profile.id}_{field.name}.jpg"
            save_relative_crop(image, path, (x1, y1, x2, y2), scale=2, contrast=1.9)
            crops[field.name] = path
    return crops


def render_profile_prompt(profile: document_profiles.DocumentProfile) -> str:
    """Render the ai_extraction_prompt template with profile data."""
    if not profile.ai_extraction_prompt:
        return ""
    prompt = profile.ai_extraction_prompt
    prompt = prompt.replace("{supplier}", profile.supplier)
    prompt = prompt.replace("{document_type}", profile.document_type)

    field_instructions = "\n".join(
        f"- {f.name}: {'Required' if f.required else 'Optional'} {f.type}"
        + (f" (pattern: {f.pattern})" if f.pattern else "")
        for f in profile.fields
    )
    prompt = prompt.replace("{field_instructions}", field_instructions)

    table_columns = " | ".join(
        f"{c.field}: {c.type}" for c in (profile.line_item_table.columns if profile.line_item_table else [])
    )
    prompt = prompt.replace("{table_columns}", table_columns)

    schema_block = "{\n"
    for f in profile.fields:
        schema_block += f"  \"{f.name}\": string | null,\n"
    schema_block += "  \"line_items\": [\n"
    for c in (profile.line_item_table.columns if profile.line_item_table else []):
        schema_block += f"    \"{c.field}\": {c.type},\n"
    schema_block += "  ]\n}"
    prompt = prompt.replace("{schema_block}", schema_block)

    return prompt


def apply_profile_normalizers(
    data: dict[str, Any],
    profile: document_profiles.DocumentProfile,
) -> dict[str, Any]:
    """Apply per-field normalizers from the profile to extracted data."""
    result = dict(data)
    for field in profile.fields:
        if field.name not in result:
            continue
        if not field.normalizers:
            continue
        result[field.name] = document_profiles.apply_normalizers(
            result[field.name], field.normalizers
        )

    # Also normalize line items
    line_items = result.get("line_items", [])
    if line_items and profile.line_item_table:
        normalized_items = []
        for item in line_items:
            normalized = dict(item)
            for col in profile.line_item_table.columns:
                if col.field not in item:
                    continue
                # Find the matching field definition for normalizers
                for field in profile.fields:
                    if field.name == col.field and field.normalizers:
                        normalized[col.field] = document_profiles.apply_normalizers(
                            item[col.field], field.normalizers
                        )
                        break
            normalized_items.append(normalized)
        result["line_items"] = normalized_items

    return result


def run_validation_rules(
    data: dict[str, Any],
    profile: document_profiles.DocumentProfile,
) -> list[str]:
    """Run validation rules from the profile and return warning messages."""
    warnings: list[str] = []
    line_items = data.get("line_items", [])
    if not isinstance(line_items, list):
        line_items = []

    for rule in profile.validation_rules:
        if rule.rule == "row_arithmetic":
            for idx, item in enumerate(line_items, start=1):
                qty = _safe_float(item.get("quantity"))
                price = _safe_float(item.get("unit_price"))
                total = _safe_float(item.get("line_total"))
                if qty is not None and price is not None and total is not None:
                    calc = qty * price
                    if abs(calc - total) > rule.tolerance:
                        warnings.append(
                            f"Row {idx} arithmetic check: {qty} x {price} = {calc:.2f}, "
                            f"extracted = {total:.2f} (diff {abs(calc - total):.2f})"
                        )

        elif rule.rule == "line_totals_sum_to":
            target = _safe_float(data.get(rule.target or ""))
            total = sum(_safe_float(item.get("line_total")) or 0 for item in line_items)
            if target is not None and total > 0:
                if abs(total - target) > rule.tolerance:
                    warnings.append(
                        f"Line items sum to {total:.2f}, but {rule.target} = {target:.2f} "
                        f"(diff {abs(total - target):.2f})"
                    )

        elif rule.rule == "field_sum":
            op_sum = sum(_safe_float(data.get(op)) or 0 for op in rule.operands)
            target = _safe_float(data.get(rule.target or ""))
            if target is not None and op_sum > 0:
                if abs(op_sum - target) > rule.tolerance:
                    warnings.append(
                        f"{' + '.join(rule.operands)} = {op_sum:.2f}, "
                        f"but {rule.target} = {target:.2f} (diff {abs(op_sum - target):.2f})"
                    )

    return warnings


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def legacy_tuju_pipeline_enabled() -> bool:
    return os.getenv("LEGACY_TUJU_PIPELINE", "0").strip().lower() in {"1", "true", "yes", "on"}


class WorkbookBusyError(PermissionError):
    pass


class LocalOCRUnavailable(RuntimeError):
    pass


class UnknownDocumentFormat(RuntimeError):
    def __init__(self, message: str, profile_id: str | None = None) -> None:
        super().__init__(message)
        self.message = message
        self.profile_id = profile_id


class DuplicateInvoiceNotice(Exception):
    def __init__(self, message: str) -> None:
        super().__init__(message)
        self.message = message


@dataclass
class LocalOCRResult:
    data: dict[str, Any]
    average_confidence: float
    accepted: bool
    reason: str
    raw_text: str


# ── Profile cache ─────────────────────────────────────────────────────
_loaded_profiles: list[document_profiles.DocumentProfile] | None = None


def reload_profiles() -> list[document_profiles.DocumentProfile]:
    """(Re)load all active profiles from disk. Safe to call at any time."""
    global _loaded_profiles
    profiles_dir = BASE_DIR / "data" / "document_profiles"
    profiles = document_profiles.load_profiles(profiles_dir)
    _loaded_profiles = profiles
    logging.info("Reloaded %d document profiles", len(profiles))
    return profiles


def get_profiles() -> list[document_profiles.DocumentProfile]:
    """Get the cached profile list, loading on first call."""
    global _loaded_profiles
    if _loaded_profiles is None:
        _loaded_profiles = reload_profiles()
    return _loaded_profiles if _loaded_profiles else document_profiles.get_default_builtin_profiles()


INVOICE_HEADERS = [
    "Tax Invoice",
    "Date",
    "Item No",
    "Description",
    "Quantity",
    "Quantity Unit",
    "Unit Price",
    "Amount",
]

SYSTEM_PROMPT = """You extract procurement document data (Quotation, Tax Invoice, Delivery Order, Cash Bill, Receipt) from images.
Return only valid JSON matching this schema:
{
  "document_type": "quotation" | "invoice" | "delivery_order" | "cash_bill" | "unknown",
  "supplier_name": string | null,
  "supplier_address": string | null,
  "supplier_phone": string | null,
  "supplier_email": string | null,
  "supplier_bank_account": string | null,
  "tax_invoice": string | null,
  "invoice_date": "YYYY-MM-DD" | null,
  "contact_person": string | null,
  "terms": string | null,
  "confidence": number,
  "notes": string | null,
  "line_items": [
    {
      "item_no": string | null,
      "description": string,
      "quantity": number | null,
      "quantity_unit": string | null,
      "unit_price": number | null,
      "line_total": number | null
    }
  ]
}
Use null when a field is not visible. Make confidence a number from 0 to 1.
For Quotations, set document_type to "quotation", put Quotation reference/number into tax_invoice and date in invoice_date.
For Delivery Orders, set document_type to "delivery_order" and put the visible Delivery Order No in tax_invoice and invoice_date.
For Tax Invoices, set document_type to "invoice" and put the visible Tax Invoice number in tax_invoice and invoice_date.
Extract supplier company name from header letterhead into supplier_name.
Extract full supplier address into supplier_address.
Extract supplier telephone/fax/mobile into supplier_phone.
Extract supplier email into supplier_email.
Extract supplier bank details into supplier_bank_account if visible.
Extract the contact person name and phone number into contact_person, for example "Feddy Sim 016-8868203" or "Farah 011-54302725".
Do not invent document numbers, dates, contact details, item details, prices, or amounts that are not visible.
Never use a phone number, fax number, address number, quantity, amount, or line-item number as tax_invoice or invoice_date."""

EXTRACTION_INSTRUCTIONS = """Extract the document details from this image.

First classify the image:
- If the image title/header says Quotation, Quote, RE: QUOTATION, or Best Price, return document_type "quotation".
- If the image title/label says Delivery Order, Delivery Order No, D.O, return document_type "delivery_order".
- If the image title/label says Tax Invoice, Invoice No, Cash Bill, Receipt, or has price/amount/tax invoice fields, return document_type "invoice".

Important extraction rules:
- Extract supplier/vendor header name, address, telephone/fax, email, and bank account if present.
- Extract document reference number (Invoice No, D.O No, Quotation Ref) into tax_invoice.
- Extract date in YYYY-MM-DD format into invoice_date. For dates printed as DD-MM-YYYY or DD.MM.YYYY, interpret as Day-Month-Year.
- Extract contact person name and phone number into contact_person (e.g. Sales Director, Person to Contact, Attn).
- Treat the line items/products as a row-by-row transcription task.
- Capture every product/item row with full description, quantity, quantity unit, unit price, and line total.
- For items like "Diesel 1600 Lts 4.96 per lts", line_total is 1600 * 4.96 = 7936.00.
- For items like "Transport Charge RM 120", description is "Transport Charge", quantity 1, unit_price 120.00, line_total 120.00.
- Ignore numbered footer notes, validity notices, payment instructions, terms, or signatures as product rows."""

RECONCILIATION_INSTRUCTIONS = """Re-check the document extraction against the image and OCR text.

The OCR product-table text shows line-item row numbers up to {expected_count}. The first JSON extraction returned {actual_count} rows.
Return the same JSON schema, but correct the line_items table so rows 1 through {expected_count} are present when visible.

Rules:
- Use the image as the source of truth.
- Transcribe the line-item table from scratch. Use the first JSON extraction only for header fields such as Delivery Order number, date, and contact person.
- Use the OCR text only to find missed product-table row numbers or to help read table values.
- Ignore numbered notes, payment instructions, terms, signatures, bank details, and any rows outside the product table.
- Do not copy obvious OCR garbage into descriptions.
- Do not preserve first-extraction line items when the image contradicts them.
- Capture any visible unit beside quantity in quantity_unit, such as pcs, boxes, box, unit, nos, set, or roll.
- Verify quantity x unit price equals amount where all three values are visible. Delivery Orders may not have price and amount columns; leave those null when not visible.
- If a row is partially unclear, include the row number with visible fields and mention uncertainty in notes."""


PLACEHOLDER_CREDENTIALS = {
    "your_openai_api_key_here",
    "your_openai_key_here",
    "put_your_openai_api_key_here",
    "your_gemini_api_key_here",
    "your_gemini_key_here",
    "put_your_gemini_api_key_here",
    "your_api_key_here",
    "sk-...",
}

OPENROUTER_BASE_URL = "https://openrouter.ai/api/v1"
DEFAULT_OPENAI_MODEL = "openai/gpt-5.4-mini"
DEFAULT_GEMINI_MODEL = "gemini-3.7-flash"


AUTH_HELP = (
    "AI authentication is not configured correctly. Set GEMINI_API_KEY (from https://aistudio.google.com) "
    "or OPENAI_API_KEY (from https://platform.openai.com/api-keys)."
)

AI_FALLBACK_DISABLED_NOTE = "Automatic AI fallback is disabled. Review the local OCR result before saving."
AI_FALLBACK_AUTH_NOTE = "AI fallback could not run because AI API authentication is not configured correctly. Review carefully before saving."
AI_FALLBACK_CREDITS_NOTE = "AI fallback could not run because the AI provider account is out of credits/quota. Review carefully before saving."
UNKNOWN_DOCUMENT_FORMAT_MESSAGE = (
    "New document format detected. I do not have a local OCR profile for this document yet. "
    "An admin can register this format with the /register_document command."
)


def configure_logging() -> None:
    log_dir = BASE_DIR / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        level=logging.INFO,
        handlers=[
            RotatingFileHandler(log_dir / "invoice_bot.log", maxBytes=5_000_000, backupCount=3, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )
    logging.getLogger("httpx").setLevel(logging.WARNING)


def line_items_from_data(data: dict[str, Any]) -> list[dict[str, Any]]:
    line_items = data.get("line_items", [])
    if not isinstance(line_items, list):
        return []
    return [item for item in line_items if isinstance(item, dict)]


def vertical_rows_from_existing_workbook(workbook: Workbook) -> list[list[Any]]:
    rows: list[list[Any]] = []

    if "Invoices" not in workbook.sheetnames:
        return rows

    invoices = workbook["Invoices"]
    headers = [cell.value for cell in invoices[1]]
    if headers[: len(INVOICE_HEADERS)] == INVOICE_HEADERS:
        for values in invoices.iter_rows(min_row=2, values_only=True):
            rows.append(list(values[: len(INVOICE_HEADERS)]))
        return rows

    legacy_invoice_rows: dict[Any, dict[str, Any]] = {}
    for row_index, values in enumerate(invoices.iter_rows(min_row=2, values_only=True), start=1):
        record = dict(zip(headers, values))
        invoice_id = record.get("invoice_id") or record.get("Invoice ID") or record.get("Tax Invoice") or f"MIGRATED-{row_index:04d}"
        legacy_invoice_rows[invoice_id] = {
            "invoice_id": invoice_id,
            "tax_invoice": record.get("Tax Invoice") or record.get("tax_invoice") or record.get("invoice_number"),
            "invoice_date": record.get("Date") or record.get("invoice_date"),
            "line_items": [],
        }

        item_index = 1
        while True:
            prefix = f"Item {item_index} "
            matching_headers = [header for header in headers if isinstance(header, str) and header.startswith(prefix)]
            if not matching_headers:
                break
            item = {
                "item_no": record.get(f"{prefix}No"),
                "description": record.get(f"{prefix}Description"),
                "quantity": record.get(f"{prefix}Quantity"),
                "quantity_unit": record.get(f"{prefix}Quantity Unit") or record.get(f"{prefix}Unit"),
                "unit_price": record.get(f"{prefix}Unit Price"),
                "line_total": record.get(f"{prefix}Amount"),
            }
            if any(value not in (None, "") for value in item.values()):
                legacy_invoice_rows[invoice_id]["line_items"].append(item)
            item_index += 1

    if "LineItems" in workbook.sheetnames:
        line_items_sheet = workbook["LineItems"]
        item_headers = [cell.value for cell in line_items_sheet[1]]
        for values in line_items_sheet.iter_rows(min_row=2, values_only=True):
            item_record = dict(zip(item_headers, values))
            invoice_id = item_record.get("invoice_id") or item_record.get("Invoice ID") or item_record.get("invoice_number")
            if invoice_id not in legacy_invoice_rows:
                legacy_invoice_rows[invoice_id] = {
                    "invoice_id": invoice_id,
                    "tax_invoice": item_record.get("invoice_number"),
                    "invoice_date": item_record.get("invoice_date"),
                    "line_items": [],
                }
            legacy_invoice_rows[invoice_id]["line_items"].append(
                {
                    "item_no": item_record.get("item_no"),
                    "description": item_record.get("description"),
                    "quantity": item_record.get("quantity"),
                    "quantity_unit": item_record.get("quantity_unit") or item_record.get("unit"),
                    "unit_price": item_record.get("unit_price"),
                    "line_total": item_record.get("line_total"),
                }
            )

    for invoice in legacy_invoice_rows.values():
        line_items = invoice["line_items"]
        if not line_items:
            rows.append(
                [
                    invoice["tax_invoice"],
                    invoice["invoice_date"],
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            continue
        for item in line_items:
            rows.append(
                [
                    invoice["tax_invoice"],
                    invoice["invoice_date"],
                    item.get("item_no"),
                    item.get("description"),
                    normalize_number(item.get("quantity")),
                    normalize_quantity_unit(item.get("quantity_unit")),
                    normalize_number(item.get("unit_price")),
                    normalize_number(item.get("line_total")),
                ]
            )

    return rows


def ensure_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        ensure_workbook_headers(path)
        return

    workbook = Workbook()
    invoices = workbook.active
    invoices.title = "Invoices"
    invoices.append(INVOICE_HEADERS)
    workbook.save(path)


def ensure_workbook_headers(path: Path) -> None:
    workbook = load_workbook(path)
    changed = False

    if "Invoices" not in workbook.sheetnames:
        worksheet = workbook.create_sheet("Invoices", 0)
        worksheet.append(INVOICE_HEADERS)
        changed = True
    else:
        worksheet = workbook["Invoices"]
        existing = [cell.value for cell in worksheet[1]]

        if existing[: len(INVOICE_HEADERS)] != INVOICE_HEADERS or worksheet.max_column != len(INVOICE_HEADERS) or len(workbook.sheetnames) > 1:
            rows = vertical_rows_from_existing_workbook(workbook)
            migrated = workbook.create_sheet("Invoices_Migrated", 0)
            migrated.append(INVOICE_HEADERS)
            for row in rows:
                migrated.append(row)
            for sheet in list(workbook.worksheets):
                if sheet.title != "Invoices_Migrated":
                    workbook.remove(sheet)
            migrated.title = "Invoices"
            changed = True
        else:
            for header in INVOICE_HEADERS:
                if header not in existing:
                    worksheet.cell(row=1, column=len(existing) + 1, value=header)
                    existing.append(header)
                    changed = True

    if changed:
        try:
            workbook.save(path)
        except PermissionError:
            logging.warning("Could not update workbook headers because the workbook is open: %s", path)


def assert_workbook_writable(path: Path) -> None:
    if not path.exists():
        return
    lock_file = path.with_name(f"~${path.name}")
    if lock_file.exists():
        raise WorkbookBusyError(f"Workbook is open or locked: {path}")
    try:
        with path.open("a+b"):
            pass
    except PermissionError as exc:
        raise WorkbookBusyError(f"Workbook is open or locked: {path}") from exc


def normalize_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace(",", "").strip()
        return float(cleaned)
    except ValueError:
        return None


def normalize_quantity_unit(value: Any) -> str:
    if value is None:
        return ""
    unit = str(value).strip()
    if not unit:
        return ""
    unit = re.sub(r"[^A-Za-z0-9./ -]+", "", unit)
    unit = re.sub(r"\s+", " ", unit).strip(" ./-")
    aliases = {
        "pcs": "pcs",
        "pes": "pcs",
        "pts": "pcs",
        "pc": "pcs",
        "piece": "pcs",
        "pieces": "pcs",
        "ton": "ton",
        "tor": "ton",
        "toll": "roll",
    }
    unit = aliases.get(unit.lower(), unit)
    return unit[:20]


def normalize_quantity_fields(item: dict[str, Any]) -> None:
    quantity = item.get("quantity")
    quantity_unit = normalize_quantity_unit(item.get("quantity_unit") or item.get("unit"))

    if isinstance(quantity, str):
        match = re.match(r"^\s*([\d,]+(?:[.]\d+)?)\s*([A-Za-z][A-Za-z0-9./ -]{0,19})\s*$", quantity)
        if match:
            quantity = match.group(1)
            if not quantity_unit:
                quantity_unit = normalize_quantity_unit(match.group(2))

    item["quantity"] = normalize_number(quantity)
    item["quantity_unit"] = quantity_unit


def normalize_extracted_data(data: dict[str, Any]) -> None:
    if not data.get("tax_invoice"):
        data["tax_invoice"] = data.get("delivery_order_no") or data.get("do_number") or data.get("invoice_number")
    if not data.get("invoice_date"):
        data["invoice_date"] = data.get("delivery_order_date") or data.get("do_date")
    data["document_type"] = normalize_document_type(data)
    for item in line_items_from_data(data):
        normalize_quantity_fields(item)


def line_item_has_value(item: dict[str, Any]) -> bool:
    return any(
        item.get(field) not in (None, "")
        for field in ["description", "quantity", "quantity_unit", "unit_price", "line_total"]
    )


def drop_blank_line_items(data: dict[str, Any]) -> None:
    line_items = line_items_from_data(data)
    if not line_items:
        return
    data["line_items"] = [item for item in line_items if line_item_has_value(item)]


def format_quantity_with_unit(item: dict[str, Any]) -> Any:
    unit = normalize_quantity_unit(item.get("quantity_unit") or item.get("unit"))
    raw_quantity = item.get("quantity")
    quantity = normalize_number(raw_quantity)
    if quantity is None:
        return raw_quantity or unit or None
    if almost_whole_number(quantity):
        quantity_text = str(int(round(quantity)))
    else:
        quantity_text = str(quantity).rstrip("0").rstrip(".")
    return f"{quantity_text} {unit}" if unit else quantity


def almost_whole_number(value: float, tolerance: float = 0.01) -> bool:
    return abs(value - round(value)) <= tolerance


def is_subsequence(needle: str, haystack: str) -> bool:
    position = 0
    for character in haystack:
        if position < len(needle) and needle[position] == character:
            position += 1
    return position == len(needle)


def repair_line_item_arithmetic(data: dict[str, Any]) -> None:
    normalize_extracted_data(data)
    drop_blank_line_items(data)
    for item in line_items_from_data(data):
        quantity = normalize_number(item.get("quantity"))
        unit_price = normalize_number(item.get("unit_price"))
        line_total = normalize_number(item.get("line_total"))
        if quantity is None or unit_price in (None, 0) or line_total is None:
            continue

        calculated_total = quantity * unit_price
        if abs(calculated_total - line_total) <= 0.02:
            continue

        implied_quantity = line_total / unit_price
        if implied_quantity <= 0 or not almost_whole_number(implied_quantity):
            continue

        implied_int = int(round(implied_quantity))
        quantity_int = int(quantity) if almost_whole_number(quantity) else None
        if quantity_int is not None and is_subsequence(str(quantity_int), str(implied_int)):
            item["quantity"] = implied_int


def env_float(name: str, default: float) -> float:
    value = os.getenv(name)
    if value in (None, ""):
        return default
    try:
        return float(value)
    except ValueError:
        logging.warning("Invalid float for %s=%r. Using default %s.", name, value, default)
        return default


def env_int(name: str, default: int) -> int:
    value = os.getenv(name)
    if value in (None, ""):
        return default
    try:
        return int(value)
    except ValueError:
        logging.warning("Invalid integer for %s=%r. Using default %s.", name, value, default)
        return default


def env_bool(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value in (None, ""):
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def ai_fallback_enabled() -> bool:
    has_ai_key = bool(
        gemini_api_key()
        or os.getenv("OPENAI_API_KEY")
        or os.getenv("OPENROUTER_API_KEY")
        or os.getenv("OPENAI_BEARER_TOKEN")
        or os.getenv("OPENAI_TOKEN_COMMAND")
    )
    return env_bool("AI_FALLBACK_ENABLED", has_ai_key)


def ai_primary_enabled() -> bool:
    has_ai_key = bool(
        gemini_api_key()
        or os.getenv("OPENAI_API_KEY")
        or os.getenv("OPENROUTER_API_KEY")
        or os.getenv("OPENAI_BEARER_TOKEN")
        or os.getenv("OPENAI_TOKEN_COMMAND")
    )
    return env_bool("AI_PRIMARY_ENABLED", has_ai_key)


def configured_allowed_chat_ids() -> set[str]:
    values: list[str] = []
    for name in ["TELEGRAM_ALLOWED_CHAT_IDS", "TELEGRAM_ALLOWED_CHAT_ID"]:
        raw_value = os.getenv(name, "")
        values.extend(part.strip() for part in re.split(r"[,;\s]+", raw_value) if part.strip())
    return set(values)


def configured_test_chat_ids() -> set[str]:
    raw_value = os.getenv("TELEGRAM_TEST_CHAT_IDS")
    if raw_value in (None, ""):
        return set(DEFAULT_TEST_CHAT_IDS)
    return {part.strip() for part in re.split(r"[,;\s]+", raw_value) if part.strip()}


def is_test_submitter(submitter_chat_id: Any) -> bool:
    return str(submitter_chat_id or "").strip() in configured_test_chat_ids()


def invoice_record_type(submitter_chat_id: Any = None, force_record: bool = False) -> str:
    if force_record:
        return "record"
    return "test" if is_test_submitter(submitter_chat_id) else "record"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def normalize_document_type(data: dict[str, Any]) -> str:
    raw_type = normalize_text(data.get("document_type") or data.get("document_kind") or data.get("type"))
    if raw_type in {"matched_pair", "matched pair", "pair", "do_invoice_pair", "d.o invoice pair"}:
        return "matched_pair"
    if raw_type in {"quotation", "quote", "re: quotation", "re quotation", "proposal", "proforma"}:
        return DOCUMENT_TYPE_QUOTATION
    if raw_type in {"cash_bill", "cash bill", "receipt", "official receipt"}:
        return DOCUMENT_TYPE_CASH_BILL
    if raw_type in {"delivery_order", "delivery order", "do", "d.o", "d/o"}:
        return DOCUMENT_TYPE_DELIVERY_ORDER
    if raw_type in {"invoice", "tax invoice", "tax_invoice"}:
        return DOCUMENT_TYPE_INVOICE

    line_items = line_items_from_data(data)
    has_prices = any(item.get("unit_price") not in (None, "") or item.get("line_total") not in (None, "") for item in line_items)
    has_contact = bool(data.get("contact_person") or data.get("person_to_contact"))
    notes = normalize_text(data.get("notes"))
    method = normalize_text(data.get("extraction_method"))
    if "quotation" in notes or "quote" in notes:
        return DOCUMENT_TYPE_QUOTATION
    if has_prices or "tax invoice" in notes:
        return DOCUMENT_TYPE_INVOICE
    if has_contact or "delivery order" in notes or "do" in method:
        return DOCUMENT_TYPE_DELIVERY_ORDER
    return DOCUMENT_TYPE_DELIVERY_ORDER


def document_type_label(document_type: str) -> str:
    if document_type == DOCUMENT_TYPE_QUOTATION:
        return "Quotation"
    if document_type == DOCUMENT_TYPE_CASH_BILL:
        return "Cash Bill"
    if document_type == DOCUMENT_TYPE_INVOICE:
        return "Invoice"
    if document_type == DOCUMENT_TYPE_DELIVERY_ORDER:
        return "D.O"
    return "Document"


def normalize_invoice_number(value: Any) -> str:
    if value is None:
        return ""
    normalized = re.sub(r"[^A-Z0-9]", "", str(value).upper())
    normalized = re.sub(r"(?<=\d)O(?=\d)", "0", normalized)
    normalized = re.sub(r"(?<=[A-Z])O(?=\d)", "0", normalized)
    return normalized


def normalize_visible_document_number(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).upper().strip()
    text = re.sub(r"[\s]+", "", text)
    text = text.strip(" .,:;-'\"()[]{}<>")
    text = re.sub(r"[\u2010-\u2015]+$", "", text)
    text = re.sub(r"[^A-Z0-9/-]", "", text)
    if not text:
        return None

    text = re.sub(r"^[1I]G(?=[-/])", "TG", text)
    text = re.sub(r"^T6(?=[-/])", "TG", text)
    if text.startswith("TG-"):
        prefix, suffix = text.split("-", 1)
        suffix = suffix.replace("O", "0")
        suffix = suffix.replace("B", "8")
        return f"{prefix}-{suffix}"
    return text


def is_plausible_document_number(value: str | None) -> bool:
    if not value:
        return False
    return bool(re.match(r"^TG-[A-Z0-9]{4,}$", value))


def normalize_item_no(value: Any) -> str:
    if value is None:
        return ""
    match = re.search(r"\d+", str(value))
    return match.group(0) if match else normalize_text(value)


def extracted_item_numbers(data: dict[str, Any]) -> set[int]:
    numbers: set[int] = set()
    for item in line_items_from_data(data):
        item_no = normalize_item_no(item.get("item_no"))
        if not item_no:
            continue
        try:
            numbers.add(int(item_no))
        except ValueError:
            continue
    return numbers


def ocr_product_table_lines(text: str) -> list[str]:
    lines: list[str] = []
    in_table = False
    header_seen = False
    terminators = [
        "ringgit malaysia",
        "gross",
        "discount",
        "total payable",
        "notes",
        "all cheque",
        "online payment",
        "authorised",
        "authorized",
        "signature",
    ]

    for raw_line in text.splitlines():
        line = " ".join(raw_line.split())
        if not line:
            continue

        lowered = line.lower()
        header_like = "product description" in lowered or (
            "quantity" in lowered and ("unit price" in lowered or "amount" in lowered)
        )
        if header_like:
            in_table = True
            header_seen = True
            continue

        if in_table and any(term in lowered for term in terminators):
            in_table = False
            continue

        if in_table:
            lines.append(line)

    return lines if header_seen else []


def expected_item_count_from_ocr(text: str) -> int | None:
    visible_numbers: set[int] = set()
    for line in ocr_product_table_lines(text):
        if not line:
            continue
        match = re.match(r"^(\d{1,2})\s*(?:[.)]|\s|[-\u2010-\u2015])+", line)
        if not match:
            continue
        number = int(match.group(1))
        if 1 <= number <= 30:
            visible_numbers.add(number)

    if not visible_numbers:
        return None

    best_max = None
    for candidate in sorted(visible_numbers):
        if candidate < 3:
            continue
        present = sum(1 for number in range(1, candidate + 1) if number in visible_numbers)
        if present >= candidate - 1:
            best_max = candidate
    return best_max


def extraction_needs_reconciliation(data: dict[str, Any], ocr_text: str) -> tuple[int, int] | None:
    expected_count = expected_item_count_from_ocr(ocr_text)
    if not expected_count:
        return None

    line_items = line_items_from_data(data)
    actual_count = len(line_items)
    item_numbers = extracted_item_numbers(data)
    has_all_expected_numbers = all(number in item_numbers for number in range(1, expected_count + 1))
    if actual_count < expected_count or not has_all_expected_numbers:
        return expected_count, actual_count
    return None


def item_signature(item: dict[str, Any]) -> tuple[str, str, float | None, str, float | None, float | None]:
    return (
        normalize_item_no(item.get("item_no")),
        normalize_text(item.get("description")),
        normalize_number(item.get("quantity")),
        normalize_quantity_unit(item.get("quantity_unit")).lower(),
        normalize_number(item.get("unit_price")),
        normalize_number(item.get("line_total")),
    )


def normalized_description_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(value))


def item_match_key(item: dict[str, Any]) -> tuple[str, str]:
    item_no = normalize_item_no(item.get("item_no"))
    description = normalized_description_key(item.get("description"))
    return item_no, description


def find_matching_invoice_item(do_item: dict[str, Any], invoice_items: list[dict[str, Any]]) -> dict[str, Any] | None:
    do_item_no, do_description = item_match_key(do_item)
    if do_item_no:
        for item in invoice_items:
            if normalize_item_no(item.get("item_no")) == do_item_no:
                return item

    if do_description:
        for item in invoice_items:
            invoice_description = normalized_description_key(item.get("description"))
            if invoice_description and (invoice_description == do_description or do_description in invoice_description or invoice_description in do_description):
                return item
    return None


def compare_and_merge_documents(
    delivery_order: dict[str, Any],
    invoice: dict[str, Any],
    item_source: str = DOCUMENT_TYPE_DELIVERY_ORDER,
) -> tuple[dict[str, Any], list[str]]:
    merged = json.loads(json.dumps(delivery_order))
    warnings: list[str] = []

    do_number = delivery_order.get("tax_invoice") or delivery_order.get("invoice_number")
    invoice_number = invoice.get("tax_invoice") or invoice.get("invoice_number")
    if do_number and invoice_number and normalize_invoice_number(do_number) != normalize_invoice_number(invoice_number):
        warnings.append(f"D.O number {do_number} does not match invoice number {invoice_number}.")
    elif not do_number:
        warnings.append("D.O number is missing.")
    elif not invoice_number:
        warnings.append("Invoice number is missing.")

    do_date = delivery_order.get("invoice_date") or delivery_order.get("delivery_order_date") or delivery_order.get("do_date")
    invoice_date = invoice.get("invoice_date") or invoice.get("invoice_document_date")
    if invoice_date:
        merged["invoice_document_date"] = invoice_date
        merged["invoice_date"] = invoice_date
    if do_date:
        merged["delivery_order_date"] = do_date
        merged["delivery_order_document_date"] = do_date
        if not invoice_date:
            merged["invoice_date"] = do_date
            warnings.append("Invoice date is missing, so D.O date was used as fallback.")
    elif not invoice_date:
        warnings.append("Both invoice date and D.O date are missing.")

    merged["tax_invoice"] = do_number or invoice_number
    merged["invoice_number"] = invoice_number or do_number
    merged["delivery_order_no"] = do_number or invoice_number
    merged["supplier_name"] = detected_supplier_name(
        {
            "delivery_order_data": delivery_order,
            "invoice_data": invoice,
            "tax_invoice": do_number or invoice_number,
            "notes": " ".join(str(part) for part in [delivery_order.get("notes"), invoice.get("notes")] if part),
        }
    )
    merged["delivery_order_contact_person"] = delivery_order.get("contact_person") or delivery_order.get("person_to_contact") or ""
    merged["contact_person"] = merged["delivery_order_contact_person"]
    merged["document_type"] = "matched_pair"

    do_items = line_items_from_data(delivery_order)
    invoice_items = line_items_from_data(invoice)
    matched_invoice_indexes: set[int] = set()
    merged_items: list[dict[str, Any]] = []

    for do_item in do_items:
        merged_item = dict(do_item)
        match = find_matching_invoice_item(do_item, invoice_items)
        if match:
            try:
                matched_invoice_indexes.add(invoice_items.index(match))
            except ValueError:
                pass
            do_quantity = normalize_number(do_item.get("quantity"))
            invoice_quantity = normalize_number(match.get("quantity"))
            if do_quantity is not None and invoice_quantity is not None and abs(do_quantity - invoice_quantity) > 0.01:
                item_label = do_item.get("item_no") or do_item.get("description") or "unknown item"
                warnings.append(f"Quantity differs for item {item_label}: D.O {format_quantity_with_unit(do_item)}, invoice {format_quantity_with_unit(match)}.")

            if not merged_item.get("unit_price") and match.get("unit_price") not in (None, ""):
                merged_item["unit_price"] = match.get("unit_price")
            if not merged_item.get("line_total") and match.get("line_total") not in (None, ""):
                merged_item["line_total"] = match.get("line_total")
            if not merged_item.get("description") and match.get("description"):
                merged_item["description"] = match.get("description")
        else:
            item_label = do_item.get("item_no") or do_item.get("description") or "unknown item"
            warnings.append(f"D.O item {item_label} was not found in the invoice.")
        merged_items.append(merged_item)

    extra_invoice_items = [
        item
        for index, item in enumerate(invoice_items)
        if index not in matched_invoice_indexes and line_item_has_value(item)
    ]
    if extra_invoice_items:
        labels = ", ".join(str(item.get("item_no") or item.get("description") or "?") for item in extra_invoice_items[:5])
        warnings.append(f"Invoice has extra item row(s) not found in D.O: {labels}.")

    if item_source == DOCUMENT_TYPE_INVOICE:
        merged["line_items"] = [dict(item) for item in invoice_items if line_item_has_value(item)] or merged_items
        merged["item_source"] = DOCUMENT_TYPE_INVOICE
    else:
        merged["line_items"] = merged_items or invoice_items
        merged["item_source"] = DOCUMENT_TYPE_DELIVERY_ORDER
    merged["pair_compare_warnings"] = warnings
    merged["extraction_method"] = "do_invoice_pair"
    note_parts = [delivery_order.get("notes"), invoice.get("notes")]
    if warnings:
        note_parts.append("Pair comparison warnings: " + "; ".join(warnings))
    merged["notes"] = " ".join(str(part).strip() for part in note_parts if part)
    repair_line_item_arithmetic(merged)
    return merged, warnings


def existing_items_for_tax_invoice(workbook: Workbook, tax_invoice: str) -> list[dict[str, Any]]:
    if "Invoices" not in workbook.sheetnames:
        return []

    worksheet = workbook["Invoices"]
    headers = [cell.value for cell in worksheet[1]]
    existing_items: list[dict[str, Any]] = []
    target_tax_invoice = normalize_invoice_number(tax_invoice)
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, values))
        if normalize_invoice_number(record.get("Tax Invoice")) != target_tax_invoice:
            continue
        existing_items.append(
            {
                "item_no": record.get("Item No"),
                "description": record.get("Description"),
                "quantity": record.get("Quantity"),
                "quantity_unit": record.get("Quantity Unit"),
                "unit_price": record.get("Unit Price"),
                "line_total": record.get("Amount"),
            }
        )
    return existing_items


def split_new_and_repeated_items(
    existing_items: list[dict[str, Any]],
    extracted_items: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    existing_signatures = {item_signature(item) for item in existing_items}
    existing_item_numbers = {normalize_item_no(item.get("item_no")) for item in existing_items if normalize_item_no(item.get("item_no"))}
    new_items = []
    repeated_items = []
    for item in extracted_items:
        item_no = normalize_item_no(item.get("item_no"))
        if item_no and existing_item_numbers:
            if item_no in existing_item_numbers:
                repeated_items.append(item)
            else:
                new_items.append(item)
        elif item_signature(item) in existing_signatures:
            repeated_items.append(item)
        else:
            new_items.append(item)
    return new_items, repeated_items


def invoice_id_from_timestamp(received_at: datetime) -> str:
    return received_at.strftime("I%Y%m%d%H%M%S%f")


def gemini_api_key() -> str | None:
    for env_var in ("GEMINI_API_KEY", "GOOGLE_API_KEY"):
        key = os.getenv(env_var)
        if key and key.strip() and key.strip() not in PLACEHOLDER_CREDENTIALS:
            return key.strip()
    return None


def gemini_model_name() -> str:
    model = os.getenv("GEMINI_MODEL", DEFAULT_GEMINI_MODEL).strip()
    if model in ("gemini-2.5-flash", "gemini-1.5-flash", "gemini-1.5-pro", "gemini-2.0-flash"):
        return "gemini-3.7-flash"
    return model


def openai_bearer_credential() -> str:
    token_command = os.getenv("OPENAI_TOKEN_COMMAND")
    if token_command:
        result = subprocess.run(
            token_command,
            check=True,
            capture_output=True,
            shell=True,
            text=True,
        )
        token = result.stdout.strip()
        if not token:
            raise RuntimeError("OPENAI_TOKEN_COMMAND returned an empty token.")
        return token

    bearer_token = os.getenv("OPENAI_BEARER_TOKEN")
    if bearer_token:
        if bearer_token.strip() in PLACEHOLDER_CREDENTIALS:
            raise RuntimeError(AUTH_HELP)
        return bearer_token

    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        if api_key.strip() in PLACEHOLDER_CREDENTIALS:
            raise RuntimeError(AUTH_HELP)
        return api_key

    openrouter_key = os.getenv("OPENROUTER_API_KEY")
    if openrouter_key:
        if openrouter_key.strip() in PLACEHOLDER_CREDENTIALS:
            raise RuntimeError(AUTH_HELP)
        return openrouter_key

    raise RuntimeError(AUTH_HELP)


def openai_base_url() -> str | None:
    """Return the base URL for the OpenAI client.

    Returns OpenRouter endpoint when OPENROUTER_API_KEY is set,
    or None to use the default OpenAI endpoint.
    """
    if os.getenv("OPENROUTER_API_KEY"):
        return OPENROUTER_BASE_URL
    return None


def openai_model_name() -> str:
    model = os.getenv("OPENAI_MODEL", DEFAULT_OPENAI_MODEL).strip()
    # OpenRouter needs the full model identifier (e.g. "openai/gpt-5.4-mini").
    # Native OpenAI API needs just the model name (e.g. "gpt-5.4-mini").
    if os.getenv("OPENROUTER_API_KEY"):
        return model
    if model.startswith("openai/"):
        return model.split("/", 1)[1]
    return model


def configured_ai_provider() -> str:
    explicit = os.getenv("AI_PROVIDER", "").strip().lower()
    if explicit in {"gemini", "google"}:
        return "gemini"
    if explicit in {"openai", "openrouter"}:
        return "openai"
    if gemini_api_key():
        return "gemini"
    return "openai"


def ai_model_name() -> str:
    if configured_ai_provider() == "gemini":
        return gemini_model_name()
    return openai_model_name()


def is_openai_auth_error(exc: Exception) -> bool:
    if isinstance(exc, AuthenticationError) or str(exc) == AUTH_HELP:
        return True
    exc_str = str(exc).lower()
    return "api_key_invalid" in exc_str or "unauthenticated" in exc_str or "permissiondenied" in exc_str or "401" in exc_str or "403" in exc_str


def is_openai_credit_error(exc: Exception) -> bool:
    """True for 402 Payment Required or 429 Quota Exceeded."""
    if isinstance(exc, APIStatusError) and exc.status_code == 402:
        return True
    exc_str = str(exc).lower()
    return "resource_exhausted" in exc_str or "quota" in exc_str or "402" in exc_str or "insufficient_quota" in exc_str


def configured_invoice_workbook_dir() -> Path:
    value = os.getenv("INVOICE_WORKBOOK_DIR", str(DEFAULT_INVOICE_WORKBOOK_DIR))
    path = Path(value)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path.resolve()


def sender_invoice_workbook_dir(submitter_chat_id: Any = None) -> Path:
    workbook_dir = configured_invoice_workbook_dir()
    procurement_dir = configured_procurement_dir()
    sender_id = str(submitter_chat_id or "").strip()
    if not sender_id:
        return workbook_dir / "unknown_sender"
    safe_sender_id = re.sub(r"[^0-9A-Za-z_-]+", "_", sender_id).strip("_") or "unknown_sender"
    return workbook_dir / safe_sender_id


def configured_invoice_template_path() -> Path | None:
    value = os.getenv("INVOICE_TEMPLATE_PATH")
    if value not in (None, ""):
        path = Path(value)
        if not path.is_absolute():
            path = BASE_DIR / path
        return path.resolve()

    for candidate in (
        DEFAULT_INVOICE_TEMPLATE_PATH,
        BASE_DIR / "templates" / "purchase_order_template.xlsx",
        BASE_DIR / "data" / "templates" / "purchase_order_template.xlsx",
    ):
        if candidate.exists():
            return candidate.resolve()
    return None


def configured_material_requisition_template_path() -> Path | None:
    value = os.getenv("MATERIAL_REQUISITION_TEMPLATE_PATH")
    if value not in (None, ""):
        path = Path(value)
        if not path.is_absolute():
            path = BASE_DIR / path
        return path.resolve()

    for candidate in (
        DEFAULT_MATERIAL_REQUISITION_TEMPLATE_PATH,
        BASE_DIR / "templates" / "material_requisition_template.xlsx",
        BASE_DIR / "data" / "templates" / "material_requisition_template.xlsx",
    ):
        if candidate.exists():
            return candidate.resolve()
    return None


def configured_invoice_register_path() -> Path:
    value = os.getenv("INVOICE_REGISTER_PATH", str(DEFAULT_INVOICE_REGISTER_PATH))
    path = Path(value)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path.resolve()


def configured_procurement_dir() -> Path:
    value = os.getenv("PROCUREMENT_DIR", str(DEFAULT_PROCUREMENT_DIR))
    path = Path(value)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path.resolve()


def configured_procurement_supplier_name() -> str:
    return os.getenv("PROCUREMENT_SUPPLIER_NAME", DEFAULT_PROCUREMENT_SUPPLIER_NAME).strip() or DEFAULT_PROCUREMENT_SUPPLIER_NAME


def configured_default_supplier() -> str:
    return configured_procurement_supplier_name()


def configured_supplier_aliases() -> dict[str, list[str]]:
    return suppliers.supplier_aliases_from_env(os.getenv("SUPPLIER_ALIASES"))


def detected_supplier_name(data: dict[str, Any]) -> str:
    return suppliers.detect_supplier_name(
        data,
        configured_procurement_supplier_name(),
        configured_supplier_aliases(),
    )


def configured_cleanup_archive_dir() -> Path:
    value = os.getenv("CLEANUP_ARCHIVE_DIR", str(DEFAULT_CLEANUP_ARCHIVE_DIR))
    path = Path(value)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path.resolve()


def configured_cleanup_retention_days() -> int:
    return max(1, env_int("CLEANUP_RETENTION_DAYS", DEFAULT_CLEANUP_RETENTION_DAYS))


def cleanup_target_dirs() -> list[Path]:
    raw_value = os.getenv("CLEANUP_TARGET_DIRS")
    if not raw_value:
        return [ENHANCED_DIR, OCR_DIR]
    paths: list[Path] = []
    for part in re.split(r"[,;]+", raw_value):
        value = part.strip()
        if not value:
            continue
        path = Path(value)
        if not path.is_absolute():
            path = BASE_DIR / path
        paths.append(path.resolve())
    return paths or [ENHANCED_DIR, OCR_DIR]


def run_cleanup_retention(dry_run: bool = False) -> retention.RetentionResult:
    return retention.archive_old_files(
        cleanup_target_dirs(),
        configured_cleanup_archive_dir(),
        configured_cleanup_retention_days(),
        dry_run=dry_run,
    )


def safe_workbook_stem(value: Any, fallback: str) -> str:
    source = str(value or "").strip() or fallback
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", source)
    stem = re.sub(r"\s+", " ", stem)
    stem = stem.strip(" ._")
    return stem[:100] or fallback


def parse_date_or_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, datetime_date):
        return datetime.combine(value, datetime.min.time())
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass
    try:
        return datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        pass
    return None


def po_month_name(source: Any = None) -> str:
    when = parse_date_or_datetime(source)
    if when is None:
        when = datetime.now(timezone.utc)
    if when.tzinfo is not None:
        when = when.astimezone()
    return when.strftime("%B").upper()


def po_month_key(source: Any = None) -> str:
    when = parse_date_or_datetime(source)
    if when is None:
        when = datetime.now(timezone.utc)
    if when.tzinfo is not None:
        when = when.astimezone()
    return when.strftime("%Y-%m")


def po_filename_pattern(month_name: str, record_type: str = "record") -> re.Pattern[str]:
    prefix = TEST_PO_FILENAME_PREFIX if record_type == "test" else PO_FILENAME_PREFIX
    return re.compile(rf"^{re.escape(prefix)} {re.escape(month_name)} (\d{{4,}})$", re.IGNORECASE)


def used_po_running_numbers(month_name: str, month_key: str, record_type: str = "record") -> set[int]:
    used: set[int] = set()
    pattern = po_filename_pattern(month_name, record_type)
    workbook_dir = configured_invoice_workbook_dir()
    if workbook_dir.exists():
        for path in workbook_dir.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in {".xlsx", ".pdf"}:
                continue
            match = pattern.match(path.stem)
            if match:
                used.add(int(match.group(1)))

    register_path = configured_invoice_register_path()
    if register_path.exists():
        with register_path.open("r", newline="", encoding="utf-8") as handle:
            for row in csv.DictReader(handle):
                if row.get("po_month_key") != month_key:
                    continue
                row_record_type = (row.get("record_type") or "record").strip().lower()
                if row_record_type != record_type:
                    continue
                value = row.get("po_running_number")
                if value and value.isdigit():
                    used.add(int(value))
                    continue
                workbook_file = row.get("workbook_file")
                if workbook_file:
                    match = pattern.match(Path(workbook_file).stem)
                    if match:
                        used.add(int(match.group(1)))

    return used


def next_po_running_number(month_name: str, month_key: str, record_type: str = "record") -> int:
    used = used_po_running_numbers(month_name, month_key, record_type)
    number = 1
    while number in used:
        number += 1
    return number


def parse_po_running_number(value: Any) -> int | None:
    text = str(value or "").strip()
    if not re.fullmatch(r"\d{1,6}", text):
        return None
    number = int(text)
    return number if number > 0 else None


def po_output_stem_for_running_number(
    running_number: int,
    received_at: Any = None,
    record_type: str = "record",
) -> tuple[str, str, str]:
    month_name = po_month_name(received_at)
    month_key = po_month_key(received_at)
    normalized_record_type = (record_type or "record").strip().lower()
    prefix = TEST_PO_FILENAME_PREFIX if normalized_record_type == "test" else PO_FILENAME_PREFIX
    return f"{prefix} {month_name} {running_number:04d}", month_name, month_key


def apply_manual_po_running_number(
    data: dict[str, Any],
    running_number: int,
    received_at: datetime | None = None,
    record_type: str = "record",
) -> str:
    doc_date = (
        data.get("invoice_date")
        or data.get("invoice_document_date")
        or data.get("delivery_order_date")
        or data.get("do_date")
        or received_at
    )
    stem, month_name, month_key = po_output_stem_for_running_number(running_number, doc_date, record_type)
    data["po_month_key"] = month_key
    data["po_month_name"] = month_name
    data["po_running_number"] = f"{running_number:04d}"
    data["po_output_stem"] = stem
    data["record_type"] = (record_type or "record").strip().lower()
    return stem


def manual_po_running_number_is_available(
    running_number: int,
    received_at: Any = None,
    record_type: str = "record",
) -> bool:
    _, month_name, month_key = po_output_stem_for_running_number(running_number, received_at, record_type)
    return running_number not in used_po_running_numbers(month_name, month_key, record_type)


def get_user_initials(submitter: Any) -> str:
    name = ""
    if isinstance(submitter, str):
        name = submitter
    elif submitter and hasattr(submitter, "full_name"):
        name = getattr(submitter, "full_name") or ""
    elif submitter and hasattr(submitter, "first_name"):
        first = getattr(submitter, "first_name") or ""
        last = getattr(submitter, "last_name") or ""
        name = f"{first} {last}".strip()

    clean_name = re.sub(r"[^A-Za-z\s]+", "", name).strip()
    words = [w for w in clean_name.split() if w.lower() not in ("mrs", "mr", "ms", "miss", "dr", "encik", "puan")]
    if not words:
        return "BFE"
    if len(words) == 1:
        return words[0][:2].upper()
    return "".join(w[0].upper() for w in words[:3])


def po_month_mmyy(value: Any) -> str:
    parsed = parse_date_or_datetime(value)
    if not parsed:
        parsed = datetime.now()
    return parsed.strftime("%m%y")


def ensure_po_output_stem(data: dict[str, Any], received_at: datetime | None = None) -> str:
    existing = data.get("po_output_stem")
    if existing:
        return safe_workbook_stem(existing, str(existing))

    doc_date = (
        data.get("invoice_date")
        or data.get("invoice_document_date")
        or data.get("delivery_order_date")
        or data.get("do_date")
        or received_at
    )
    supplier_profile = data.get("supplier_profile") or suppliers.detect_supplier_profile(
        data, configured_default_supplier(), configured_supplier_aliases()
    )
    data["supplier_profile"] = supplier_profile
    category = (supplier_profile.get("category") or "TUJU").upper()
    supplier_display_name = supplier_profile.get("display_name") or "TUJU GALAKSI SDN BHD"

    record_type = (data.get("record_type") or invoice_record_type(data.get("submitter_chat_id"))).strip().lower()
    is_test = record_type == "test"
    test_prefix = "TEST " if is_test else ""

    submitter_name = data.get("submitter_name") or data.get("requested_by") or ""
    initials = get_user_initials(submitter_name)

    if category == "TECH":
        mmyy = po_month_mmyy(doc_date)
        running_key = f"TECH_{mmyy}"
        running_number = next_po_running_number(running_key, running_key, record_type)
        stem = f"{test_prefix}BFE PO TECH {mmyy} {running_number:04d} {supplier_display_name}"
        pr_number = f"BFE/PO/TECH/{initials}/{mmyy}-{running_number:04d}"
        data["po_month_key"] = running_key
        data["po_month_name"] = mmyy
    elif category == "TUJU":
        month_name = po_month_name(doc_date)
        month_key = po_month_key(doc_date)
        prefix = TEST_PO_FILENAME_PREFIX if is_test else PO_FILENAME_PREFIX
        running_number = next_po_running_number(month_name, month_key, record_type)
        stem = f"{prefix} {month_name} {running_number:04d}"
        pr_number = stem
        data["po_month_key"] = month_key
        data["po_month_name"] = month_name
    else:
        mmyy = po_month_mmyy(doc_date)
        running_key = f"{category}_{mmyy}"
        running_number = next_po_running_number(running_key, running_key, record_type)
        stem = f"{test_prefix}BFE PO {category} {mmyy} {running_number:04d} {supplier_display_name}"
        pr_number = f"BFE/PO/{category}/{initials}/{mmyy}-{running_number:04d}"
        data["po_month_key"] = running_key
        data["po_month_name"] = mmyy

    data["po_running_number"] = f"{running_number:04d}"
    data["po_output_stem"] = stem
    data["pr_number"] = pr_number
    data["record_type"] = record_type
    return stem


def invoice_workbook_path(
    invoice_id: str,
    data: dict[str, Any],
    received_at: datetime | None = None,
    submitter_chat_id: Any = None,
) -> Path:
    stem = ensure_po_output_stem(data, received_at)
    filename = f"{safe_workbook_stem(stem, invoice_id)}.xlsx"
    return sender_invoice_workbook_dir(submitter_chat_id or data.get("submitter_chat_id")) / filename


def image_stat_mean_and_std(image: Image.Image) -> tuple[float, float]:
    stat = ImageStat.Stat(image)
    mean = float(stat.mean[0]) if stat.mean else 0.0
    stddev = float(stat.stddev[0]) if stat.stddev else 0.0
    return mean, stddev


def image_sharpness_score(gray_image: Image.Image) -> float:
    """Approximate sharpness without adding OpenCV as a hard dependency.

    Higher edge variance usually means text strokes are clearer. This is not a
    perfect blur detector, but it is fast and good enough to warn users before OCR.
    """
    edges = gray_image.filter(ImageFilter.FIND_EDGES)
    stat = ImageStat.Stat(edges)
    return float(stat.var[0]) if stat.var else 0.0


def inspect_document_image_quality(image_path: Path) -> dict[str, Any]:
    """Return lightweight image quality metrics and warnings before OCR."""
    min_short_side = env_int("IMAGE_QUALITY_MIN_SHORT_SIDE", DEFAULT_IMAGE_QUALITY_MIN_SHORT_SIDE)
    min_contrast = env_float("IMAGE_QUALITY_MIN_CONTRAST", DEFAULT_IMAGE_QUALITY_MIN_CONTRAST)
    min_sharpness = env_float("IMAGE_QUALITY_MIN_SHARPNESS", DEFAULT_IMAGE_QUALITY_MIN_SHARPNESS)
    min_brightness = env_float("IMAGE_QUALITY_MIN_BRIGHTNESS", DEFAULT_IMAGE_QUALITY_MIN_BRIGHTNESS)
    max_brightness = env_float("IMAGE_QUALITY_MAX_BRIGHTNESS", DEFAULT_IMAGE_QUALITY_MAX_BRIGHTNESS)

    with Image.open(image_path) as image:
        oriented = orient_document_image(image)
        gray = ImageOps.grayscale(oriented)
        brightness, contrast = image_stat_mean_and_std(gray)
        sharpness = image_sharpness_score(gray)
        width, height = oriented.size
        short_side = min(width, height)

    warnings: list[str] = []
    if short_side < min_short_side:
        warnings.append(f"low resolution: short side {short_side}px; target at least {min_short_side}px")
    if brightness < min_brightness:
        warnings.append(f"too dark: brightness {brightness:.0f}; target {min_brightness:.0f}-{max_brightness:.0f}")
    elif brightness > max_brightness:
        warnings.append(f"too bright: brightness {brightness:.0f}; target {min_brightness:.0f}-{max_brightness:.0f}")
    if contrast < min_contrast:
        warnings.append(f"low contrast: {contrast:.0f}; target at least {min_contrast:.0f}")
    if sharpness < min_sharpness:
        warnings.append(f"possibly blurry: sharpness {sharpness:.0f}; target at least {min_sharpness:.0f}")

    return {
        "width": width,
        "height": height,
        "brightness": round(brightness, 1),
        "contrast": round(contrast, 1),
        "sharpness": round(sharpness, 1),
        "warnings": warnings,
        "status": "review" if warnings else "pass",
    }


def format_image_quality_warning(report: dict[str, Any]) -> str:
    warnings = report.get("warnings") or []
    if not warnings:
        return ""
    bullet_lines = "\n".join(f"- {warning}" for warning in warnings)
    return (
        "Image quality warning. I will still try OCR, but extraction may be less accurate.\n"
        f"{bullet_lines}\n\n"
        "If the review looks wrong, retake the photo with the full page visible, camera parallel to the paper, and stronger lighting."
    )


def orient_document_image(image: Image.Image) -> Image.Image:
    image = ImageOps.exif_transpose(image)
    if image.width > image.height:
        image = image.rotate(90, expand=True)
    return image


def threshold_for_text(gray_image: Image.Image) -> Image.Image:
    """Simple adaptive-like threshold using image brightness.

    This avoids a hard OpenCV dependency while still helping faint invoice text.
    """
    brightness, _ = image_stat_mean_and_std(gray_image)
    threshold = int(max(145, min(205, brightness * 0.92)))
    return gray_image.point(lambda pixel: 255 if pixel > threshold else 0, mode="1").convert("L")


def enhance_ocr_crop(crop: Image.Image, scale: int = 2, contrast: float = 1.8, mode: str = "standard") -> Image.Image:
    crop = ImageOps.grayscale(crop)
    crop = ImageOps.autocontrast(crop)
    target_size = (max(1, crop.width * scale), max(1, crop.height * scale))

    if mode == "mild":
        crop = ImageEnhance.Contrast(crop).enhance(max(1.15, contrast * 0.75))
        crop = crop.resize(target_size, Image.Resampling.LANCZOS)
        return crop.filter(ImageFilter.UnsharpMask(radius=1.2, percent=120, threshold=3))

    if mode == "strong":
        crop = ImageEnhance.Contrast(crop).enhance(max(2.1, contrast * 1.25))
        crop = crop.resize(target_size, Image.Resampling.LANCZOS)
        crop = crop.filter(ImageFilter.UnsharpMask(radius=1.5, percent=180, threshold=2))
        return crop

    if mode == "binary":
        crop = ImageEnhance.Contrast(crop).enhance(max(1.8, contrast))
        crop = crop.resize(target_size, Image.Resampling.LANCZOS)
        crop = crop.filter(ImageFilter.MedianFilter(size=3))
        return threshold_for_text(crop)

    # Standard keeps your original behavior, but uses UnsharpMask for clearer text strokes.
    crop = ImageEnhance.Contrast(crop).enhance(contrast)
    crop = crop.resize(target_size, Image.Resampling.LANCZOS)
    return crop.filter(ImageFilter.UnsharpMask(radius=1.3, percent=150, threshold=3))


def ocr_variant_paths_for(image_path: Path) -> list[Path]:
    variants = [image_path]
    for suffix in ["mild", "strong", "binary"]:
        candidate = image_path.with_name(f"{image_path.stem}__{suffix}{image_path.suffix}")
        if candidate.exists():
            variants.append(candidate)
    return variants


def save_enhanced_image_variants(crop: Image.Image, output_path: Path, scale: int, contrast: float) -> None:
    modes = ["standard"]
    if env_bool("OCR_ENHANCEMENT_VARIANTS_ENABLED", True):
        modes.extend(["mild", "strong", "binary"])

    for mode in modes:
        enhanced = enhance_ocr_crop(crop, scale=scale, contrast=contrast, mode=mode)
        candidate_path = output_path if mode == "standard" else output_path.with_name(f"{output_path.stem}__{mode}{output_path.suffix}")
        enhanced.save(candidate_path, quality=95)


def save_relative_crop(
    image: Image.Image,
    output_path: Path,
    box: tuple[float, float, float, float],
    scale: int = 2,
    contrast: float = 1.8,
) -> Path:
    image = orient_document_image(image)
    width, height = image.size
    crop_box = (
        int(width * box[0]),
        int(height * box[1]),
        int(width * box[2]),
        int(height * box[3]),
    )
    crop = image.crop(crop_box)
    save_enhanced_image_variants(crop, output_path, scale=scale, contrast=contrast)
    return output_path

def tuju_profile_enabled() -> bool:
    if not TUJU_PROFILE_ENABLED:
        return False
    return os.getenv("TUJU_PROFILE_ENABLED", "1").strip().lower() in {"1", "true", "yes", "on"}


def create_enhanced_table_crop(image_path: Path) -> Path:
    ENHANCED_DIR.mkdir(parents=True, exist_ok=True)
    output_path = ENHANCED_DIR / f"{image_path.stem}_table.jpg"

    with Image.open(image_path) as image:
        save_relative_crop(image, output_path, (0.03, 0.24, 0.995, 0.61), scale=2, contrast=1.7)

    return output_path


def create_tuju_focused_crops(image_path: Path) -> dict[str, Path]:
    ENHANCED_DIR.mkdir(parents=True, exist_ok=True)
    with Image.open(image_path) as image:
        crops = {
            "vendor": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_vendor.jpg",
                (0.03, 0.04, 0.58, 0.16),
                scale=2,
                contrast=1.9,
            ),
            "details": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_details.jpg",
                (0.54, 0.05, 0.94, 0.16),
                scale=2,
                contrast=1.9,
            ),
            "contact": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_contact.jpg",
                (0.07, 0.28, 0.72, 0.35),
                scale=3,
                contrast=2.0,
            ),
            "table": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_table.jpg",
                (0.04, 0.34, 0.98, 0.50),
                scale=3,
                contrast=2.0,
            ),
            "total": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_total.jpg",
                (0.02, 0.58, 0.995, 0.72),
                scale=2,
                contrast=1.8,
            ),
        }
    return crops


def create_tuju_classifier_crop(image_path: Path) -> Path:
    ENHANCED_DIR.mkdir(parents=True, exist_ok=True)
    output_path = ENHANCED_DIR / f"{image_path.stem}_tuju_classifier.jpg"

    with Image.open(image_path) as image:
        save_relative_crop(image, output_path, (0.02, 0.00, 0.98, 0.30), scale=2, contrast=2.0)

    return output_path


def create_tuju_invoice_focused_crops(image_path: Path) -> dict[str, Path]:
    ENHANCED_DIR.mkdir(parents=True, exist_ok=True)
    with Image.open(image_path) as image:
        crops = {
            "header": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_invoice_header.jpg",
                (0.02, 0.04, 0.98, 0.27),
                scale=2,
                contrast=2.0,
            ),
            "contact": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_invoice_contact.jpg",
                (0.02, 0.25, 0.98, 0.31),
                scale=3,
                contrast=2.0,
            ),
            "table": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_invoice_table.jpg",
                (0.03, 0.32, 0.98, 0.52),
                scale=3,
                contrast=2.0,
            ),
            "total": save_relative_crop(
                image,
                ENHANCED_DIR / f"{image_path.stem}_tuju_invoice_total.jpg",
                (0.02, 0.54, 0.99, 0.72),
                scale=2,
                contrast=1.8,
            ),
        }
    return crops


def create_ocr_ready_image(image_path: Path) -> Path:
    OCR_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OCR_DIR / f"{image_path.stem}_ocr.jpg"

    with Image.open(image_path) as image:
        image = orient_document_image(image)
        save_enhanced_image_variants(image, output_path, scale=2, contrast=1.8)

    return output_path


def import_pytesseract() -> Any:
    try:
        import pytesseract
    except ImportError as exc:
        raise LocalOCRUnavailable("pytesseract is not installed.") from exc

    configured_cmd = os.getenv("TESSERACT_CMD")
    if configured_cmd:
        pytesseract.pytesseract.tesseract_cmd = configured_cmd
    return pytesseract


def parse_ocr_confidence(value: Any) -> float | None:
    try:
        confidence = float(value)
    except (TypeError, ValueError):
        return None
    if confidence < 0:
        return None
    return confidence


def ocr_config_for_zone(zone: str) -> str:
    env_name = "TESSERACT_CONFIG_" + re.sub(r"[^A-Z0-9]+", "_", zone.upper()).strip("_")
    default_config = os.getenv("TESSERACT_CONFIG", "--oem 3 --psm 6")
    return os.getenv(env_name, default_config)


def ocr_selection_score(text: str, confidence: float) -> float:
    normalized = re.sub(r"\s+", "", text or "")
    numeric_hits = len(re.findall(r"\d", text or ""))
    keyword_hits = len(re.findall(r"\b(?:tuju|invoice|delivery|order|qty|quantity|amount|total|contact)\b", text or "", flags=re.IGNORECASE))
    return confidence + min(len(normalized) / 8, 25) + min(numeric_hits / 3, 10) + min(keyword_hits * 2, 12)


def ocr_single_text_and_confidence(image_path: Path, config: str | None = None) -> tuple[str, float]:
    pytesseract = import_pytesseract()
    config = config or os.getenv("TESSERACT_CONFIG", "--oem 3 --psm 6")
    try:
        data = pytesseract.image_to_data(
            str(image_path),
            lang=os.getenv("TESSERACT_LANG", "eng"),
            config=config,
            output_type=pytesseract.Output.DICT,
        )
    except Exception as exc:
        raise LocalOCRUnavailable(f"Local OCR failed: {exc}") from exc

    lines_by_key: dict[tuple[int, int, int], list[tuple[int, str]]] = {}
    confidences: list[float] = []
    for index, text in enumerate(data.get("text", [])):
        word = str(text).strip()
        if not word:
            continue
        confidence = parse_ocr_confidence(data.get("conf", [])[index])
        if confidence is not None:
            confidences.append(confidence)
        key = (
            int(data.get("block_num", [0])[index]),
            int(data.get("par_num", [0])[index]),
            int(data.get("line_num", [0])[index]),
        )
        left = int(data.get("left", [0])[index])
        lines_by_key.setdefault(key, []).append((left, word))

    lines = []
    for key in sorted(lines_by_key):
        words = [word for _, word in sorted(lines_by_key[key], key=lambda item: item[0])]
        lines.append(" ".join(words))

    average_confidence = sum(confidences) / len(confidences) if confidences else 0.0
    return "\n".join(lines), average_confidence


def ocr_text_and_confidence(image_path: Path, config: str | None = None) -> tuple[str, float]:
    config = config or os.getenv("TESSERACT_CONFIG", "--oem 3 --psm 6")
    candidate_paths = ocr_variant_paths_for(image_path) if env_bool("OCR_VARIANT_SELECTION_ENABLED", True) else [image_path]
    best_text = ""
    best_confidence = 0.0
    best_score = -1.0
    best_path = image_path

    for candidate_path in candidate_paths:
        text, confidence = ocr_single_text_and_confidence(candidate_path, config=config)
        score = ocr_selection_score(text, confidence)
        logging.info(
            "OCR variant candidate path=%s confidence=%.1f chars=%s score=%.1f",
            candidate_path.name,
            confidence,
            len(text),
            score,
        )
        if score > best_score:
            best_text = text
            best_confidence = confidence
            best_score = score
            best_path = candidate_path

    if best_path != image_path:
        logging.info(
            "OCR variant selected base=%s selected=%s confidence=%.1f score=%.1f",
            image_path.name,
            best_path.name,
            best_confidence,
            best_score,
        )
    return best_text, best_confidence


def cached_ocr_text_and_confidence(
    cache: dict[tuple[str, str], tuple[str, float]],
    image_path: Path,
    zone: str,
) -> tuple[str, float]:
    config = ocr_config_for_zone(zone)
    cache_key = (str(image_path.resolve()), config)
    cached = cache.get(cache_key)
    if cached is not None:
        logging.info("Local OCR cache hit for zone=%s path=%s", zone, image_path.name)
        return cached

    started_at = time.perf_counter()
    result = ocr_text_and_confidence(image_path, config=config)
    elapsed = time.perf_counter() - started_at
    text, confidence = result
    logging.info(
        "Local OCR zone=%s path=%s elapsed=%.2fs confidence=%.1f chars=%s",
        zone,
        image_path.name,
        elapsed,
        confidence,
        len(text),
    )
    cache[cache_key] = result
    return result


def text_looks_like_tuju_invoice(text: str) -> bool:
    normalized = normalize_text(text)
    compact = re.sub(r"[^a-z0-9]+", "", normalized.lower())
    return (
        ("tuju" in compact and ("galaksi" in compact or "galaks" in compact))
        or "tujugalaksi" in compact
        or bool(re.search(r"\bTG[-\s]?[KR]\d{4,}\b", text, flags=re.IGNORECASE))
        or bool(re.search(r"tgk[0-9o]{4,}", compact, flags=re.IGNORECASE))
    )


def classify_document(text: str) -> tuple[str | None, str, int, int]:
    """Classify OCR text against all loaded document profiles.

    Returns:
        (profile_id, status, score, runner_up_score)
        status: "matched" | "ambiguous" | "below_threshold"
    """
    profiles = get_profiles()
    return document_profiles.classify_best_profile(profiles, text)


def detect_tuju_invoice_from_crops(crops: dict[str, Path], cache: dict[tuple[str, str], tuple[str, float]] | None = None) -> tuple[bool, str, float]:
    probe_texts: list[str] = []
    confidences: list[float] = []
    for name in ("vendor", "details"):
        path = crops.get(name)
        if not path:
            continue
        if cache is not None:
            text, confidence = cached_ocr_text_and_confidence(cache, path, f"tuju_probe_{name}")
        else:
            text, confidence = ocr_text_and_confidence(path)
        probe_texts.append(text)
        confidences.append(confidence)
    probe_text = "\n".join(probe_texts)
    confidence = sum(confidences) / len(confidences) if confidences else 0.0
    return text_looks_like_tuju_invoice(probe_text), probe_text, confidence


def focused_ai_image_paths(image_path: Path) -> tuple[list[Path], str]:
    return [image_path], "Full document image."


def unique_existing_paths(paths: list[Path]) -> list[Path]:
    seen: set[str] = set()
    unique: list[Path] = []
    for path in paths:
        try:
            resolved = str(path.resolve())
        except OSError:
            continue
        if resolved in seen or not path.exists():
            continue
        seen.add(resolved)
        unique.append(path)
    return unique


def ai_primary_image_paths(image_path: Path) -> tuple[list[Path], str]:
    return [image_path], "Full document image."


def parse_ocr_date(text: str) -> str | None:
    for date_match in re.finditer(r"\b(\d{1,3})[.,/;:-]+(\d{1,2})[.,/;:-]+(\d{2,4})\b", text):
        first, second, year = date_match.groups()
        day = int(first)
        if day > 31 and day % 100 <= 31:
            day = day % 100
        month = int(second)
        year_int = int(year)
        if year_int < 100:
            year_int += 2000
        try:
            return datetime(year_int, month, day).date().isoformat()
        except ValueError:
            continue
    return None


def normalize_ai_date(value: Any, notes: Any = None) -> tuple[str | None, str | None]:
    text = str(value or "").strip()
    note_text = str(notes or "")
    visible_date_match = re.search(r"\b(\d{1,2})[.](\d{1,2})[.](\d{4})\b", note_text)
    if visible_date_match:
        day, month, year = (int(part) for part in visible_date_match.groups())
        try:
            return datetime(year, month, day).date().isoformat(), None
        except ValueError:
            pass
    if not text:
        return None, None
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", text):
        try:
            return datetime.fromisoformat(text).date().isoformat(), None
        except ValueError:
            return None, f"Rejected invalid AI date: {text}"
    parsed = parse_ocr_date(text)
    if parsed:
        return parsed, None
    return None, f"Rejected unclear AI date: {text}"


def normalize_ai_document_number(value: Any) -> tuple[str | None, str | None]:
    text = str(value or "").strip()
    if not text:
        return None, None
    if re.search(r"\b(?:tel|fax|h/?p|phone|mobile)\b", text, flags=re.IGNORECASE):
        return None, f"Rejected phone/contact value as document number: {text}"
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 8 and not re.search(r"\b[T1I]G[-/ ]?[KXR]", text, flags=re.IGNORECASE):
        return None, f"Rejected numeric contact-like value as document number: {text}"
    parsed = parse_ocr_tax_invoice(text)
    normalized = normalize_visible_document_number(parsed or text)
    if is_plausible_document_number(normalized):
        return normalized, None
    return None, f"Rejected invalid AI document number: {text}"


def validate_ai_extraction(data: dict[str, Any]) -> dict[str, Any]:
    warnings = list(data.get("validation_warnings") or [])
    normalized_number, number_warning = normalize_ai_document_number(data.get("tax_invoice") or data.get("invoice_number"))
    data["tax_invoice"] = normalized_number
    data.pop("invoice_number", None)
    if number_warning:
        warnings.append(number_warning)

    normalized_date, date_warning = normalize_ai_date(data.get("invoice_date"), data.get("notes"))
    data["invoice_date"] = normalized_date
    if date_warning:
        warnings.append(date_warning)

    if warnings:
        data["validation_warnings"] = warnings
        notes = str(data.get("notes") or "").strip()
        warning_text = "Validation warnings: " + "; ".join(warnings)
        data["notes"] = f"{notes} {warning_text}".strip() if notes else warning_text
    return data


def parse_ocr_tax_invoice(text: str) -> str | None:
    document_number = r"([A-Z0-9]{1,5}[-/ ]?[A-Z0-9]{4,})"
    labeled_patterns = [
        rf"tax[^\S\r\n]*invoice[^\S\r\n]*(?:no\.?|number|#)?[^\S\r\n]*[:>.,-]?[^\S\r\n]*{document_number}",
        rf"invoice[^\S\r\n]*(?:no\.?|number|#)?[^\S\r\n]*[:>.,-]?[^\S\r\n]*{document_number}",
        rf"delivery[^\S\r\n]*order[^\S\r\n]*(?:no\.?|number|#)?[^\S\r\n]*[:>.,-]?[^\S\r\n]*{document_number}",
        rf"\bD[./\s-]*O[^\S\r\n]*(?:no\.?|number|#)?[^\S\r\n]*[:>.,-]?[^\S\r\n]*{document_number}",
    ]
    for line in text.splitlines():
        for pattern in labeled_patterns:
            match = re.search(pattern, line, flags=re.IGNORECASE)
            if match:
                number = normalize_visible_document_number(match.group(1))
                if is_plausible_document_number(number):
                    return number

    fallback_match = re.search(r"\b([T1I]G[-/ ]?[KXR][A-Z0-9]{4,})\b", text, flags=re.IGNORECASE)
    if fallback_match:
        return normalize_visible_document_number(fallback_match.group(1))
    return None


def parse_ocr_contact_person(text: str) -> str | None:
    lines = [" ".join(line.split()) for line in text.splitlines() if line.strip()]
    compact_text = "\n".join(lines)
    phone_pattern = r"(?:\+?6?0)?1\d[-\s]?\d{3,4}[-\s]?\d{4}"

    def clean_contact(value: str) -> str | None:
        value = re.sub(r"\b(?:tel|fax|h/?p|hp|attn|contact person)\b\s*:?", "", value, flags=re.IGNORECASE)
        value = re.sub(r"[^A-Za-z0-9+()'/@.,& -]+", " ", value)
        value = " ".join(value.split(" -:"))
        value = re.sub(r"\s+", " ", value).strip(" :-,")
        value = re.sub(r"^(?:[A-Za-z]\s*[:.-]?\s+)+(?=[A-Z][a-z])", "", value).strip(" :-,")
        return value or None

    for index, line in enumerate(lines):
        lowered = line.lower()
        if "contact person" not in lowered:
            continue
        candidate = re.sub(r"^.*?contact\s*person\s*:?", "", line, flags=re.IGNORECASE).strip()
        if not re.search(phone_pattern, candidate) and index + 1 < len(lines):
            candidate = f"{candidate} {lines[index + 1]}".strip()
        cleaned = clean_contact(candidate)
        if cleaned:
            return cleaned

    attn_match = re.search(r"\bAttn\s*:?\s*([A-Za-z][A-Za-z .'/@-]{1,60})", compact_text, flags=re.IGNORECASE)
    hp_match = re.search(r"\bH\s*/?\s*P\s*:?\s*(" + phone_pattern + r")", compact_text, flags=re.IGNORECASE)
    if attn_match or hp_match:
        parts = []
        if attn_match:
            parts.append(attn_match.group(1).strip(" :-,"))
        if hp_match:
            parts.append(hp_match.group(1).replace(" ", ""))
        cleaned = clean_contact(" ".join(parts))
        if cleaned:
            return cleaned

    phone_match = re.search(phone_pattern, compact_text)
    if phone_match:
        return clean_contact(phone_match.group(0))
    return None


def clean_ocr_description(value: str) -> str:
    description = " ".join(value.split())
    replacements = {
        "RAS": "RHS",
        "MSSHS": "MS SHS",
        "MSPlate": "MS Plate",
        "Plate4": "Plate 4",
        "RHS:": "RHS",
        "SHS50mm": "SHS 50mm",
        "'6m": "6m",
        "’x": "' x",
        "''": "'",
        "x:": "x ",
        ":x": " x",
        "3:0mm": "3.0mm",
        "3:0 mm": "3.0mm",
        "2.3mmx": "2.3mm x",
        "0mmx": "0mm x",
    }
    for old, new in replacements.items():
        description = description.replace(old, new)
    return " ".join(description.split())


def has_suspicious_ocr_description(line_items: list[dict[str, Any]]) -> bool:
    suspicious_markers = ["{", "}", "|", "’", "“", "”", "‘", "€", "¢", "?", "oe"]
    for item in line_items:
        description = str(item.get("description") or "")
        if any(marker in description for marker in suspicious_markers):
            return True
        if description.count(":") > 0:
            return True
    return False


def clean_ocr_table_line(value: str) -> str:
    line = " ".join(value.split())
    replacements = {
        "â€˜": "'",
        "â€™": "'",
        "â€œ": '"',
        "â€": '"',
        "‘": "'",
        "’": "'",
        "“": '"',
        "”": '"',
        "stee!": "steel",
        "PVCplug": "PVC plug",
        "PVCplugÃ©": "PVC plug ",
        "PVCplugé": "PVC plug ",
        "plug6mm": "plug 6mm",
        "plugé6mm": "plug 6mm",
        "plug:6mm": "plug 6mm",
        "6mmx": "6mm x",
        "x40mm": "x 40mm",
        "x.40mm": "x 40mm",
        "x:40mm": "x 40mm",
        ";x": " x",
        "10misteel": "10m steel",
        "10mssteel": "10m steel",
        "x.10m": "x 10m",
        "x:10m": "x 10m",
        "toll": "roll",
        "Toll": "roll",
    }
    for old, new in replacements.items():
        line = line.replace(old, new)
    line = re.sub(r"(?<=\d)\\(?=[A-Za-z])", " ", line)
    line = re.sub(r"(?<=\s)['`]+(?=[A-Za-z])", "", line)
    line = re.sub(r"(?<=\d)[.:]+(?=\s+[A-Za-z])", "", line)
    line = re.sub(r"(\d+[.:]\d{1,2})[.;:]+(?=\s|$)", r"\1", line)
    line = re.sub(r"(?<=\d)[);,]+(?=\s|$)", "", line)
    line = re.sub(r"(?<=[A-Za-z])[.:]+(?=\s|$)", "", line)
    line = re.sub(r"\s+", " ", line).strip()
    return line


def parse_ocr_line_items(text: str) -> list[dict[str, Any]]:
    line_items: list[dict[str, Any]] = []
    money = r"[\d,]+(?:[.:]\d{1,2})?"
    quantity_unit = r"[A-Za-z][A-Za-z0-9./-]{0,19}"
    priced_row_pattern = re.compile(
        rf"^\s*(\d{{1,3}})\s*(?:[.)=-]\s*)?(.+?)\s+({money})\s*(?:[\\/|]\s*)?({quantity_unit})?\s+({money})\s+({money})\s*$",
        flags=re.IGNORECASE,
    )
    delivery_order_row_pattern = re.compile(
        rf"^\s*(\d{{1,3}})\s*(?:[.)=-]\s*)?(.+?)\s+({money})\s*(?:[\\/|]\s*)?({quantity_unit})\s*$",
        flags=re.IGNORECASE,
    )
    skip_words = {
        "delivery order",
        "description",
        "quantity",
        "unit price",
        "amount",
        "total",
        "subtotal",
    }

    candidate_lines = ocr_product_table_lines(text) or [" ".join(raw_line.split()) for raw_line in text.splitlines()]
    for line in candidate_lines:
        if not line:
            continue
        line = clean_ocr_table_line(line)
        line = re.sub(r"(?<=\s)\D(?=\d+[.:]\d{1,2}\s+\d)", "", line)
        lowered = line.lower()
        if any(word in lowered for word in skip_words):
            continue

        match = priced_row_pattern.match(line)
        if match:
            item_no, description, quantity, unit, unit_price, line_total = match.groups()
        else:
            match = delivery_order_row_pattern.match(line)
            if not match:
                continue
            item_no, description, quantity, unit = match.groups()
            unit_price = None
            line_total = None

        unit = normalize_quantity_unit(unit)
        if not unit:
            continue

        description = clean_ocr_description(description.strip(" -:|"))
        if not description:
            continue
        line_items.append(
            {
                "item_no": item_no,
                "description": description,
                "quantity": normalize_number(quantity.replace(":", ".")),
                "quantity_unit": unit,
                "unit_price": normalize_number(unit_price.replace(":", ".")) if unit_price else None,
                "line_total": normalize_number(line_total.replace(":", ".")) if line_total else None,
            }
        )

    deduped: dict[tuple[str, float | None, str, float | None, float | None], dict[str, Any]] = {}
    for item in line_items:
        key = (
            normalize_item_no(item.get("item_no")),
            normalize_number(item.get("quantity")),
            normalize_quantity_unit(item.get("quantity_unit")).lower(),
            normalize_number(item.get("unit_price")),
            normalize_number(item.get("line_total")),
        )
        existing = deduped.get(key)
        if not existing or len(str(item.get("description") or "")) >= len(str(existing.get("description") or "")):
            deduped[key] = item

    line_items = list(deduped.values())
    return line_items


def merge_ocr_line_item_candidates(*candidate_groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    ordered: list[dict[str, Any]] = []

    def item_score(item: dict[str, Any]) -> int:
        score = 0
        for key in ("description", "quantity", "quantity_unit", "unit_price", "line_total"):
            value = item.get(key)
            if value not in (None, ""):
                score += 1
        score += min(len(str(item.get("description") or "")), 80)
        return score

    for group in candidate_groups:
        for item in group:
            item_no = normalize_item_no(item.get("item_no"))
            if not item_no:
                ordered.append(item)
                continue
            existing = merged.get(item_no)
            if not existing or item_score(item) > item_score(existing):
                merged[item_no] = item

    for item_no in sorted(merged, key=lambda value: int(value) if value.isdigit() else 999):
        ordered.append(merged[item_no])
    return ordered


def parse_ocr_document_total(text: str) -> float | None:
    compact_lines = [" ".join(clean_ocr_table_line(line).split()) for line in text.splitlines() if line.strip()]
    patterns = [
        r"(?<![A-Za-z0-9])(?:net\s*)?total\s*payable\b.*?\bRM(?![A-Za-z0-9])[\W_]{0,20}([\d,]+(?:[.:]\d{2})?)",
        r"(?<![A-Za-z0-9])net\s*total\b.*?\bRM(?![A-Za-z0-9])[\W_]{0,20}([\d,]+(?:[.:]\d{2})?)",
        r"(?<![A-Za-z0-9])gross\b.*?\bRM(?![A-Za-z0-9])[\W_]{0,20}([\d,]+(?:[.:]\d{2})?)",
    ]
    for line in compact_lines:
        lowered = line.lower()
        if not any(marker in lowered for marker in ("total", "gross")):
            continue
        for pattern in patterns:
            match = re.search(pattern, line, flags=re.IGNORECASE)
            if match:
                return normalize_number(match.group(1).replace(":", "."))
    return None


def extract_invoice_with_local_ocr(image_path: Path) -> LocalOCRResult:
    extraction_started_at = time.perf_counter()
    ocr_cache: dict[tuple[str, str], tuple[str, float]] = {}
    tuju_used = False
    focused_profile = ""
    matched_profile_id: str | None = None

    # ── Profile-based classification ──────────────────────────────────
    profiles = get_profiles()
    if profiles:
        # Build OCR text from classifier crop
        classifier_path = create_tuju_classifier_crop(image_path)
        classifier_text, classifier_confidence = cached_ocr_text_and_confidence(ocr_cache, classifier_path, "classifier_probe")
        probe_full_page_path = create_ocr_ready_image(image_path)
        probe_full_text, probe_full_confidence = cached_ocr_text_and_confidence(ocr_cache, probe_full_page_path, "classifier_full_page_probe")
        probe_text = f"{classifier_text}\n{probe_full_text}"

        profile_id, status, score, runner_up = classify_document(probe_text)
        logging.info(
            "Profile classification image=%s profile_id=%s status=%s score=%d runner_up=%d",
            image_path.name, profile_id, status, score, runner_up,
        )

        if status == "matched" and profile_id:
            matched_profile_id = profile_id
            logging.info("Document classified as profile=%s score=%d", profile_id, score)
        elif status == "ambiguous":
            logging.info(
                "Ambiguous classification for image=%s (best=%s score=%d runner_up=%d); "
                "falling through to generic local OCR",
                image_path.name, profile_id, score, runner_up,
            )
        else:
            if text_looks_like_tuju_invoice(probe_text):
                logging.info(
                    "TUJU-looking document fell back to generic local OCR image=%s classifier_confidence=%.1f probe_confidence=%.1f",
                    image_path.name, classifier_confidence, probe_full_confidence,
                )
            else:
                logging.info(
                    "Unknown document format for image=%s classifier_confidence=%.1f",
                    image_path.name, classifier_confidence,
                )
                raise UnknownDocumentFormat(UNKNOWN_DOCUMENT_FORMAT_MESSAGE, profile_id)

    else:
        # No profiles loaded — use legacy TUJU detection as fallback
        if tuju_profile_enabled():
            logging.info("No profiles loaded; using legacy TUJU detection")
            classifier_path = create_tuju_classifier_crop(image_path)
            classifier_text, classifier_confidence = cached_ocr_text_and_confidence(ocr_cache, classifier_path, "tuju_classifier")

    # ── Profile-based extraction ──────────────────────────────────────
    matched_profile: document_profiles.DocumentProfile | None = None
    if matched_profile_id:
        matched_profile = profile_by_id(matched_profile_id)

    # Legacy TUJU path (always used for TUJU profiles for backward compatibility)
    use_legacy_tuju = legacy_tuju_pipeline_enabled() or matched_profile_id in ("tuju_galaxy_delivery_order", "tuju_galaxy_invoice")

    if use_legacy_tuju and matched_profile_id == "tuju_galaxy_delivery_order":
        tuju_crops = create_tuju_focused_crops(image_path)
        contact_text, contact_confidence = cached_ocr_text_and_confidence(ocr_cache, tuju_crops["contact"], "tuju_do_contact")
        table_text, table_confidence = cached_ocr_text_and_confidence(ocr_cache, tuju_crops["table"], "tuju_do_table")
        full_page_path = create_ocr_ready_image(image_path)
        full_text, full_confidence = cached_ocr_text_and_confidence(ocr_cache, full_page_path, "tuju_do_full_page")
        text = f"{classifier_text}\n{contact_text}\n{table_text}\n{full_text}"
        average_confidence = (classifier_confidence + contact_confidence + table_confidence + full_confidence) / 4
        tuju_used = True
        focused_profile = "tuju_delivery_order_focused"

    elif use_legacy_tuju and matched_profile_id == "tuju_galaxy_invoice":
        invoice_crops = create_tuju_invoice_focused_crops(image_path)
        contact_text, contact_confidence = cached_ocr_text_and_confidence(ocr_cache, invoice_crops["contact"], "tuju_invoice_contact")
        table_text, table_confidence = cached_ocr_text_and_confidence(ocr_cache, invoice_crops["table"], "tuju_invoice_table")
        total_text, total_confidence = cached_ocr_text_and_confidence(ocr_cache, invoice_crops["total"], "tuju_invoice_total")
        text = f"{classifier_text}\n{contact_text}\n{table_text}\n{total_text}"
        average_confidence = (classifier_confidence + contact_confidence + table_confidence + total_confidence) / 4
        tuju_used = True
        focused_profile = "tuju_invoice_focused"

    elif matched_profile:
        # Profile-driven extraction
        profile_crops = create_profile_crops(image_path, matched_profile)
        if profile_crops:
            crop_texts = []
            crop_confidences = []
            for field_name, crop_path in profile_crops.items():
                ct, cc = cached_ocr_text_and_confidence(ocr_cache, crop_path, f"profile_{matched_profile.id}_{field_name}")
                crop_texts.append(ct)
                crop_confidences.append(cc)
            full_page_path = create_ocr_ready_image(image_path)
            full_text, full_confidence = cached_ocr_text_and_confidence(ocr_cache, full_page_path, "profile_full_page")
            text = "\n".join([classifier_text] + crop_texts + [full_text])
            all_confidences = [classifier_confidence] + crop_confidences + [full_confidence]
            average_confidence = sum(all_confidences) / len(all_confidences)
        else:
            # No crop hints — use full-page OCR only
            full_page_path = create_ocr_ready_image(image_path)
            full_text, full_confidence = cached_ocr_text_and_confidence(ocr_cache, full_page_path, "profile_full_page")
            text = f"{classifier_text}\n{full_text}"
            average_confidence = (classifier_confidence + full_confidence) / 2

        tuju_used = False
        focused_profile = ""

    if not tuju_used:
        ocr_image_path = create_ocr_ready_image(image_path)
        table_crop_path = create_enhanced_table_crop(image_path)
        full_text, full_confidence = cached_ocr_text_and_confidence(ocr_cache, ocr_image_path, "local_full_page")
        table_text, table_confidence = cached_ocr_text_and_confidence(ocr_cache, table_crop_path, "local_table")
        text = f"{full_text}\n{table_text}"
        average_confidence = (full_confidence + table_confidence) / 2

    tax_invoice = parse_ocr_tax_invoice(text)
    invoice_date = parse_ocr_date(text)
    contact_person = parse_ocr_contact_person(text)
    line_items = merge_ocr_line_item_candidates(parse_ocr_line_items(table_text), parse_ocr_line_items(text))
    document_total = parse_ocr_document_total(text)

    # ── Apply profile normalizers ─────────────────────────────────────
    validation_warnings: list[str] = []
    if matched_profile:
        raw_data: dict[str, Any] = {
            "tax_invoice": tax_invoice,
            "invoice_date": invoice_date,
            "contact_person": contact_person,
            "document_total": document_total,
            "line_items": line_items,
        }
        normalized = apply_profile_normalizers(raw_data, matched_profile)
        tax_invoice = normalized.get("tax_invoice", tax_invoice)
        invoice_date = normalized.get("invoice_date", invoice_date)
        contact_person = normalized.get("contact_person", contact_person)
        document_total = normalized.get("document_total", document_total)
        line_items = normalized.get("line_items", line_items)

        # ── Run validation rules ──────────────────────────────────────
        validation_warnings = run_validation_rules(raw_data, matched_profile)

    min_confidence = env_float("LOCAL_OCR_MIN_CONFIDENCE", DEFAULT_LOCAL_OCR_MIN_CONFIDENCE)
    review_confidence = env_float("LOCAL_OCR_REVIEW_CONFIDENCE", DEFAULT_LOCAL_OCR_REVIEW_CONFIDENCE)
    min_items = env_int("LOCAL_OCR_MIN_ITEMS", DEFAULT_LOCAL_OCR_MIN_ITEMS)
    missing_fields = []
    if not tax_invoice:
        missing_fields.append("delivery order number")
    if not invoice_date:
        missing_fields.append("date")
    if len(line_items) < min_items:
        missing_fields.append("line items")
    if line_items and has_suspicious_ocr_description(line_items):
        missing_fields.append("clean item descriptions")

    accepted = average_confidence >= min_confidence and not missing_fields
    if accepted and average_confidence < review_confidence:
        reason = (
            f"{'TUJU focused OCR' if tuju_used else 'Local OCR'} parsed with confidence {average_confidence:.1f}; "
            "review carefully before saving."
        )
    elif accepted:
        reason = f"{'TUJU focused OCR' if tuju_used else 'Local OCR'} accepted with confidence {average_confidence:.1f}."
    elif average_confidence < min_confidence:
        reason = f"{'TUJU focused OCR' if tuju_used else 'Local OCR'} confidence {average_confidence:.1f} is below {min_confidence:.1f}."
    else:
        reason = f"{'TUJU focused OCR' if tuju_used else 'Local OCR'} missing required fields: " + ", ".join(missing_fields) + "."

    data = {
        "tax_invoice": tax_invoice,
        "invoice_date": invoice_date,
        "contact_person": contact_person,
        "confidence": round(min(average_confidence / 100, 1), 3),
        "notes": reason,
        "line_items": line_items,
        "extraction_method": "tuju_focused_ocr" if tuju_used else "local_ocr",
    }
    if focused_profile:
        data["extraction_profile"] = focused_profile
        data["document_profile"] = focused_profile.removesuffix("_focused")
    if matched_profile_id:
        data["profile_id"] = matched_profile_id
        data["profile_supplier"] = matched_profile.supplier if matched_profile else ""
        data["profile_document_type"] = matched_profile.document_type if matched_profile else ""
    if document_total is not None:
        data["document_total"] = document_total
    if validation_warnings:
        data["validation_warnings"] = validation_warnings
    repair_line_item_arithmetic(data)
    logging.info(
        "Local OCR extraction complete image=%s profile=%s elapsed=%.2fs ocr_calls=%s",
        image_path.name,
        focused_profile or "local_full_page",
        time.perf_counter() - extraction_started_at,
        len(ocr_cache),
    )
    return LocalOCRResult(data=data, average_confidence=average_confidence, accepted=accepted, reason=reason, raw_text=text)


async def extract_invoice_hybrid(image_path: Path, model: str) -> tuple[dict[str, Any], str, str | None]:
    if ai_primary_enabled():
        try:
            data = await extract_invoice(image_path, model, primary=True)
            data["extraction_method"] = "ai_primary"
            if os.getenv("LOCAL_OCR_VERIFICATION", "0").strip().lower() in {"1", "true", "yes", "on"}:
                try:
                    local_result = await asyncio.to_thread(extract_invoice_with_local_ocr, image_path)
                    data["local_ocr_verification"] = {
                        "accepted": local_result.accepted,
                        "confidence": round(local_result.average_confidence, 1),
                        "reason": local_result.reason,
                        "tax_invoice": local_result.data.get("tax_invoice"),
                        "invoice_date": local_result.data.get("invoice_date"),
                    }
                except Exception as exc:
                    data["local_ocr_verification"] = {"accepted": False, "reason": str(exc)}
            return data, "AI primary", "AI primary extraction is enabled."
        except Exception as exc:
            if os.getenv("LOCAL_OCR_ENABLED", "0").strip().lower() in {"1", "true", "yes", "on"}:
                logging.exception("AI primary extraction failed; trying local OCR path")
            else:
                raise

    if os.getenv("LOCAL_OCR_ENABLED", "1").strip().lower() not in {"1", "true", "yes", "on"}:
        if ai_fallback_enabled():
            data = await extract_invoice(image_path, model)
            data["extraction_method"] = "ai_fallback"
            return data, "AI fallback", "Local OCR is disabled; AI extraction was used."
        raise LocalOCRUnavailable("Local OCR is disabled and AI fallback is disabled.")

    try:
        local_result = await asyncio.to_thread(extract_invoice_with_local_ocr, image_path)
    except (LocalOCRUnavailable, UnknownDocumentFormat) as exc:
        if ai_fallback_enabled():
            data = await extract_invoice(image_path, model)
            data["extraction_method"] = "ai_fallback"
            return data, "AI fallback", f"{exc} AI extraction was used."
        logging.info("Local OCR could not extract the document and AI fallback is disabled: %s", exc)
        raise
    except Exception as exc:
        if ai_fallback_enabled():
            logging.exception("Local OCR failed unexpectedly; trying AI fallback")
            data = await extract_invoice(image_path, model)
            data["extraction_method"] = "ai_fallback"
            return data, "AI fallback", f"Local OCR failed: {exc}. AI extraction was used."
        logging.exception("Local OCR failed unexpectedly and AI fallback is disabled")
        raise RuntimeError(f"Local OCR failed: {exc}") from exc

    OCR_DIR.mkdir(parents=True, exist_ok=True)
    (OCR_DIR / f"{image_path.stem}.txt").write_text(local_result.raw_text, encoding="utf-8")

    reconciliation_needed = extraction_needs_reconciliation(local_result.data, local_result.raw_text)
    if reconciliation_needed:
        expected_count, actual_count = reconciliation_needed
        note = None
        if ai_fallback_enabled():
            try:
                data = await reconcile_invoice_extraction(image_path, model, local_result.data, local_result.raw_text, expected_count, actual_count)
                return (
                    data,
                    "AI/OCR reconciliation",
                    f"Local OCR row scan found {expected_count} visible item rows, but {actual_count} were parsed.",
                )
            except Exception as exc:
                if is_openai_auth_error(exc):
                    note = f"{local_result.reason} {AI_FALLBACK_AUTH_NOTE}"
                    logging.info(note)
                elif is_openai_credit_error(exc):
                    note = f"{local_result.reason} {AI_FALLBACK_CREDITS_NOTE}"
                    logging.info(note)
                else:
                    raise
        if note is None:
            note = (
                f"{local_result.reason} OCR row scan suggests {expected_count} visible item rows, "
                f"but {actual_count} were parsed. {AI_FALLBACK_DISABLED_NOTE}"
            )
        logging.info(note)
        local_result.data["notes"] = note
        local_result.accepted = False
        local_result.reason = note
    elif not local_result.accepted:
        note = None
        if ai_fallback_enabled():
            try:
                data = await extract_invoice(image_path, model)
                data["extraction_method"] = "ai_fallback"
                return data, "AI fallback", f"{local_result.reason} AI extraction was used."
            except Exception as exc:
                if is_openai_auth_error(exc):
                    note = f"{local_result.reason} {AI_FALLBACK_AUTH_NOTE}"
                elif is_openai_credit_error(exc):
                    note = f"{local_result.reason} {AI_FALLBACK_CREDITS_NOTE}"
                else:
                    raise
        if note is None:
            note = f"{local_result.reason} {AI_FALLBACK_DISABLED_NOTE}"
        local_result.data["notes"] = note
        local_result.reason = note

    return (
        local_result.data,
        "TUJU focused OCR" if local_result.data.get("extraction_method") == "tuju_focused_ocr" else "Local OCR",
        local_result.reason,
    )


async def is_authorized(update: Update) -> bool:
    allowed_chat_ids = configured_allowed_chat_ids()
    if not allowed_chat_ids:
        return True
    chat_id = str(update.effective_chat.id) if update.effective_chat else None
    if chat_id and chat_id in allowed_chat_ids:
        return True
    logging.warning("Rejected message from unauthorized chat_id=%s", chat_id)
    if update.message:
        await update.message.reply_text(
            f"⚠️ *This bot is not authorized for this chat.*\n\n"
            f"Your Chat ID is: `{chat_id}`\n\n"
            f"To authorize your account, add your Chat ID to `.env` on your server:\n"
            f"`TELEGRAM_ALLOWED_CHAT_IDS={chat_id}`\n\n"
            f"Or leave `TELEGRAM_ALLOWED_CHAT_IDS=` empty in `.env` to allow all chats.",
            parse_mode="Markdown",
        )
    return False


async def extract_invoice_gemini(
    image_paths: list[Path],
    prompt_text: str,
    model: str,
    system_instruction: str = SYSTEM_PROMPT,
) -> dict[str, Any]:
    try:
        from google import genai
        from google.genai import types
    except ImportError as exc:
        raise RuntimeError("google-genai SDK is not installed. Run 'pip install google-genai'.") from exc

    api_key = gemini_api_key()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY or GOOGLE_API_KEY is not set.")

    client = genai.Client(api_key=api_key)

    contents: list[Any] = []
    for candidate_path in image_paths:
        encoded_bytes = candidate_path.read_bytes()
        contents.append(types.Part.from_bytes(data=encoded_bytes, mime_type="image/jpeg"))
    contents.append(prompt_text)

    config = types.GenerateContentConfig(
        system_instruction=system_instruction,
        response_mime_type="application/json",
        temperature=0.0,
    )

    models_to_try = [model]
    for fallback in ("gemini-3.5-flash-lite", "gemini-3.7-flash", "gemini-3.1-pro-preview"):
        if fallback not in models_to_try:
            models_to_try.append(fallback)

    last_error: Exception | None = None
    for attempt_model in models_to_try:
        for attempt in range(2):
            try:
                response = await client.aio.models.generate_content(
                    model=attempt_model,
                    contents=contents,
                    config=config,
                )
                text = response.text
                if not text:
                    raise RuntimeError("Gemini API returned an empty extraction result.")
                return json.loads(text)
            except Exception as exc:
                last_error = exc
                exc_str = str(exc).lower()
                is_transient = any(
                    err in exc_str
                    for err in (
                        "503",
                        "unavailable",
                        "high demand",
                        "overloaded",
                        "429",
                        "resource_exhausted",
                        "rate_limit",
                        "quota",
                        "500",
                        "internal",
                        "404",
                        "not_found",
                    )
                )
                if is_transient:
                    logging.warning(
                        "Gemini model %s attempt %d returned transient error: %s; trying fallback...",
                        attempt_model,
                        attempt + 1,
                        exc,
                    )
                    await asyncio.sleep(1.0)
                    if "404" in exc_str or "not_found" in exc_str:
                        break  # Move immediately to the next model
                    continue
                raise

    if last_error:
        raise last_error
    raise RuntimeError("Gemini extraction failed on all attempted models.")


async def extract_invoice_openai(
    image_paths: list[Path],
    prompt_text: str,
    model: str,
    system_instruction: str = SYSTEM_PROMPT,
) -> dict[str, Any]:
    base_url = openai_base_url()
    client = AsyncOpenAI(api_key=openai_bearer_credential(), base_url=base_url)

    content: list[dict[str, Any]] = [{"type": "text", "text": prompt_text}]
    for candidate_path in image_paths:
        encoded = base64.b64encode(candidate_path.read_bytes()).decode("ascii")
        content.append(
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{encoded}",
                    "detail": "high",
                },
            }
        )

    response = await client.chat.completions.create(
        model=model,
        temperature=0,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system_instruction},
            {"role": "user", "content": content},
        ],
    )

    content_str = response.choices[0].message.content
    if not content_str:
        raise RuntimeError("OpenAI returned an empty extraction result.")
    return json.loads(content_str)


async def extract_invoice(image_path: Path, model: str | None = None, primary: bool = False) -> dict[str, Any]:
    provider = configured_ai_provider()
    image_paths, image_note = ai_primary_image_paths(image_path) if primary else focused_ai_image_paths(image_path)

    prompt_text = (
        f"{EXTRACTION_INSTRUCTIONS}\n\n"
        f"{image_note} Use the product table crop for line_items and the document "
        "details crop/full document for Delivery Order number, date, and contact person. "
        "For dates printed as DD.MM.YYYY, interpret them as day.month.year."
    )

    if provider == "gemini":
        target_model = model or gemini_model_name()
        try:
            data = await extract_invoice_gemini(image_paths, prompt_text, target_model, SYSTEM_PROMPT)
        except Exception as exc:
            if os.getenv("OPENAI_API_KEY") or os.getenv("OPENROUTER_API_KEY"):
                logging.warning("Gemini extraction failed (%s), attempting OpenAI fallback...", exc)
                openai_model = openai_model_name()
                data = await extract_invoice_openai(image_paths, prompt_text, openai_model, SYSTEM_PROMPT)
            else:
                raise
    else:
        target_model = model or openai_model_name()
        data = await extract_invoice_openai(image_paths, prompt_text, target_model, SYSTEM_PROMPT)

    data.setdefault("line_items", [])
    repair_line_item_arithmetic(data)
    validate_ai_extraction(data)
    if "TUJU focused" in image_note:
        data["extraction_profile"] = "tuju_focused"
    return data


async def reconcile_invoice_extraction(
    image_path: Path,
    model: str | None,
    first_data: dict[str, Any],
    ocr_text: str,
    expected_count: int,
    actual_count: int,
) -> dict[str, Any]:
    provider = configured_ai_provider()
    image_paths, image_note = focused_ai_image_paths(image_path)
    first_header_json = json.dumps(
        {
            "tax_invoice": first_data.get("tax_invoice") or first_data.get("invoice_number"),
            "invoice_date": first_data.get("invoice_date"),
            "contact_person": first_data.get("contact_person"),
            "confidence": first_data.get("confidence"),
            "notes": first_data.get("notes"),
        },
        ensure_ascii=False,
        indent=2,
    )
    trimmed_ocr_text = ocr_text[-8000:]

    prompt_text = (
        RECONCILIATION_INSTRUCTIONS.format(
            expected_count=expected_count,
            actual_count=actual_count,
        )
        + "\n\nFirst extraction header fields:\n"
        + first_header_json
        + "\n\nOCR text:\n"
        + trimmed_ocr_text
        + "\n\nImage set:\n"
        + image_note
    )

    if provider == "gemini":
        target_model = model or gemini_model_name()
        data = await extract_invoice_gemini(image_paths, prompt_text, target_model, SYSTEM_PROMPT)
    else:
        target_model = model or openai_model_name()
        data = await extract_invoice_openai(image_paths, prompt_text, target_model, SYSTEM_PROMPT)

    data.setdefault("line_items", [])
    repair_line_item_arithmetic(data)
    validate_ai_extraction(data)
    data["extraction_method"] = "ai_reconciled"
    if "TUJU focused" in image_note:
        data["extraction_profile"] = "tuju_focused"
    if extraction_needs_reconciliation(data, ocr_text):
        notes = data.get("notes")
        warning = f"OCR suggests {expected_count} visible item rows. Review carefully before saving."
        data["notes"] = f"{notes} {warning}".strip() if notes else warning
    return data


def save_extraction_json(invoice_id: str, data: dict[str, Any]) -> None:
    EXTRACTION_DIR.mkdir(parents=True, exist_ok=True)
    output_path = EXTRACTION_DIR / f"{invoice_id}.json"
    output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def cached_extraction_path(source_image_hash: str) -> Path:
    return EXTRACTION_CACHE_DIR / f"{source_image_hash}.json"


def load_cached_extraction(source_image_hash: str | None) -> dict[str, Any] | None:
    if not source_image_hash:
        return None
    path = cached_extraction_path(source_image_hash)
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        logging.exception("Could not load cached extraction for image hash %s", source_image_hash[:12])
        return None
    if not isinstance(data, dict):
        return None
    if data.get("extraction_method") in ("local_ocr", "tuju_focused_ocr", "legacy", "failed"):
        return None
    if not isinstance(data.get("line_items"), list):
        return None
    return data


def save_cached_extraction(source_image_hash: str | None, data: dict[str, Any]) -> None:
    if not source_image_hash:
        return
    EXTRACTION_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    payload = dict(data)
    payload["source_image_hash"] = source_image_hash
    cached_extraction_path(source_image_hash).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def extraction_review_warnings(data: dict[str, Any]) -> list[str]:
    warnings: list[str] = []
    document_type = normalize_document_type(data)
    for warning in data.get("pair_compare_warnings") or []:
        warnings.append(f"Pair check: {warning}")

    # Add profile validation warnings
    for warning in data.get("validation_warnings") or []:
        warnings.append(f"Profile validation: {warning}")

    confidence = normalize_number(data.get("confidence"))
    if confidence is not None and confidence < 0.75:
        warnings.append(f"Low OCR confidence ({confidence:.2f}); review document number, date, contact, and rows before saving.")

    method = normalize_text(data.get("extraction_method"))
    notes = normalize_text(data.get("notes"))
    if "reconciled" in method or "reconciled" in notes:
        warnings.append("AI/OCR reconciliation was used; check row count carefully.")
    if "auth" in notes or "fallback" in notes or "missing" in notes:
        warnings.append(f"Extraction note: {data.get('notes')}")

    if not data.get("tax_invoice") and not data.get("invoice_number"):
        warnings.append("Delivery Order number is missing.")
    if not data.get("invoice_date"):
        warnings.append("Delivery Order date is missing.")
    if not data.get("contact_person") and data.get("extraction_profile") == "tuju_focused":
        warnings.append("Contact person is missing.")

    for index, item in enumerate(line_items_from_data(data), start=1):
        item_no = item.get("item_no") or index
        required_item_fields = [
            ("description", item.get("description")),
            ("quantity", item.get("quantity")),
        ]
        if document_type != DOCUMENT_TYPE_DELIVERY_ORDER:
            required_item_fields.extend(
                [
                    ("unit price", item.get("unit_price")),
                    ("amount", item.get("line_total")),
                ]
            )
        missing = [label for label, value in required_item_fields if value in (None, "")]
        if missing:
            warnings.append(f"Row {item_no} is missing {', '.join(missing)}.")

        quantity = normalize_number(item.get("quantity"))
        unit_price = normalize_number(item.get("unit_price"))
        line_total = normalize_number(item.get("line_total"))
        if quantity is not None and unit_price is not None and line_total is not None:
            calculated = quantity * unit_price
            if abs(calculated - line_total) > 0.05:
                warnings.append(
                    f"Row {item_no} amount check differs: qty x unit = {calculated:.2f}, extracted amount = {line_total:.2f}."
                )

    if document_type != DOCUMENT_TYPE_DELIVERY_ORDER:
        document_total = normalize_number(data.get("document_total"))
        line_totals = [normalize_number(item.get("line_total")) for item in line_items_from_data(data)]
        summed_total = sum(total for total in line_totals if total is not None)
        if document_total is not None and line_totals and abs(summed_total - document_total) > 0.05:
            warnings.append(f"Invoice total check differs: rows sum to {summed_total:.2f}, document total is {document_total:.2f}.")

    return warnings


def pair_item_mismatch_warnings(data: dict[str, Any]) -> list[str]:
    warnings = []
    for warning in data.get("pair_compare_warnings") or []:
        warning_text = normalize_text(warning)
        if (
            "quantity differs" in warning_text
            or "item" in warning_text
            or "row" in warning_text
        ):
            warnings.append(str(warning))
    return warnings


def pair_item_source_label(data: dict[str, Any]) -> str:
    source = data.get("item_source") or DOCUMENT_TYPE_DELIVERY_ORDER
    if source == DOCUMENT_TYPE_INVOICE:
        return "Invoice"
    return "D.O"


def pair_item_source_instruction(data: dict[str, Any]) -> str:
    mismatches = pair_item_mismatch_warnings(data)
    if not mismatches:
        return ""
    return (
        "D.O and Invoice item rows are not the same.\n"
        f"Current item source: {pair_item_source_label(data)}\n"
        "Send /usedo to use D.O items, /useinvoice to use Invoice items, or /cancel to discard."
    )


def format_review_warnings(data: dict[str, Any]) -> str:
    warnings = extraction_review_warnings(data)
    if not warnings:
        return "Review warnings: none detected."
    return "Review warnings:\n" + "\n".join(f"- {warning}" for warning in warnings)


def format_line_item_value(value: Any) -> str:
    if value is None or value == "":
        return "-"
    return str(value).strip()


def format_item_review(data: dict[str, Any]) -> str:
    line_items = line_items_from_data(data)
    if not line_items:
        return "No line items were extracted."

    lines = []
    for index, item in enumerate(line_items, start=1):
        item_no = format_line_item_value(item.get("item_no") or index)
        description = format_line_item_value(item.get("description"))
        quantity = format_line_item_value(format_quantity_with_unit(item))
        unit_price = format_line_item_value(item.get("unit_price"))
        line_total = format_line_item_value(item.get("line_total"))
        if unit_price != "-" and line_total != "-":
            lines.append(f"• {description}\n  Qty: {quantity}  |  Price: RM {unit_price}  |  Total: RM {line_total}")
        else:
            lines.append(f"• {description}\n  Qty: {quantity}")

    return "\n\n".join(lines) if lines else "No valid line items were extracted."


def get_update_message(update: Update):
    if update.message:
        return update.message
    if update.callback_query and update.callback_query.message:
        return update.callback_query.message
    return None


def review_action_keyboard(is_test: bool = False) -> InlineKeyboardMarkup:
    keyboard = [
        [
            InlineKeyboardButton("💾 Save & Generate PO", callback_data="btn_save"),
        ],
        [
            InlineKeyboardButton("📋 Review Items", callback_data="btn_review"),
            InlineKeyboardButton("❌ Discard", callback_data="btn_cancel"),
        ],
    ]
    if is_test:
        keyboard.insert(1, [InlineKeyboardButton("📁 Save Official Record", callback_data="btn_saverecord")])
    return InlineKeyboardMarkup(keyboard)


async def reply_long_text(
    update: Update,
    text: str,
    reply_markup: InlineKeyboardMarkup | None = None,
) -> None:
    message = get_update_message(update)
    if not message:
        return

    max_length = 3900
    remaining = text
    while remaining:
        if len(remaining) <= max_length:
            await safe_reply_text(update, remaining, "long reply", reply_markup=reply_markup)
            return
        split_at = remaining.rfind("\n", 0, max_length)
        if split_at <= 0:
            split_at = max_length
        await safe_reply_text(update, remaining[:split_at], "long reply chunk")
        remaining = remaining[split_at:].lstrip()


async def safe_reply_text(
    update: Update,
    text: str,
    label: str,
    reply_markup: InlineKeyboardMarkup | None = None,
) -> bool:
    message = get_update_message(update)
    if not message:
        return False
    try:
        await message.reply_text(
            text,
            reply_markup=reply_markup,
            connect_timeout=15,
            read_timeout=45,
            write_timeout=45,
            pool_timeout=15,
        )
        return True
    except TimedOut:
        logging.warning("Telegram timed out while sending %s; continuing invoice flow", label)
        return False
    except TelegramError:
        logging.exception("Telegram failed while sending %s; continuing invoice flow", label)
        return False


def save_pending_review(
    context: ContextTypes.DEFAULT_TYPE,
    invoice_id: str,
    received_at: datetime,
    image_path: Path,
    data: dict[str, Any],
    submitter_chat_id: int | None = None,
    submitter_name: str | None = None,
) -> dict[str, Any]:
    data["submitter_chat_id"] = str(submitter_chat_id or "")
    data["submitter_name"] = submitter_name or ""
    data["record_type"] = invoice_record_type(submitter_chat_id)
    document_type = normalize_document_type(data)
    data["document_type"] = document_type

    pending = context.chat_data.get(PENDING_REVIEW_KEY)
    if not isinstance(pending, dict) or pending.get("mode") != "pair":
        pending = {
            "mode": "pair",
            "documents": {},
            "submitter_chat_id": str(submitter_chat_id or ""),
            "submitter_name": submitter_name or "",
            "force_record": False,
        }

    document_record = {
        "invoice_id": invoice_id,
        "received_at": received_at.isoformat(),
        "image_path": str(image_path),
        "data": data,
        "submitter_chat_id": str(submitter_chat_id or ""),
        "submitter_name": submitter_name or "",
    }
    pending.setdefault("documents", {})[document_type] = document_record
    pending["submitter_chat_id"] = str(submitter_chat_id or pending.get("submitter_chat_id") or "")
    pending["submitter_name"] = submitter_name or pending.get("submitter_name") or ""

    # Check if standalone quotation/single document workflow applies
    supplier_profile = data.get("supplier_profile") or suppliers.detect_supplier_profile(
        data, configured_default_supplier(), configured_supplier_aliases()
    )
    data["supplier_profile"] = supplier_profile
    is_standalone = (
        document_type in (DOCUMENT_TYPE_QUOTATION, DOCUMENT_TYPE_CASH_BILL)
        or supplier_profile.get("category") == "TECH"
        or "walihin" in str(supplier_profile.get("display_name", "")).lower()
    )
    if is_standalone:
        pending["data"] = data
        pending["invoice_id"] = invoice_id
        pending["received_at"] = received_at.isoformat()
        pending["image_path"] = str(image_path)
        pending["item_source_confirmed"] = True

    documents = pending.get("documents") or {}
    delivery_order_record = documents.get(DOCUMENT_TYPE_DELIVERY_ORDER)
    invoice_record = documents.get(DOCUMENT_TYPE_INVOICE)
    if delivery_order_record and invoice_record:
        item_source_choice = pending.get("item_source_choice") or DOCUMENT_TYPE_DELIVERY_ORDER
        merged_data, compare_warnings = compare_and_merge_documents(
            delivery_order_record["data"],
            invoice_record["data"],
            item_source_choice,
        )
        merged_data["submitter_chat_id"] = pending["submitter_chat_id"]
        merged_data["submitter_name"] = pending["submitter_name"]
        merged_data["record_type"] = invoice_record_type(submitter_chat_id)
        merged_data["delivery_order_image_path"] = delivery_order_record["image_path"]
        merged_data["invoice_image_path"] = invoice_record["image_path"]
        pending["data"] = merged_data
        pending["pair_compare_warnings"] = compare_warnings
        pending["item_source_choice"] = merged_data.get("item_source") or DOCUMENT_TYPE_DELIVERY_ORDER
        if pair_item_mismatch_warnings(merged_data):
            pending["item_source_confirmed"] = False
        pending["invoice_id"] = delivery_order_record["invoice_id"]
        pending["received_at"] = delivery_order_record["received_at"]
        pending["image_path"] = delivery_order_record["image_path"]

    context.chat_data[PENDING_REVIEW_KEY] = pending
    if submitter_chat_id:
        pending_store.save_pending(DATA_DIR, submitter_chat_id, pending)
    return pending


def pending_chat_id(pending: dict[str, Any]) -> str:
    return str(pending.get("submitter_chat_id") or "").strip()


def persist_pending_review(pending: dict[str, Any]) -> None:
    chat_id = pending_chat_id(pending)
    if chat_id:
        pending_store.save_pending(DATA_DIR, chat_id, pending)


def get_pending_review(context: ContextTypes.DEFAULT_TYPE, chat_id: Any = None) -> dict[str, Any] | None:
    pending = context.chat_data.get(PENDING_REVIEW_KEY)
    if isinstance(pending, dict):
        return pending
    if chat_id:
        pending = pending_store.load_pending(DATA_DIR, chat_id)
        if isinstance(pending, dict):
            context.chat_data[PENDING_REVIEW_KEY] = pending
            return pending
    return None


def pending_review_data(pending: dict[str, Any]) -> dict[str, Any] | None:
    data = pending.get("data")
    return data if isinstance(data, dict) else None


def pending_missing_document_types(pending: dict[str, Any]) -> list[str]:
    documents = pending.get("documents") or {}
    missing = []
    if DOCUMENT_TYPE_DELIVERY_ORDER not in documents:
        missing.append(DOCUMENT_TYPE_DELIVERY_ORDER)
    if DOCUMENT_TYPE_INVOICE not in documents:
        missing.append(DOCUMENT_TYPE_INVOICE)
    return missing


def pending_pair_summary(pending: dict[str, Any]) -> str:
    documents = pending.get("documents") or {}
    lines = []
    for document_type in (DOCUMENT_TYPE_DELIVERY_ORDER, DOCUMENT_TYPE_INVOICE):
        record = documents.get(document_type)
        if not record:
            lines.append(f"{document_type_label(document_type)}: missing")
            continue
        data = record.get("data") or {}
        number = data.get("tax_invoice") or data.get("invoice_number") or "number unknown"
        date = data.get("invoice_date") or "date unknown"
        rows = len(line_items_from_data(data))
        lines.append(f"{document_type_label(document_type)}: {number}, {date}, {rows} row(s)")
    return "\n".join(lines)


def rebuild_pending_pair_with_item_source(pending: dict[str, Any], item_source: str) -> dict[str, Any] | None:
    documents = pending.get("documents") or {}
    delivery_order_record = documents.get(DOCUMENT_TYPE_DELIVERY_ORDER)
    invoice_record = documents.get(DOCUMENT_TYPE_INVOICE)
    if not delivery_order_record or not invoice_record:
        return None

    merged_data, compare_warnings = compare_and_merge_documents(
        delivery_order_record["data"],
        invoice_record["data"],
        item_source,
    )
    merged_data["submitter_chat_id"] = pending.get("submitter_chat_id") or ""
    merged_data["submitter_name"] = pending.get("submitter_name") or ""
    merged_data["record_type"] = invoice_record_type(pending.get("submitter_chat_id"))
    merged_data["delivery_order_image_path"] = delivery_order_record["image_path"]
    merged_data["invoice_image_path"] = invoice_record["image_path"]
    pending["data"] = merged_data
    pending["pair_compare_warnings"] = compare_warnings
    pending["item_source_choice"] = merged_data.get("item_source") or DOCUMENT_TYPE_DELIVERY_ORDER
    pending["item_source_confirmed"] = True
    pending["invoice_id"] = delivery_order_record["invoice_id"]
    pending["received_at"] = delivery_order_record["received_at"]
    pending["image_path"] = delivery_order_record["image_path"]
    return merged_data


def clear_pending_review(context: ContextTypes.DEFAULT_TYPE, chat_id: Any = None) -> None:
    pending = context.chat_data.pop(PENDING_REVIEW_KEY, None)
    target_chat_id = chat_id or (pending_chat_id(pending) if isinstance(pending, dict) else None)
    if target_chat_id:
        pending_store.clear_pending(DATA_DIR, target_chat_id)


def mark_duplicate_reviewed(pending: dict[str, Any]) -> None:
    pending["duplicate_reviewed"] = True


def mark_auto_save_started(pending: dict[str, Any]) -> bool:
    if pending.get("auto_save_started"):
        return False
    pending["auto_save_started"] = True
    return True


def user_facing_error(exc: Exception) -> str:
    if isinstance(exc, UnknownDocumentFormat):
        return str(exc)
    if isinstance(exc, AuthenticationError):
        return AUTH_HELP
    if isinstance(exc, WorkbookBusyError):
        return "Excel is open, so I kept this document waiting. Close the Excel file, then send /save again. No need to capture the image again."
    if isinstance(exc, PermissionError):
        return "Excel blocked saving the workbook. I kept this document waiting. Close the Excel file, then send /save again."
    return f"Extraction failed: {exc}"


def duplicate_notice_text(
    tax_invoice: str,
    existing_count: int,
    extracted_count: int,
    new_count: int,
    repeated_count: int,
) -> str:
    if new_count:
        return (
            f"D.O / Invoice No {tax_invoice} already exists in Excel.\n"
            f"Existing rows: {existing_count}\n"
            f"Extracted rows now: {extracted_count}\n"
            f"Existing item numbers detected: {repeated_count}\n"
            f"Additional/new rows detected: {new_count}\n\n"
            "Send /save again to add only the additional rows, /saveall to add all extracted rows anyway, or /cancel to stop."
        )
    return (
        f"D.O / Invoice No {tax_invoice} already exists in Excel.\n"
        f"Existing rows: {existing_count}\n"
        f"Extracted rows now: {extracted_count}\n"
        "No additional item rows were detected.\n\n"
        "Send /save again if you still want to add this duplicate document again, or /cancel to stop."
    )


def template_display_date(value: Any) -> Any:
    if not value:
        return value
    text = str(value).strip()
    try:
        parsed = datetime.fromisoformat(text).date()
    except ValueError:
        return value
    return parsed.strftime("%d/%m/%Y")


def purchase_order_date_from_invoice(value: Any) -> Any:
    if not value:
        return value
    text = str(value).strip()
    try:
        parsed = datetime.fromisoformat(text).date()
    except ValueError:
        return value

    po_date = parsed - timedelta(days=4)
    if po_date.weekday() == 6:
        po_date -= timedelta(days=1)
    return po_date.isoformat()


def date_request_from_purchase_order_date(value: Any) -> Any:
    if not value:
        return value
    text = str(value).strip()
    try:
        parsed = datetime.fromisoformat(text).date()
    except ValueError:
        return value
    return (parsed - timedelta(days=2)).isoformat()


def clear_template_items(worksheet: Any) -> None:
    for row in range(TEMPLATE_FIRST_ITEM_ROW, TEMPLATE_LAST_ITEM_ROW + 1):
        for column in ("B", "C", "G", "H", "J"):
            worksheet[f"{column}{row}"] = None


def clear_material_requisition_items(worksheet: Any) -> None:
    for row in range(MR_FIRST_ITEM_ROW, MR_LAST_ITEM_ROW + 1):
        for col in ("B", "C", "H", "K", "L", "N"):
            worksheet[f"{col}{row}"] = None


def save_template_workbook(
    target_path: Path,
    template_path: Path,
    data: dict[str, Any],
) -> int:
    line_items = line_items_from_data(data)
    if len(line_items) > TEMPLATE_LAST_ITEM_ROW - TEMPLATE_FIRST_ITEM_ROW + 1:
        raise RuntimeError(
            f"Template only has space for {TEMPLATE_LAST_ITEM_ROW - TEMPLATE_FIRST_ITEM_ROW + 1} item rows, "
            f"but extraction has {len(line_items)} rows."
        )

    assert_workbook_writable(target_path)
    workbook = load_workbook(template_path)
    if TEMPLATE_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Template sheet {TEMPLATE_SHEET_NAME!r} was not found.")

    worksheet = workbook[TEMPLATE_SHEET_NAME]
    clear_template_items(worksheet)

    tax_invoice = data.get("tax_invoice") or data.get("invoice_number") or data.get("quotation_number")
    invoice_date = data.get("invoice_date")
    po_date = data.get("po_document_date") or purchase_order_date_from_invoice(invoice_date)
    data["po_document_date"] = po_date

    # PR Number, Dates & References
    pr_number = data.get("pr_number") or target_path.stem
    worksheet["J15"] = pr_number
    worksheet["J16"] = template_display_date(po_date)
    worksheet["J17"] = tax_invoice or ""
    worksheet["J18"] = data.get("delivery_order_no") or (tax_invoice if data.get("document_type") == DOCUMENT_TYPE_DELIVERY_ORDER else "")

    # Dynamic Supplier Profile into Vendor Block (B15:B20)
    supplier_profile = data.get("supplier_profile") or suppliers.detect_supplier_profile(
        data, configured_default_supplier(), configured_supplier_aliases()
    )
    if supplier_profile:
        worksheet["B15"] = supplier_profile.get("display_name") or "SUPPLIER"
        worksheet["B16"] = supplier_profile.get("address_line1") or ""
        worksheet["B17"] = supplier_profile.get("address_line2") or ""
        worksheet["B18"] = supplier_profile.get("tel_fax") or ""
        worksheet["B19"] = supplier_profile.get("email") or ""
        worksheet["B20"] = supplier_profile.get("bank_account") or ""

    contact = delivery_order_requested_by(data) or (supplier_profile.get("default_contact") if supplier_profile else "") or "Lukman 018-9414868"
    worksheet["H59"] = contact

    for offset, item in enumerate(line_items):
        row = TEMPLATE_FIRST_ITEM_ROW + offset
        worksheet[f"B{row}"] = item.get("item_no") or offset + 1
        desc = str(item.get("description") or "").strip()
        desc_cell = worksheet[f"C{row}"]
        desc_cell.value = desc
        desc_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        worksheet[f"G{row}"] = format_quantity_with_unit(item)
        worksheet[f"H{row}"] = normalize_number(item.get("unit_price"))
        worksheet[f"J{row}"] = normalize_number(item.get("line_total"))
        if desc:
            lines = max(1, math.ceil(len(desc) / 38))
            if lines > 1:
                worksheet.row_dimensions[row].height = max(24.0, lines * 16.0)

    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_path)
    workbook.close()
    return len(line_items)


def material_requisition_workbook_path(po_workbook_path: Path) -> Path:
    return po_workbook_path.with_name(f"{po_workbook_path.stem} - MR.xlsx")


def delivery_order_requested_by(data: dict[str, Any]) -> str:
    return str(data.get("delivery_order_contact_person") or data.get("contact_person") or "").strip()


def save_material_requisition_workbook(
    target_path: Path,
    template_path: Path,
    data: dict[str, Any],
    po_reference: str,
) -> int:
    line_items = line_items_from_data(data)
    if len(line_items) > MR_LAST_ITEM_ROW - MR_FIRST_ITEM_ROW + 1:
        raise RuntimeError(
            f"Material Requisition template only has space for {MR_LAST_ITEM_ROW - MR_FIRST_ITEM_ROW + 1} item rows, "
            f"but extraction has {len(line_items)} rows."
        )

    assert_workbook_writable(target_path)
    workbook = load_workbook(template_path)
    if MR_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Material Requisition template sheet {MR_SHEET_NAME!r} was not found.")

    worksheet = workbook[MR_SHEET_NAME]
    clear_material_requisition_items(worksheet)

    po_date = data.get("po_document_date") or purchase_order_date_from_invoice(data.get("invoice_date"))
    date_request = date_request_from_purchase_order_date(po_date)
    data["material_requisition_date_request"] = date_request
    data["material_requisition_reference"] = po_reference

    supplier_profile = data.get("supplier_profile") or suppliers.detect_supplier_profile(
        data, configured_default_supplier(), configured_supplier_aliases()
    )
    requested_by = delivery_order_requested_by(data) or (supplier_profile.get("default_contact") if supplier_profile else "") or "Lukman 018-9414868"

    worksheet["N10"] = str(data.get("pr_number") or po_reference)
    worksheet["N12"] = str(template_display_date(date_request))
    worksheet["D15"] = requested_by
    worksheet["D52"] = requested_by

    if supplier_profile:
        worksheet["E44"] = supplier_profile.get("display_name") or ""
        worksheet["E45"] = supplier_profile.get("address_line1") or ""
        worksheet["E46"] = supplier_profile.get("address_line2") or ""

    for offset, item in enumerate(line_items):
        row = MR_FIRST_ITEM_ROW + offset
        worksheet[f"B{row}"] = item.get("item_no") or offset + 1
        desc = str(item.get("description") or "").strip()
        desc_cell = worksheet[f"C{row}"]
        desc_cell.value = desc
        desc_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        worksheet[f"K{row}"] = format_quantity_with_unit(item) or ""
        unit_price = normalize_number(item.get("unit_price"))
        if unit_price is not None:
            worksheet[f"L{row}"] = unit_price
        line_total = normalize_number(item.get("line_total"))
        if line_total is not None:
            worksheet[f"N{row}"] = line_total
        if desc:
            lines = max(1, math.ceil(len(desc) / 38))
            if lines > 1:
                worksheet.row_dimensions[row].height = max(24.0, lines * 16.0)

    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_path)
    workbook.close()
    return len(line_items)


def powershell_single_quote(value: Path) -> str:
    return str(value).replace("'", "''")


def powershell_single_quote_text(value: Any) -> str:
    return str(value).replace("'", "''")


def find_libreoffice_cmd() -> str | None:
    configured = os.getenv("LIBREOFFICE_CMD")
    if configured and (shutil.which(configured) or Path(configured).exists()):
        return configured
    for candidate in ("libreoffice", "soffice"):
        found = shutil.which(candidate)
        if found:
            return found
    if os.name == "nt":
        for default_win in (
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ):
            if Path(default_win).exists():
                return default_win
    return None


def export_workbook_to_pdf_libreoffice(workbook_path: Path, output_dir: Path) -> Path:
    cmd = find_libreoffice_cmd()
    if not cmd:
        raise RuntimeError("LibreOffice command not found.")
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = output_dir / f"{workbook_path.stem}.pdf"
    if pdf_path.exists():
        pdf_path.unlink()
    res = subprocess.run(
        [cmd, "--headless", "--convert-to", "pdf", "--outdir", str(output_dir), str(workbook_path)],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if res.returncode != 0 or not pdf_path.exists():
        error = res.stderr.strip() or res.stdout.strip() or "LibreOffice conversion failed"
        raise RuntimeError(f"LibreOffice PDF export failed: {error}")
    return pdf_path


def export_workbook_to_pdf(workbook_path: Path, sheet_name: str = TEMPLATE_SHEET_NAME) -> Path:
    pdf_path = workbook_path.with_suffix(".pdf")
    workbook_path = workbook_path.resolve()
    pdf_path = pdf_path.resolve()
    if pdf_path.exists():
        assert_workbook_writable(pdf_path)
        pdf_path.unlink()

    # 1. Try LibreOffice first if available (cross-platform, Linux GCP VM standard)
    if find_libreoffice_cmd():
        try:
            return export_workbook_to_pdf_libreoffice(workbook_path, workbook_path.parent)
        except Exception as exc:
            if os.name != "nt":
                raise
            logging.warning("LibreOffice PDF export failed, falling back to Excel COM: %s", exc)

    # 2. Fallback to Windows PowerShell Excel COM if on Windows
    if os.name == "nt" and shutil.which("powershell"):
        command = f"""
$ErrorActionPreference = 'Stop'
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open('{powershell_single_quote(workbook_path)}', 3, $true)
    $worksheet = $workbook.Worksheets.Item('{sheet_name}')
    $worksheet.PageSetup.Zoom = $false
    $worksheet.PageSetup.FitToPagesWide = 1
    $worksheet.PageSetup.FitToPagesTall = 1
    $workbook.ExportAsFixedFormat(0, '{powershell_single_quote(pdf_path)}')
    $workbook.Close($false)
}} finally {{
    if ($workbook -ne $null) {{
        try {{ $workbook.Close($false) }} catch {{ }}
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }}
    if ($excel -ne $null) {{
        $excel.Quit()
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
    }}
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}}
"""
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", command],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            error = result.stderr.strip() or result.stdout.strip() or "unknown error"
            raise RuntimeError(f"PDF export failed: {error}")
        if not pdf_path.exists():
            raise RuntimeError("PDF export failed: output file was not created.")
        return pdf_path

    raise RuntimeError("No PDF exporter found. Install LibreOffice (on Linux/Windows) or Excel (on Windows).")


def export_pdf_if_enabled(workbook_path: Path) -> Path | None:
    if not env_bool("EXPORT_PDF", DEFAULT_EXPORT_PDF):
        return None
    return export_workbook_to_pdf(workbook_path)


def export_pdf_safely(workbook_path: Path) -> tuple[Path | None, str | None]:
    if not env_bool("EXPORT_PDF", DEFAULT_EXPORT_PDF):
        return None, None
    try:
        return export_workbook_to_pdf(workbook_path), None
    except Exception as exc:
        logging.exception("PDF export failed for %s", workbook_path)
        return None, str(exc)


def export_material_requisition_pdf_safely(workbook_path: Path) -> tuple[Path | None, str | None]:
    if not env_bool("EXPORT_PDF", DEFAULT_EXPORT_PDF):
        return None, None
    try:
        return export_workbook_to_pdf(workbook_path, MR_SHEET_NAME), None
    except Exception as exc:
        logging.exception("Material Requisition PDF export failed for %s", workbook_path)
        return None, str(exc)


def create_material_requisition_outputs(
    po_workbook_path: Path,
    data: dict[str, Any],
) -> tuple[Path | None, Path | None, str | None]:
    template_path = configured_material_requisition_template_path()
    if not template_path:
        return None, None, "Material Requisition template is not configured."
    if not template_path.exists():
        return None, None, f"Material Requisition template file was not found: {template_path}"

    mr_workbook_path = material_requisition_workbook_path(po_workbook_path)
    try:
        save_material_requisition_workbook(
            mr_workbook_path,
            template_path,
            data,
            po_workbook_path.stem,
        )
    except Exception as exc:
        logging.exception("Material Requisition workbook creation failed for %s", po_workbook_path)
        return None, None, str(exc)

    mr_pdf_path, mr_pdf_error = export_material_requisition_pdf_safely(mr_workbook_path)
    return mr_workbook_path, mr_pdf_path, mr_pdf_error


def create_procurement_bundle(
    data: dict[str, Any],
    po_pdf_path: Path | None,
    mr_pdf_path: Path | None,
) -> tuple[Path, list[Path], list[str]]:
    invoice_number = data.get("tax_invoice") or data.get("invoice_number") or data.get("delivery_order_no") or "UNKNOWN"
    po_reference = data.get("po_output_stem") or invoice_number
    supplier_name = detected_supplier_name(data)
    data["supplier_name"] = supplier_name
    folder, created, issues = create_procurement_packet(
        configured_procurement_dir(),
        supplier_name,
        invoice_number,
        po_reference,
        po_pdf_path,
        mr_pdf_path,
        Path(data["invoice_image_path"]) if data.get("invoice_image_path") else None,
        Path(data["delivery_order_image_path"]) if data.get("delivery_order_image_path") else None,
    )

    data["procurement_folder"] = str(folder)
    data["procurement_files"] = "; ".join(str(path) for path in created)
    if issues:
        data["procurement_issues"] = "; ".join(issues)
    return folder, created, issues


def record_invoice_save(
    invoice_id: str,
    received_at: datetime,
    image_path: Path,
    data: dict[str, Any],
    workbook_path: Path,
    pdf_path: Path | None,
    pdf_error: str | None,
    saved_count: int,
    submitter_chat_id: str = "",
    submitter_name: str = "",
    material_requisition_path: Path | None = None,
    material_requisition_pdf_path: Path | None = None,
    material_requisition_error: str | None = None,
) -> None:
    register_path = configured_invoice_register_path()
    register_path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "processed_at",
        "received_at",
        "invoice_id",
        "tax_invoice",
        "invoice_date",
        "po_document_date",
        "record_type",
        "po_month_key",
        "po_month_name",
        "po_running_number",
        "po_output_stem",
        "supplier_name",
        "workbook_file",
        "pdf_file",
        "pdf_status",
        "material_requisition_file",
        "material_requisition_pdf_file",
        "material_requisition_status",
        "procurement_folder",
        "procurement_files",
        "procurement_issues",
        "rows_saved",
        "line_items_extracted",
        "extraction_method",
        "submitter_chat_id",
        "submitter_name",
        "contact_person",
        "source_image",
        "notes",
    ]
    if register_path.exists():
        with register_path.open("r", newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            existing_headers = reader.fieldnames or []
            existing_rows = list(reader)
        if existing_headers != headers:
            with register_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=headers)
                writer.writeheader()
                for existing_row in existing_rows:
                    writer.writerow({header: existing_row.get(header, "") for header in headers})
    row = {
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "received_at": received_at.isoformat(),
        "invoice_id": invoice_id,
        "tax_invoice": data.get("tax_invoice") or data.get("invoice_number") or "",
        "invoice_date": data.get("invoice_date") or "",
        "po_document_date": data.get("po_document_date") or purchase_order_date_from_invoice(data.get("invoice_date")) or "",
        "record_type": data.get("record_type") or invoice_record_type(submitter_chat_id),
        "po_month_key": data.get("po_month_key") or "",
        "po_month_name": data.get("po_month_name") or "",
        "po_running_number": data.get("po_running_number") or "",
        "po_output_stem": data.get("po_output_stem") or workbook_path.stem,
        "supplier_name": data.get("supplier_name") or detected_supplier_name(data),
        "workbook_file": str(workbook_path),
        "pdf_file": str(pdf_path) if pdf_path else "",
        "pdf_status": "disabled" if not env_bool("EXPORT_PDF", DEFAULT_EXPORT_PDF) else ("ok" if pdf_path else f"failed: {pdf_error}"),
        "material_requisition_file": str(material_requisition_path) if material_requisition_path else "",
        "material_requisition_pdf_file": str(material_requisition_pdf_path) if material_requisition_pdf_path else "",
        "material_requisition_status": (
            "ok"
            if material_requisition_path and (material_requisition_pdf_path or not env_bool("EXPORT_PDF", DEFAULT_EXPORT_PDF))
            else (f"failed: {material_requisition_error}" if material_requisition_error else "not created")
        ),
        "procurement_folder": data.get("procurement_folder") or "",
        "procurement_files": data.get("procurement_files") or "",
        "procurement_issues": data.get("procurement_issues") or "",
        "rows_saved": saved_count,
        "line_items_extracted": len(line_items_from_data(data)),
        "extraction_method": data.get("extraction_method") or "",
        "submitter_chat_id": submitter_chat_id,
        "submitter_name": submitter_name,
        "contact_person": data.get("contact_person") or "",
        "source_image": str(image_path),
        "notes": data.get("notes") or "",
    }
    file_exists = register_path.exists()
    with register_path.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)


def read_last_register_row() -> dict[str, str] | None:
    register_path = configured_invoice_register_path()
    if not register_path.exists():
        return None
    with register_path.open("r", newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    return rows[-1] if rows else None


def excel_com_available() -> tuple[bool, str | None]:
    command = """
$ErrorActionPreference = 'Stop'
$excel = $null
try {
    $excel = New-Object -ComObject Excel.Application
} finally {
    if ($excel -ne $null) {
        $excel.Quit()
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
    }
}
"""
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", command],
            capture_output=True,
            text=True,
            timeout=30,
        )
    except Exception as exc:
        return False, str(exc)
    if result.returncode != 0:
        return False, result.stderr.strip() or result.stdout.strip() or "Excel COM check failed."
    return True, None


def pdf_exporter_available() -> tuple[bool, str | None]:
    if find_libreoffice_cmd():
        return True, None
    if os.name == "nt" and shutil.which("powershell"):
        return excel_com_available()
    return False, "No PDF exporter found. Install LibreOffice on Linux or Excel on Windows."


def startup_self_check(test_excel: bool = False) -> list[str]:
    issues: list[str] = []

    if not os.getenv("TELEGRAM_BOT_TOKEN"):
        issues.append("TELEGRAM_BOT_TOKEN is missing.")
    if not configured_allowed_chat_ids():
        issues.append("TELEGRAM_ALLOWED_CHAT_IDS is not set; any chat can use the bot.")

    workbook_dir = configured_invoice_workbook_dir()
    try:
        workbook_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        issues.append(f"Invoice output folder is not writable: {exc}")

    template_path = configured_invoice_template_path()
    if template_path and not template_path.exists():
        issues.append(f"Invoice template file is missing: {template_path}")
    elif not template_path:
        issues.append("Invoice template is not configured; plain workbook output will be used.")

    mr_template_path = configured_material_requisition_template_path()
    if mr_template_path and not mr_template_path.exists():
        issues.append(f"Material Requisition template file is missing: {mr_template_path}")
    elif not mr_template_path:
        issues.append("Material Requisition template is not configured; MR output will not be created.")

    register_path = configured_invoice_register_path()
    try:
        register_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        issues.append(f"Invoice register folder is not writable: {exc}")

    try:
        configured_procurement_dir().mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        issues.append(f"Procurement folder is not writable: {exc}")

    if env_bool("CLEANUP_ENABLED", True):
        try:
            configured_cleanup_archive_dir().mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            issues.append(f"Cleanup archive folder is not writable: {exc}")

    if env_bool("LOCAL_OCR_ENABLED", True):
        tesseract_cmd = os.getenv("TESSERACT_CMD")
        if tesseract_cmd and not Path(tesseract_cmd).exists():
            issues.append(f"TESSERACT_CMD does not exist: {tesseract_cmd}")
        try:
            import_pytesseract()
        except Exception as exc:
            issues.append(f"Local OCR is enabled but unavailable: {exc}")

    if env_bool("EXPORT_PDF", DEFAULT_EXPORT_PDF) and test_excel:
        ok, error = pdf_exporter_available()
        if not ok:
            issues.append(f"PDF export is enabled but PDF automation is unavailable: {error}")

    return issues


def latest_invoice_files(submitter_chat_id: Any = None) -> list[Path]:
    workbook_dir = sender_invoice_workbook_dir(submitter_chat_id) if submitter_chat_id else configured_invoice_workbook_dir()
    if not workbook_dir.exists():
        return []
    candidates = [
        path
        for path in workbook_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in {".xlsx", ".pdf"} and not path.name.startswith("~$")
    ]
    if not candidates:
        return []
    latest = max(candidates, key=lambda path: path.stat().st_mtime)
    return [path for path in [latest.parent / f"{latest.stem}.xlsx", latest.parent / f"{latest.stem}.pdf"] if path.exists()]


def status_text(context: ContextTypes.DEFAULT_TYPE) -> str:
    workbook_dir = configured_invoice_workbook_dir()
    procurement_dir = configured_procurement_dir()
    template_path = configured_invoice_template_path()
    mr_template_path = configured_material_requisition_template_path()
    register_path = configured_invoice_register_path()
    pending = get_pending_review(context)
    last_row = read_last_register_row()
    issues = startup_self_check(test_excel=False)
    latest_files = latest_invoice_files()

    lines = [
        "Invoice bot status",
        f"Output folder: {workbook_dir}",
        "Output grouping: one subfolder per sender chat ID",
        f"Procurement folder: {procurement_dir}",
        f"Default supplier folder: {configured_procurement_supplier_name()}",
        f"Supplier aliases: {len(configured_supplier_aliases())} configured",
        f"Cleanup retention: {'enabled' if env_bool('CLEANUP_ENABLED', True) else 'disabled'}, archive OCR/temp files older than {configured_cleanup_retention_days()} day(s)",
        f"Template: {template_path if template_path else 'not configured'}",
        f"Template found: {'yes' if template_path and template_path.exists() else 'no'}",
        f"MR template: {mr_template_path if mr_template_path else 'not configured'}",
        f"MR template found: {'yes' if mr_template_path and mr_template_path.exists() else 'no'}",
        f"Authorized chats: {len(configured_allowed_chat_ids()) if configured_allowed_chat_ids() else 'any'}",
        f"Testing chats: {len(configured_test_chat_ids())}",
        f"PDF export: {'enabled' if env_bool('EXPORT_PDF', DEFAULT_EXPORT_PDF) else 'disabled'}",
        f"AI provider: {configured_ai_provider().upper()} ({ai_model_name()})",
        f"AI fallback: {'enabled' if ai_fallback_enabled() else 'disabled'}",
        f"AI primary: {'enabled' if ai_primary_enabled() else 'disabled'}",
        f"Local OCR: {'enabled' if env_bool('LOCAL_OCR_ENABLED', True) else 'disabled'}",
        f"Register: {register_path}",
        f"Pending review: {'yes' if pending else 'no'}",
        f"Pending pair: {pending_pair_summary(pending) if pending else 'none'}",
        f"Latest files: {', '.join(path.name for path in latest_files) if latest_files else 'none'}",
    ]
    if last_row:
        lines.append(f"Last saved invoice: {last_row.get('tax_invoice') or last_row.get('invoice_id') or 'unknown'}")
        lines.append(f"Last PDF status: {last_row.get('pdf_status') or 'unknown'}")
    if issues:
        lines.append("Checks:")
        lines.extend(f"- {issue}" for issue in issues)
    else:
        lines.append("Checks: no obvious configuration issues.")
    return "\n".join(lines)


def append_to_workbook(
    path: Path,
    invoice_id: str,
    received_at: datetime,
    image_path: Path,
    data: dict[str, Any],
    duplicate_mode: str = "check",
) -> int:
    del received_at, image_path
    template_path = configured_invoice_template_path()
    if template_path:
        if not template_path.exists():
            raise RuntimeError(f"Invoice template file was not found: {template_path}")
        return save_template_workbook(path, template_path, data)

    assert_workbook_writable(path)
    line_items = line_items_from_data(data)
    ensure_workbook(path)
    ensure_workbook_headers(path)
    workbook = load_workbook(path)
    invoices = workbook["Invoices"]

    tax_invoice = data.get("tax_invoice") or data.get("invoice_number")
    invoice_date = data.get("invoice_date")
    rows_to_save = line_items

    if tax_invoice and duplicate_mode in {"check", "new_only"}:
        existing_items = existing_items_for_tax_invoice(workbook, str(tax_invoice))
        if existing_items:
            new_items, repeated_items = split_new_and_repeated_items(existing_items, line_items)
            if duplicate_mode == "check":
                raise DuplicateInvoiceNotice(
                    duplicate_notice_text(
                        str(tax_invoice),
                        len(existing_items),
                        len(line_items),
                        len(new_items),
                        len(repeated_items),
                    )
                )
            rows_to_save = new_items or line_items

    if not rows_to_save:
        invoices.append([tax_invoice, invoice_date, None, None, None, None, None, None])
        saved_count = 1
    else:
        saved_count = 0
        for item in rows_to_save:
            invoices.append(
                [
                    tax_invoice,
                    invoice_date,
                    item.get("item_no"),
                    item.get("description"),
                    normalize_number(item.get("quantity")),
                    normalize_quantity_unit(item.get("quantity_unit")),
                    normalize_number(item.get("unit_price")),
                    normalize_number(item.get("line_total")),
                ]
            )
            saved_count += 1

    workbook.save(path)
    return saved_count


async def auto_save_when_workbook_available(
    application: Application,
    chat_id: int,
    invoice_id: str,
    force_record: bool = False,
) -> None:
    for _ in range(AUTO_SAVE_MAX_ATTEMPTS):
        await asyncio.sleep(AUTO_SAVE_RETRY_SECONDS)
        pending = application.chat_data.get(chat_id, {}).get(PENDING_REVIEW_KEY)
        if not isinstance(pending, dict) or pending.get("invoice_id") != invoice_id:
            return

        try:
            received_at = datetime.fromisoformat(pending["received_at"])
            image_path = Path(pending["image_path"])
            data = pending["data"]
            duplicate_mode = "new_only" if pending.get("duplicate_reviewed") else "check"
            data["submitter_chat_id"] = str(pending.get("submitter_chat_id") or "")
            data["submitter_name"] = str(pending.get("submitter_name") or "")
            data["record_type"] = invoice_record_type(data.get("submitter_chat_id"), force_record or bool(pending.get("force_record")))
            target_path = invoice_workbook_path(invoice_id, data, received_at)
            saved_count = await asyncio.to_thread(append_to_workbook, target_path, invoice_id, received_at, image_path, data, duplicate_mode)
            pdf_path, pdf_error = await asyncio.to_thread(export_pdf_safely, target_path)
            mr_path, mr_pdf_path, mr_error = await asyncio.to_thread(create_material_requisition_outputs, target_path, data)
            procurement_folder, procurement_files, procurement_issues = await asyncio.to_thread(create_procurement_bundle, data, pdf_path, mr_pdf_path)
            record_invoice_save(
                invoice_id,
                received_at,
                image_path,
                data,
                target_path,
                pdf_path,
                pdf_error,
                saved_count,
                str(pending.get("submitter_chat_id") or ""),
                str(pending.get("submitter_name") or ""),
                mr_path,
                mr_pdf_path,
                mr_error,
            )
        except (WorkbookBusyError, PermissionError):
            continue
        except DuplicateInvoiceNotice as exc:
            pending["auto_save_started"] = False
            mark_duplicate_reviewed(pending)
            persist_pending_review(pending)
            await application.bot.send_message(chat_id=chat_id, text=exc.message)
            return
        except Exception:
            logging.exception("Automatic save for pending document failed")
            pending["auto_save_started"] = False
            persist_pending_review(pending)
            await application.bot.send_message(
                chat_id=chat_id,
                text="I still have this document waiting, but automatic saving hit an error. Send /save to try again.",
            )
            return

        application.chat_data[chat_id].pop(PENDING_REVIEW_KEY, None)
        pending_store.clear_pending(DATA_DIR, chat_id)
        tax_invoice = data.get("tax_invoice") or data.get("invoice_number") or "number unknown"
        line_item_count = len(line_items_from_data(data))
        record_label = "Testing invoice" if data.get("record_type") == "test" else "Recorded invoice"
        target_path = invoice_workbook_path(invoice_id, data, received_at)
        pdf_line = f"PDF: {pdf_path.name if pdf_path else 'PDF export disabled'}"
        if pdf_error:
            pdf_line = f"PDF export failed: {pdf_error}"
        mr_line = f"MR: {mr_path.name if mr_path else 'not created'}"
        if mr_pdf_path:
            mr_line += f"\nMR PDF: {mr_pdf_path.name}"
        if mr_error:
            mr_line += f"\nMR issue: {mr_error}"
        procurement_line = f"Procurement folder: {procurement_folder}"
        if procurement_issues:
            procurement_line += "\nProcurement issues: " + "; ".join(procurement_issues)
        await application.bot.send_message(
            chat_id=chat_id,
            text=(
                "Excel is closed now, so I saved the document.\n"
                f"Type: {record_label}\n"
                f"D.O / Invoice No: {tax_invoice}\n"
                f"File: {target_path.name}\n"
                f"{pdf_line}\n"
                f"{mr_line}\n"
                f"{procurement_line}\n"
                f"Extracted line items: {line_item_count}\n"
                f"Rows saved: {saved_count}\n"
                "Sending a copy of the saved file(s) now."
            ),
        )
        files_to_send = [target_path] + ([pdf_path] if pdf_path else []) + ([mr_path] if mr_path else []) + ([mr_pdf_path] if mr_pdf_path else [])
        sent_files = await send_saved_invoice_files(application, chat_id, files_to_send)
        if not sent_files:
            await application.bot.send_message(chat_id=chat_id, text="Saved successfully, but I could not send the generated file copy.")
        return

    pending = application.chat_data.get(chat_id, {}).get(PENDING_REVIEW_KEY)
    if isinstance(pending, dict) and pending.get("invoice_id") == invoice_id:
        pending["auto_save_started"] = False
        persist_pending_review(pending)
        await application.bot.send_message(
            chat_id=chat_id,
            text="I kept the document waiting, but Excel stayed locked too long. Close the Excel file and send /save again.",
        )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    del context
    if update.effective_chat:
        user = update.effective_user
        logging.info(
            "Start command from chat_id=%s user=%s username=%s",
            update.effective_chat.id,
            user.full_name if user else None,
            user.username if user else None,
        )
    if update.message:
        await update.message.reply_text(
            "Send me the D.O photo and the matching invoice photo. I will compare both documents first, then wait for /save before writing the P.O to Excel.\n\n"
            "Commands (type / to see all):\n"
            "/save \u2014 Save with auto filename\n"
            "/savewithdifferentname <name> \u2014 Save with custom filename\n"
            "/editnothensave <number> \u2014 Save with specific running number\n"
            "/review \u2014 Show current pending pair\n"
            "/cancel \u2014 Cancel pending review\n"
            "/status \u2014 Show bot configuration"
        )


async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await reply_long_text(update, status_text(context))


async def procurement_query_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message:
        return
    if not await is_authorized(update):
        return

    question = " ".join(context.args or []).strip()
    if not question:
        await update.message.reply_text(
            "Use /procure followed by a question, for example:\n"
            "/procure what item was bought the most\n"
            "/procure most expensive item\n"
            "/procure what did we buy from Southern Cable"
        )
        return

    result = await asyncio.to_thread(answer_question, question)
    table = format_table(result.get("rows", []), limit=8)
    response = str(result.get("answer", "No answer found."))
    if table:
        response = f"{response}\n\n{table}"
    await reply_long_text(update, response)


async def cleanup_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message:
        return
    if not await is_authorized(update):
        return
    dry_run = bool(context.args and context.args[0].lower() in {"check", "dry-run", "dryrun", "preview"})
    try:
        result = run_cleanup_retention(dry_run=dry_run)
    except Exception as exc:
        logging.exception("Cleanup retention failed")
        await update.message.reply_text(f"Cleanup failed: {exc}")
        return

    lines = [
        "Cleanup retention complete.",
        f"Mode: {'preview' if dry_run else 'archive'}",
        f"Rule: OCR/temp files older than {configured_cleanup_retention_days()} day(s)",
        f"Files scanned: {result.scanned}",
        f"Files {'would archive' if dry_run else 'archived'}: {result.scanned if dry_run else result.archived}",
        f"Archive folder: {result.archive_dir if result.archive_dir else 'none'}",
    ]
    if result.issues:
        lines.append("Issues:")
        lines.extend(f"- {issue}" for issue in result.issues[:10])
    await reply_long_text(update, "\n".join(lines))


async def last_invoice_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message:
        return
    if not await is_authorized(update):
        return

    files = latest_invoice_files(update.effective_chat.id if update.effective_chat else None)
    if not files:
        await update.message.reply_text("No saved invoice Excel or PDF files were found yet.")
        return

    await update.message.reply_text("Sending latest saved invoice file(s).")
    for path in files:
        try:
            with path.open("rb") as handle:
                await update.message.reply_document(
                    document=handle,
                    filename=path.name,
                    connect_timeout=15,
                    read_timeout=60,
                    write_timeout=60,
                    pool_timeout=15,
                )
        except TelegramError:
            logging.exception("Telegram failed while sending latest invoice file: %s", path)
            await update.message.reply_text(f"Could not send {path.name}.")


async def send_saved_invoice_files(application: Application, chat_id: int, files: list[Path]) -> list[str]:
    sent: list[str] = []
    for path in files:
        if not path or not path.exists():
            continue
        try:
            with path.open("rb") as handle:
                await application.bot.send_document(
                    chat_id=chat_id,
                    document=handle,
                    filename=path.name,
                    connect_timeout=15,
                    read_timeout=60,
                    write_timeout=60,
                    pool_timeout=15,
                )
            sent.append(path.name)
        except TelegramError:
            logging.exception("Telegram failed while sending saved invoice file %s to chat_id=%s", path, chat_id)
    return sent


async def whoami(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    del context
    if update.message and update.effective_chat:
        user = update.effective_user
        logging.info(
            "Whoami command from chat_id=%s user=%s username=%s",
            update.effective_chat.id,
            user.full_name if user else None,
            user.username if user else None,
        )
        await update.message.reply_text(f"Your Telegram chat ID is: {update.effective_chat.id}")


async def process_invoice_image(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    image_path: Path,
    invoice_id: str,
    received_at: datetime,
    source_image_hash: str | None = None,
) -> None:
    if not update.message:
        return

    pending_before = get_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    if pending_before and pending_review_data(pending_before):
        await update.message.reply_text(
            "You already have a matched D.O + invoice pair waiting. Use /save to write it to Excel, /review to see it again, or /cancel to discard it before sending another pair."
        )
        return

    user_name = update.effective_user.first_name if update.effective_user else "there"
    await safe_reply_text(update, f"Hi {user_name}, document received. Please hold while we process your file(s)...", "document received acknowledgement")

    image_quality_report = await asyncio.to_thread(inspect_document_image_quality, image_path)
    logging.info("Image quality image=%s report=%s", image_path.name, image_quality_report)
    image_quality_warning = format_image_quality_warning(image_quality_report)
    if image_quality_warning and env_bool("IMAGE_QUALITY_WARN_USER", False):
        await safe_reply_text(update, image_quality_warning, "image quality warning")

    # ── Multi-version OCR comparison (only needed for legacy local OCR) ──
    if not ai_primary_enabled():
        enhanced_dir = image_path.parent
        if (enhanced_dir / "clean.jpg").exists():
            try:
                best_text, best_result = await asyncio.to_thread(
                    ocr_enhanced.run_multi_ocr_comparison,
                    enhanced_dir,
                    ocr_single_text_and_confidence,
                )
                if best_result and best_result.get("image_version"):
                    best_version = best_result["image_version"]
                    best_img = enhanced_dir / f"{best_version}.jpg"
                    if best_img.exists():
                        image_path = best_img
            except Exception:
                logging.exception("Multi-OCR comparison failed; keeping original image_path")

    try:
        cached_data = load_cached_extraction(source_image_hash)
        if cached_data:
            data = dict(cached_data)
            data["notes"] = f"{data.get('notes') or ''} Reused prior extraction for identical image.".strip()
            extraction_method = str(data.get("extraction_method") or "cached")
            extraction_reason = "Identical source image was processed before, so the previous extraction was reused."
            logging.info("Reused cached extraction for invoice_id=%s source_hash=%s", invoice_id, source_image_hash[:12] if source_image_hash else None)
        else:
            model = ai_model_name()
            data, extraction_method, extraction_reason = await extract_invoice_hybrid(image_path, model)
            if source_image_hash:
                data["source_image_hash"] = source_image_hash
                save_cached_extraction(source_image_hash, data)
        save_extraction_json(invoice_id, data)
    except Exception as exc:
        logging.exception("Invoice extraction failed")
        await update.message.reply_text(user_facing_error(exc))
        return

    tax_invoice = data.get("tax_invoice") or data.get("invoice_number") or "number unknown"
    line_item_count = len(line_items_from_data(data))
    submitter = update.effective_user
    pending = save_pending_review(
        context,
        invoice_id,
        received_at,
        image_path,
        data,
        update.effective_chat.id if update.effective_chat else None,
        submitter.full_name if submitter else "",
    )
    document_type = normalize_document_type(data)
    paired_data = pending_review_data(pending)
    supplier_profile = data.get("supplier_profile") or suppliers.detect_supplier_profile(
        data, configured_default_supplier(), configured_supplier_aliases()
    )
    is_standalone = (
        document_type in (DOCUMENT_TYPE_QUOTATION, DOCUMENT_TYPE_CASH_BILL)
        or supplier_profile.get("category") == "TECH"
        or "walihin" in str(supplier_profile.get("display_name", "")).lower()
    )

    if not paired_data and not is_standalone:
        missing = pending_missing_document_types(pending)
        next_needed = document_type_label(missing[0]) if missing else "matching document"
        is_album = bool(update.message and update.message.media_group_id)

        if is_album:
            await safe_reply_text(
                update,
                f"📥 {document_type_label(document_type)} ({tax_invoice}) received. Merging with matching album document...",
                "album document acknowledgement",
            )
            return

        supplier_title = supplier_profile.get("display_name") or data.get("supplier_name") or "Supplier"
        date_str = data.get("invoice_date") or "Date unknown"
        contact_str = data.get("contact_person") or "Contact unknown"

        single_summary = (
            f"📄 *{document_type_label(document_type)} Extracted*\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"• *Supplier:* {supplier_title}\n"
            f"• *Document No:* `{tax_invoice}`\n"
            f"• *Date:* {date_str}\n"
            f"• *Contact:* {contact_str}\n"
            f"• *Line Items:* {line_item_count}\n\n"
            f"📦 *Extracted Items:*\n{format_item_review(data)}\n\n"
            f"⏳ *Next Step:* Please upload the matching *{next_needed}* photo to generate the Purchase Order."
        )
        await safe_reply_text(update, single_summary, "single document extraction summary")
        return

    data = paired_data or data
    tax_invoice = data.get("tax_invoice") or data.get("invoice_number") or data.get("quotation_number") or "number unknown"
    line_item_count = len(line_items_from_data(data))

    supplier_title = supplier_profile.get("display_name") or data.get("supplier_name") or "Supplier"
    doc_label = "Quotation" if is_standalone else "D.O + Invoice Pair"
    date_str = data.get("invoice_date") or "Date unknown"
    contact_str = data.get("contact_person") or "Contact unknown"

    summary_text = (
        f"📋 *{doc_label} Ready for Review*\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"• *Supplier:* {supplier_title}\n"
        f"• *Reference No:* `{tax_invoice}`\n"
        f"• *PO Date:* {date_str}\n"
        f"• *Attention:* {contact_str}\n"
        f"• *Total Items:* {line_item_count}\n\n"
        f"📦 *Extracted Items:*\n{format_item_review(data)}\n\n"
        f"Please review the items above and tap *Save & Generate PO* below to proceed."
    )

    is_test = data.get("record_type") == "test"
    keyboard = review_action_keyboard(is_test=is_test)

    warnings_text = format_review_warnings(data)
    if warnings_text and "No critical issues" not in warnings_text:
        await safe_reply_text(update, f"⚠️ *Notes:*\n{warnings_text}", "review warnings")

    source_instruction = pair_item_source_instruction(data) if not is_standalone else ""
    if source_instruction:
        await safe_reply_text(update, source_instruction, "item source instructions")

    await safe_reply_text(update, summary_text, "document review summary", reply_markup=keyboard)


async def review_pending(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not get_update_message(update):
        return
    if not await is_authorized(update):
        return

    pending = get_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    if not pending:
        await safe_reply_text(update, "ℹ️ No document is currently waiting for review. Send a D.O or Quotation photo first.", "empty review notice")
        return

    data = pending_review_data(pending)
    if not data:
        await safe_reply_text(
            update,
            "ℹ️ I have one document stored, waiting for the matching pair to complete the review.\n" + pending_pair_summary(pending),
            "partial pair review notice",
        )
        return

    supplier_profile = data.get("supplier_profile") or suppliers.detect_supplier_profile(
        data, configured_default_supplier(), configured_supplier_aliases()
    )
    document_type = normalize_document_type(data)
    is_standalone = (
        document_type in (DOCUMENT_TYPE_QUOTATION, DOCUMENT_TYPE_CASH_BILL)
        or supplier_profile.get("category") == "TECH"
        or "walihin" in str(supplier_profile.get("display_name", "")).lower()
    )
    supplier_title = supplier_profile.get("display_name") or data.get("supplier_name") or "Supplier"
    doc_label = "Quotation" if is_standalone else "D.O + Invoice Pair"
    tax_invoice = data.get("tax_invoice") or data.get("invoice_number") or data.get("quotation_number") or "number unknown"
    date_str = data.get("invoice_date") or "Date unknown"
    contact_str = data.get("contact_person") or "Contact unknown"
    line_item_count = len(line_items_from_data(data))

    summary_text = (
        f"📋 *{doc_label} Review*\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"• *Supplier:* {supplier_title}\n"
        f"• *Reference No:* `{tax_invoice}`\n"
        f"• *PO Date:* {date_str}\n"
        f"• *Attention:* {contact_str}\n"
        f"• *Total Items:* {line_item_count}\n\n"
        f"📦 *Extracted Items:*\n{format_item_review(data)}\n\n"
        f"Tap *Save & Generate PO* below to confirm."
    )

    is_test = data.get("record_type") == "test"
    keyboard = review_action_keyboard(is_test=is_test)

    warnings_text = format_review_warnings(data)
    if warnings_text and "No critical issues" not in warnings_text:
        await safe_reply_text(update, f"⚠️ *Notes:*\n{warnings_text}", "review warnings")

    source_instruction = pair_item_source_instruction(data) if not is_standalone else ""
    if source_instruction:
        await safe_reply_text(update, source_instruction, "item source instructions")

    await safe_reply_text(update, summary_text, "document review summary", reply_markup=keyboard)


async def choose_pair_item_source(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    item_source: str,
) -> None:
    if not update.message:
        return
    if not await is_authorized(update):
        return

    pending = get_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    if not pending:
        await safe_reply_text(update, "No D.O + invoice pair is waiting. Send both images first.", "empty item source notice")
        return

    data = rebuild_pending_pair_with_item_source(pending, item_source)
    if not data:
        await safe_reply_text(
            update,
            "The P.O pair is not complete yet.\n" + pending_pair_summary(pending),
            "partial item source notice",
        )
        return

    persist_pending_review(pending)
    await safe_reply_text(
        update,
        f"Item source set to {pair_item_source_label(data)}.\nP.O line items found: {len(line_items_from_data(data))}",
        "item source selected",
    )
    await reply_long_text(update, f"Extracted items:\n{format_item_review(data)}")
    record_instruction = (
        " This sender is treated as testing by default; use /saverecord only when this document should become an official record."
        if data.get("record_type") == "test"
        else ""
    )
    await safe_reply_text(update, "Send /save to write it to Excel or /cancel to discard it." + record_instruction, "review instructions")


async def use_do_items(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await choose_pair_item_source(update, context, DOCUMENT_TYPE_DELIVERY_ORDER)


async def use_invoice_items(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await choose_pair_item_source(update, context, DOCUMENT_TYPE_INVOICE)


async def save_pending_with_mode(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    force_all: bool = False,
    force_record: bool = False,
    manual_running_number: int | None = None,
) -> None:
    message = get_update_message(update)
    if not message:
        return
    if not await is_authorized(update):
        return

    chat_id = update.effective_chat.id if update.effective_chat else None
    pending = get_pending_review(context, chat_id)
    if not pending:
        await safe_reply_text(update, "ℹ️ No document is waiting to be saved. Please upload your document photo first.", "empty save notice")
        return

    if not pending_review_data(pending):
        await safe_reply_text(
            update,
            "ℹ️ The document pair is not complete yet.\n"
            + pending_pair_summary(pending)
            + "\n\nPlease send the missing D.O or invoice image before saving.",
            "incomplete pair save notice",
        )
        return

    try:
        invoice_id = pending["invoice_id"]
        received_at = datetime.fromisoformat(pending["received_at"])
        image_path = Path(pending["image_path"])
        data = pending["data"]
        if force_all:
            duplicate_mode = "all"
        elif pending.get("duplicate_reviewed"):
            duplicate_mode = "new_only"
        else:
            duplicate_mode = "check"
        data["submitter_chat_id"] = str(pending.get("submitter_chat_id") or "")
        data["submitter_name"] = str(pending.get("submitter_name") or "")
        data["record_type"] = invoice_record_type(data.get("submitter_chat_id"), force_record)
        pending["force_record"] = force_record
        if manual_running_number is not None:
            if not manual_po_running_number_is_available(manual_running_number, received_at, data["record_type"]):
                _, month_name, _ = po_output_stem_for_running_number(manual_running_number, received_at, data["record_type"])
                await safe_reply_text(
                    update,
                    (
                        f"⚠️ P.O running number {manual_running_number:04d} is already used for "
                        f"{month_name} {data['record_type']} invoices. Please try another number."
                    ),
                    "manual running number collision notice",
                )
                return
            apply_manual_po_running_number(data, manual_running_number, received_at, data["record_type"])
        if pair_item_mismatch_warnings(data) and not pending.get("item_source_confirmed"):
            await safe_reply_text(
                update,
                pair_item_source_instruction(data),
                "item source required before save",
            )
            return
        target_path = invoice_workbook_path(invoice_id, data, received_at)
        saved_count = await asyncio.to_thread(append_to_workbook, target_path, invoice_id, received_at, image_path, data, duplicate_mode)
        pdf_path, pdf_error = await asyncio.to_thread(export_pdf_safely, target_path)
        mr_path, mr_pdf_path, mr_error = await asyncio.to_thread(create_material_requisition_outputs, target_path, data)
        procurement_folder, procurement_files, procurement_issues = await asyncio.to_thread(create_procurement_bundle, data, pdf_path, mr_pdf_path)
        record_invoice_save(
            invoice_id,
            received_at,
            image_path,
            data,
            target_path,
            pdf_path,
            pdf_error,
            saved_count,
            str(pending.get("submitter_chat_id") or ""),
            str(pending.get("submitter_name") or ""),
            mr_path,
            mr_pdf_path,
            mr_error,
        )
    except (WorkbookBusyError, PermissionError) as exc:
        logging.info("Workbook is locked; keeping document %s pending", pending.get("invoice_id"))
        if update.effective_chat and mark_auto_save_started(pending):
            persist_pending_review(pending)
            context.application.create_task(
                auto_save_when_workbook_available(
                    context.application,
                    update.effective_chat.id,
                    str(pending["invoice_id"]),
                    force_record,
                )
            )
        await safe_reply_text(
            update,
            user_facing_error(exc) + " Automatic retries are active.",
            "workbook locked notice",
        )
        return
    except DuplicateInvoiceNotice as exc:
        mark_duplicate_reviewed(pending)
        persist_pending_review(pending)
        await safe_reply_text(update, exc.message, "duplicate invoice notice")
        return
    except Exception as exc:
        logging.exception("Saving reviewed invoice failed")
        await safe_reply_text(update, user_facing_error(exc), "save failure notice")
        return

    clear_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    tax_invoice = data.get("tax_invoice") or data.get("invoice_number") or data.get("quotation_number") or "number unknown"
    line_item_count = len(line_items_from_data(data))

    logging.info("Saved reviewed invoice %s to %s; sending confirmation", invoice_id, target_path)
    await safe_reply_text(
        update,
        (
            f"✅ *Purchase Order & MR Generated!*\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"• *Document No:* `{tax_invoice}`\n"
            f"• *PO File:* `{target_path.name}`\n"
            f"• *Line Items Saved:* {saved_count}\n\n"
            f"📦 *Your files are ready below:*"
        ),
        "saved invoice confirmation",
    )
    sender_chat_id = int(pending.get("submitter_chat_id") or (update.effective_chat.id if update.effective_chat else 0))
    files_to_send = [target_path] + ([pdf_path] if pdf_path else []) + ([mr_path] if mr_path else []) + ([mr_pdf_path] if mr_pdf_path else [])
    sent_files = await send_saved_invoice_files(context.application, sender_chat_id, files_to_send)
    if not sent_files:
        await safe_reply_text(update, "Saved successfully to server, but could not deliver the file copy in chat.", "saved file send failure notice")


async def save_pending(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await save_pending_with_mode(update, context, force_all=False)


async def save_pending_as_record(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await save_pending_with_mode(update, context, force_all=False, force_record=True)


async def save_all_pending(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await save_pending_with_mode(update, context, force_all=True)


async def edit_number_then_save(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not get_update_message(update):
        return
    if not await is_authorized(update):
        return

    pending = get_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    if not pending or not pending_review_data(pending):
        await safe_reply_text(update, "ℹ️ No completed D.O + invoice pair is waiting. Send both images first.", "edit number notice")
        return

    raw_number = " ".join(context.args or []).strip()
    if not raw_number:
        context.chat_data[EDIT_PO_RUNNING_NUMBER_KEY] = {"force_record": False}
        await safe_reply_text(update, "Send the P.O running number to use, for example 7 or 0007. Send /cancel to stop.", "edit number prompt")
        return

    running_number = parse_po_running_number(raw_number)
    if running_number is None:
        await safe_reply_text(update, "Use a positive running number, for example /editnothensave 7 or /editnothensave 0007.", "invalid running number")
        return

    await save_pending_with_mode(update, context, force_all=False, force_record=False, manual_running_number=running_number)


async def save_with_custom_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not get_update_message(update):
        return
    if not await is_authorized(update):
        return

    pending = get_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    if not pending or not pending_review_data(pending):
        await safe_reply_text(update, "ℹ️ No completed D.O + invoice pair is waiting. Send both images first.", "custom name notice")
        return

    raw_name = " ".join(context.args or []).strip()
    if not raw_name:
        context.chat_data[SAVE_WITH_CUSTOM_NAME_KEY] = {"force_record": False}
        await safe_reply_text(
            update,
            "Send the custom filename to use for this invoice, for example: \"BFE PO TUJU JULY 9999\"\n"
            "The filename will be used as-is (with .xlsx appended). Send /cancel to stop.",
            "custom name prompt",
        )
        return

    await _do_save_with_custom_name(update, context, raw_name, force_record=False)


async def _do_save_with_custom_name(
    update: Update, context: ContextTypes.DEFAULT_TYPE, custom_name: str, force_record: bool = False
) -> None:
    pending = get_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    if not pending or not pending_review_data(pending):
        await safe_reply_text(update, "ℹ️ No completed D.O + invoice pair is waiting. Send both images first.", "custom name notice")
        return

    # Override the output stem with the custom name
    pending["data"]["po_output_stem"] = custom_name.strip()
    persist_pending_review(pending)

    await save_pending_with_mode(update, context, force_all=False, force_record=force_record, manual_running_number=None)


async def cancel_pending(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not get_update_message(update):
        return
    if not await is_authorized(update):
        return
    context.chat_data.pop(EDIT_PO_RUNNING_NUMBER_KEY, None)
    context.chat_data.pop(SAVE_WITH_CUSTOM_NAME_KEY, None)

    pending = get_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    if not pending:
        await safe_reply_text(update, "ℹ️ No document is currently waiting for review.", "cancel empty notice")
        return

    invoice_id = pending.get("invoice_id", "unknown")
    clear_pending_review(context, update.effective_chat.id if update.effective_chat else None)
    await safe_reply_text(
        update,
        f"🗑️ *Document Discarded*\n━━━━━━━━━━━━━━━━━━━\nPending document `{invoice_id}` has been discarded. You can upload new documents at any time.",
        "cancel confirmation",
    )


async def handle_action_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not query:
        return
    if not await is_authorized(update):
        await query.answer("Unauthorized.", show_alert=True)
        return

    data = query.data or ""
    if data == "btn_save":
        await query.answer("Generating Purchase Order & MR...", show_alert=False)
        await save_pending_with_mode(update, context, force_all=False)
    elif data == "btn_saverecord":
        await query.answer("Saving Official Record...", show_alert=False)
        await save_pending_with_mode(update, context, force_all=False, force_record=True)
    elif data == "btn_review":
        await query.answer()
        await review_pending(update, context)
    elif data == "btn_cancel":
        await query.answer("Discarded.")
        await cancel_pending(update, context)


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.message.photo:
        return
    if not await is_authorized(update):
        return

    # Check if in registration mode
    if registration_flow.is_in_registration_mode(context):
        received_at = datetime.now(timezone.utc)
        invoice_id = "reg_" + invoice_id_from_timestamp(received_at)
        IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        image_path = IMAGE_DIR / f"{invoice_id}.jpg"
        photo = update.message.photo[-1]
        telegram_file = await context.bot.get_file(photo.file_id)
        await telegram_file.download_to_drive(custom_path=image_path)
        await registration_flow.handle_registration_file(update, context, image_path)
        return

    # Check if in test mode
    if profile_management.get_test_state(context):
        received_at = datetime.now(timezone.utc)
        invoice_id = "test_" + invoice_id_from_timestamp(received_at)
        IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        image_path = IMAGE_DIR / f"{invoice_id}.jpg"
        photo = update.message.photo[-1]
        telegram_file = await context.bot.get_file(photo.file_id)
        await telegram_file.download_to_drive(custom_path=image_path)
        handled = await profile_management.handle_test_file(update, context, image_path)
        if handled:
            return

    received_at = datetime.now(timezone.utc)
    invoice_id = invoice_id_from_timestamp(received_at)
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    image_path = IMAGE_DIR / f"{invoice_id}.jpg"

    photo = update.message.photo[-1]
    telegram_file = await context.bot.get_file(photo.file_id)
    await telegram_file.download_to_drive(custom_path=image_path)
    source_image_hash = file_sha256(image_path)

    # ── OpenCV preprocessing pipeline (only needed for legacy local OCR) ──
    if not ai_primary_enabled():
        try:
            enhanced_dir = ENHANCED_OUTPUTS_DIR / invoice_id
            preprocess_result = await asyncio.to_thread(
                image_processor.preprocess_invoice, image_path, enhanced_dir,
            )
            image_path = preprocess_result["preprocessed_ocr"]
            logging.info("OpenCV preprocessing applied for %s.", invoice_id)
        except Exception:
            logging.exception(
                "OpenCV preprocessing failed for %s; falling back to raw image",
                invoice_id,
            )

    await process_invoice_image(update, context, image_path, invoice_id, received_at, source_image_hash)


async def handle_document_image(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.message.document:
        return
    if not await is_authorized(update):
        return

    mime_type = update.message.document.mime_type or ""
    file_name = update.message.document.file_name or "document"
    file_ext = Path(file_name).suffix.lower()

    # Support XLSX files for registration
    if registration_flow.is_in_registration_mode(context) and file_ext in (".xlsx", ".xls"):
        received_at = datetime.now(timezone.utc)
        invoice_id = "reg_" + invoice_id_from_timestamp(received_at)
        suffix = file_ext or ".xlsx"
        IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        file_path = IMAGE_DIR / f"{invoice_id}{suffix}"
        telegram_file = await context.bot.get_file(update.message.document.file_id)
        await telegram_file.download_to_drive(custom_path=file_path)
        await registration_flow.handle_registration_file(update, context, file_path)
        return

    if not mime_type.startswith("image/"):
        await update.message.reply_text("Please send an image file or photo of the D.O or invoice.")
        return

    if registration_flow.is_in_registration_mode(context):
        received_at = datetime.now(timezone.utc)
        invoice_id = "reg_" + invoice_id_from_timestamp(received_at)
        suffix = file_ext or ".jpg"
        IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        image_path = IMAGE_DIR / f"{invoice_id}{suffix}"
        telegram_file = await context.bot.get_file(update.message.document.file_id)
        await telegram_file.download_to_drive(custom_path=image_path)
        await registration_flow.handle_registration_file(update, context, image_path)
        return

    # Check if in test mode
    if profile_management.get_test_state(context):
        received_at = datetime.now(timezone.utc)
        invoice_id = "test_" + invoice_id_from_timestamp(received_at)
        suffix = file_ext or ".jpg"
        IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        image_path = IMAGE_DIR / f"{invoice_id}{suffix}"
        telegram_file = await context.bot.get_file(update.message.document.file_id)
        await telegram_file.download_to_drive(custom_path=image_path)
        handled = await profile_management.handle_test_file(update, context, image_path)
        if handled:
            return

    received_at = datetime.now(timezone.utc)
    invoice_id = invoice_id_from_timestamp(received_at)
    suffix = file_ext or ".jpg"
    suffix = Path(update.message.document.file_name or "document.jpg").suffix or ".jpg"
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    image_path = IMAGE_DIR / f"{invoice_id}{suffix}"

    telegram_file = await context.bot.get_file(update.message.document.file_id)
    await telegram_file.download_to_drive(custom_path=image_path)
    source_image_hash = file_sha256(image_path)

    # ── OpenCV preprocessing pipeline (only needed for legacy local OCR) ──
    if not ai_primary_enabled():
        try:
            enhanced_dir = ENHANCED_OUTPUTS_DIR / invoice_id
            preprocess_result = await asyncio.to_thread(
                image_processor.preprocess_invoice, image_path, enhanced_dir,
            )
            image_path = preprocess_result["preprocessed_ocr"]
            logging.info("OpenCV preprocessing applied for %s.", invoice_id)
        except Exception:
            logging.exception(
                "OpenCV preprocessing failed for %s; falling back to raw image",
                invoice_id,
            )

    await process_invoice_image(update, context, image_path, invoice_id, received_at, source_image_hash)


async def handle_other_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle non-image document files (DOCX, PDF, etc.) — primarily for registration mode."""
    if not update.message or not update.message.document:
        return
    if not await is_authorized(update):
        return

    mime_type = update.message.document.mime_type or ""
    file_name = update.message.document.file_name or "document"
    file_ext = Path(file_name).suffix.lower()

    # Only handle during registration mode
    if not registration_flow.is_in_registration_mode(context):
        await update.message.reply_text(
            "I can only process image files (photos or scanned images) for invoices. "
            "Please send a photo of the document."
        )
        return

    # In registration mode, accept any file type
    received_at = datetime.now(timezone.utc)
    invoice_id = "reg_" + invoice_id_from_timestamp(received_at)
    suffix = file_ext or ".bin"
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    file_path = IMAGE_DIR / f"{invoice_id}{suffix}"

    telegram_file = await context.bot.get_file(update.message.document.file_id)
    await telegram_file.download_to_drive(custom_path=file_path)

    await registration_flow.handle_registration_file(update, context, file_path)


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await is_authorized(update):
        return

    # Check registration mode text commands first
    if registration_flow.is_in_registration_mode(context):
        handled = await registration_flow.handle_registration_text(update, context)
        if handled:
            return

    if update.message:
        edit_state = context.chat_data.get(EDIT_PO_RUNNING_NUMBER_KEY)
        if isinstance(edit_state, dict):
            raw_number = update.message.text or ""
            running_number = parse_po_running_number(raw_number)
            if running_number is None:
                await update.message.reply_text("That is not a valid running number. Send a positive number like 7 or 0007, or /cancel to stop.")
                return
            context.chat_data.pop(EDIT_PO_RUNNING_NUMBER_KEY, None)
            await save_pending_with_mode(
                update,
                context,
                force_all=False,
                force_record=bool(edit_state.get("force_record")),
                manual_running_number=running_number,
            )
            return

        custom_name_state = context.chat_data.get(SAVE_WITH_CUSTOM_NAME_KEY)
        if isinstance(custom_name_state, dict):
            custom_name = (update.message.text or "").strip()
            if not custom_name:
                await update.message.reply_text("The filename cannot be empty. Send a valid filename or /cancel to stop.")
                return
            context.chat_data.pop(SAVE_WITH_CUSTOM_NAME_KEY, None)
            await _do_save_with_custom_name(
                update, context, custom_name, force_record=bool(custom_name_state.get("force_record")),
            )
            return
        await update.message.reply_text("Send me the D.O photo and the matching invoice photo, or attach them as image files.")


async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logging.exception("Unhandled error while processing update", exc_info=context.error)
    chat_id = None
    if isinstance(update, Update) and update.effective_chat:
        chat_id = update.effective_chat.id
    if chat_id is not None:
        try:
            await context.bot.send_message(
                chat_id=chat_id,
                text="Something went wrong handling that. Please try again, or send /status.",
            )
        except Exception:
            logging.exception("Failed to notify chat about an error")


async def reload_profiles_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /reload_profiles — reload document profiles from disk."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    try:
        count = len(reload_profiles())
        await update.message.reply_text(f"Reloaded {count} document profiles.")
    except Exception as exc:
        await update.message.reply_text(f"Failed to reload profiles: {exc}")


async def register_document_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /register_document — start the registration flow."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await registration_flow.register_document_command(update, context)


async def register_blank_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /blank — mark next file as blank template."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await registration_flow.register_blank_command(update, context)


async def register_supplier_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /supplier — set supplier name."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await registration_flow.register_supplier_command(update, context)


async def list_profiles_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /list_profiles — list all registered profiles."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await profile_management.list_profiles_command(update, context)


async def profile_info_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /profile_info <id> — show detailed profile info."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await profile_management.profile_info_command(update, context)


async def disable_profile_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /disable_profile <id> — soft-disable a profile."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await profile_management.disable_profile_command(update, context)


async def enable_profile_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /enable_profile <id> — re-enable a disabled profile."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await profile_management.enable_profile_command(update, context)


async def remove_profile_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /remove_profile <id> — move profile to removed/ directory."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await profile_management.remove_profile_command(update, context)


async def test_profile_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /test_profile <id> — start a test flow for a profile."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await profile_management.test_profile_command(update, context)


async def export_profiles_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /export_profiles — export all profiles as ZIP."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await profile_management.export_profiles_command(update, context)


async def import_profiles_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /import_profiles — import profiles from a ZIP."""
    if not update.message:
        return
    if not await is_authorized(update):
        return
    await profile_management.import_profiles_command(update, context)


BOT_COMMANDS = [
    BotCommand("start", "Start the bot and see instructions"),
    BotCommand("whoami", "Show your Telegram chat ID"),
    BotCommand("status", "Show bot configuration and status"),
    BotCommand("register_document", "Register a new document type profile"),
    BotCommand("reload_profiles", "Reload document profiles from disk"),
    BotCommand("list_profiles", "List all registered document profiles"),
    BotCommand("profile_info", "Show detailed info about a profile"),
    BotCommand("test_profile", "Test extraction with a profile"),
    BotCommand("export_profiles", "Export all profiles as ZIP"),
    BotCommand("import_profiles", "Import profiles from a ZIP"),
    BotCommand("save", "Save the reviewed D.O + invoice pair to Excel"),
    BotCommand("savewithdifferentname", "Save with a custom filename"),
    BotCommand("editnothensave", "Save with a specific running number"),
    BotCommand("saveall", "Force save all items including duplicates"),
    BotCommand("saverecord", "Save as official record (testing users)"),
    BotCommand("review", "Show the current pending pair review"),
    BotCommand("usedo", "Use D.O items as the item source"),
    BotCommand("useinvoice", "Use invoice items as the item source"),
    BotCommand("last", "Resend the latest saved files"),
    BotCommand("cancel", "Cancel pending review"),
    BotCommand("procure", "Query procurement data"),
    BotCommand("cleanup", "Archive old OCR/temp files"),
]


def build_application(token: str) -> Application:
    application = Application.builder().token(token).build()

    async def set_commands(app: Application) -> None:
        try:
            await app.bot.set_my_commands(BOT_COMMANDS)
        except Exception:
            logging.exception("Failed to set bot commands")

    application.post_init = set_commands

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("whoami", whoami))
    application.add_handler(CommandHandler("status", status_command))
    application.add_handler(CommandHandler("register_document", register_document_command))
    application.add_handler(CommandHandler("reload_profiles", reload_profiles_command))
    application.add_handler(CommandHandler("blank", register_blank_command))
    application.add_handler(CommandHandler("supplier", register_supplier_command))
    application.add_handler(CommandHandler("list_profiles", list_profiles_command))
    application.add_handler(CommandHandler("profile_info", profile_info_command))
    application.add_handler(CommandHandler("disable_profile", disable_profile_command))
    application.add_handler(CommandHandler("enable_profile", enable_profile_command))
    application.add_handler(CommandHandler("remove_profile", remove_profile_command))
    application.add_handler(CommandHandler("test_profile", test_profile_command))
    application.add_handler(CommandHandler("export_profiles", export_profiles_command))
    application.add_handler(CommandHandler("import_profiles", import_profiles_command))
    application.add_handler(CommandHandler("procure", procurement_query_command))
    application.add_handler(CommandHandler("cleanup", cleanup_command))
    application.add_handler(CommandHandler("last", last_invoice_command))
    application.add_handler(CommandHandler("review", review_pending))
    application.add_handler(CommandHandler("usedo", use_do_items))
    application.add_handler(CommandHandler("useinvoice", use_invoice_items))
    application.add_handler(CommandHandler("save", save_pending))
    application.add_handler(CommandHandler("approve", save_pending))
    application.add_handler(CommandHandler("saverecord", save_pending_as_record))
    application.add_handler(CommandHandler("approverecord", save_pending_as_record))
    application.add_handler(CommandHandler("saveall", save_all_pending))
    application.add_handler(CommandHandler("editnothensave", edit_number_then_save))
    application.add_handler(CommandHandler("savewithdifferentname", save_with_custom_name))
    application.add_handler(CommandHandler("cancel", cancel_pending))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(MessageHandler(filters.Document.IMAGE, handle_document_image))
    application.add_handler(MessageHandler(filters.Document.ALL & ~filters.Document.IMAGE, handle_other_document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    application.add_handler(CallbackQueryHandler(handle_action_callback, pattern="^btn_"))
    application.add_handler(CallbackQueryHandler(registration_flow.registration_callback, pattern="^reg_"))
    application.add_error_handler(on_error)
    return application


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Telegram invoice extractor to Excel.")
    parser.add_argument("--init-workbook", action="store_true", help="Create the Excel workbook and exit.")
    return parser.parse_args()


def main() -> None:
    load_dotenv(BASE_DIR / ".env", override=True)
    configure_logging()
    args = parse_args()

    workbook_dir = configured_invoice_workbook_dir()
    workbook_dir.mkdir(parents=True, exist_ok=True)

    if args.init_workbook:
        print(f"Invoice workbook folder ready: {workbook_dir}")
        return

    token = os.getenv("TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN is not set.")

    issues = startup_self_check(test_excel=True)
    if issues:
        for issue in issues:
            logging.warning("Startup check: %s", issue)
    else:
        logging.info("Startup checks passed.")

    if env_bool("CLEANUP_ENABLED", True):
        try:
            cleanup_result = run_cleanup_retention()
            if cleanup_result.scanned or cleanup_result.issues:
                logging.info(
                    "Cleanup retention archived %s/%s old OCR/temp file(s) to %s. Issues: %s",
                    cleanup_result.archived,
                    cleanup_result.scanned,
                    cleanup_result.archive_dir,
                    len(cleanup_result.issues),
                )
        except Exception:
            logging.exception("Startup cleanup retention failed.")

    application = build_application(token)
    logging.info("Invoice bot is running. Invoice workbook folder: %s", workbook_dir)
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()

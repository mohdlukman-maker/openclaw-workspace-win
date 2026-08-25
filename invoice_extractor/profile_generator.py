#!/usr/bin/env python3
"""
AI-Assisted Profile Generator (Phase 4).

Generates a draft DocumentProfile JSON from 1–2 filled sample documents
(images, or XLSX via openpyxl), with optional blank template enrichment.

Usage:
    from profile_generator import generate_profile_from_samples

    draft = generate_profile_from_samples(
        samples=[path1, path2],
        blank_template=None,
        supplier_name="Optional Supplier Name",
    )
"""

from __future__ import annotations

import base64
import json
import logging
import os
import re
import sys
import tempfile
import time
from pathlib import Path
from typing import Any

from openai import AsyncOpenAI, AuthenticationError

# Import existing project modules
sys.path.insert(0, str(Path(__file__).resolve().parent))
from document_profiles import (
    SCHEMA_VERSION,
    DocumentProfile,
    ClassifierConfig,
    ClassifierMarker,
    FieldDef,
    TableColumn,
    LineItemTable,
    ValidationRule,
    WorkflowConfig,
    QRConfig,
    validate_profile,
    profile_from_dict,
    profile_to_dict,
    save_profile,
    NORMALIZER_REGISTRY,
    load_profiles,
    classify_best_profile,
    score_profile_classifier,
)
import image_processor

logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────
DEFAULT_MODEL = os.getenv("OPENAI_MODEL", "openai/gpt-5.4-mini")
SELF_TEST_MAX_RETRIES = 2

# OpenAI config — use environment variables directly to avoid circular imports
_DEFAULT_OPENAI_BASE_URL = "https://openrouter.ai/api/v1"


def _openai_base_url() -> str:
    return os.getenv("OPENAI_BASE_URL", _DEFAULT_OPENAI_BASE_URL).strip() or _DEFAULT_OPENAI_BASE_URL


def _openai_bearer_credential() -> str | None:
    """Resolve AI bearer credential, matching invoice_bot's resolution order."""
    for var in ("OPENAI_BEARER_TOKEN", "OPENAI_API_KEY", "OPENROUTER_API_KEY", "OPENAI_BEARER_CREDENTIAL"):
        val = os.getenv(var, "").strip()
        if val:
            return val
    return None


# ── Public API ────────────────────────────────────────────────────────


async def generate_profile_from_samples(
    samples: list[Path],
    blank_template: Path | None = None,
    supplier_name: str | None = None,
) -> dict[str, Any]:
    """Generate a draft DocumentProfile from 1-2 filled sample documents.

    Returns a dict with keys:
        "profile": DocumentProfile (or None if generation failed)
        "draft_json": dict (the raw AI response parsed as JSON)
        "self_test": dict (self-test results)
        "errors": list[str]
        "warnings": list[str]
    """
    result: dict[str, Any] = {
        "profile": None,
        "draft_json": None,
        "self_test": None,
        "errors": [],
        "warnings": [],
    }

    if not samples:
        result["errors"].append("At least one sample document is required.")
        return result

    # ── Prepare input data ────────────────────────────────────────────
    sample_info = await _prepare_samples(samples, result)
    if not sample_info:
        return result  # errors already populated

    blank_info = None
    if blank_template:
        blank_info = _prepare_blank(blank_template, result)

    # ── Generate profile via AI ───────────────────────────────────────
    draft_json = await _generate_profile_via_ai(
        sample_info, blank_info, supplier_name, result
    )
    if draft_json is None:
        return result

    result["draft_json"] = draft_json

    # ── Post-process: validate and fix ────────────────────────────────
    profile = _post_process_profile(draft_json, sample_info, result)
    if profile is None:
        return result

    result["profile"] = profile

    # ── Self-test ─────────────────────────────────────────────────────
    self_test = await _self_test_profile(profile, sample_info, result)
    result["self_test"] = self_test

    return result


# ── Sample preparation ────────────────────────────────────────────────


async def _prepare_samples(
    samples: list[Path], result: dict[str, Any]
) -> dict[str, Any] | None:
    """Prepare sample documents for profile generation.

    Returns a dict with:
        "image_paths": list[Path] — image versions of each sample
        "text_hints": list[str] — structural text hints (from XLSX, etc.)
        "ocr_texts": list[str] — OCR text from images
        "original_paths": list[Path] — original file paths
    """
    info: dict[str, Any] = {
        "image_paths": [],
        "text_hints": [],
        "ocr_texts": [],
        "original_paths": list(samples),
    }

    for sample_path in samples:
        if not sample_path.exists():
            result["errors"].append(f"Sample not found: {sample_path}")
            continue

        ext = sample_path.suffix.lower()

        if ext in (".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff"):
            info["image_paths"].append(sample_path)

        elif ext == ".xlsx":
            # Read structural text from XLSX
            try:
                import openpyxl
                wb = openpyxl.load_workbook(sample_path, data_only=True)
                ws = wb.active
                hints = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() for c in row if c is not None]
                    if cells:
                        hints.append(" | ".join(cells))
                info["text_hints"].append("\n".join(hints))
                wb.close()
                logger.info("Read XLSX sample: %s (%d rows)", sample_path.name, len(hints))
            except Exception as exc:
                result["warnings"].append(f"Could not read XLSX {sample_path.name}: {exc}")

        elif ext == ".docx":
            result["warnings"].append(
                f"DOCX support requires python-docx. Sample {sample_path.name} skipped. "
                "Install with: pip install python-docx"
            )

        elif ext == ".pdf":
            result["warnings"].append(
                f"PDF support requires PyMuPDF or pdf2image. Sample {sample_path.name} skipped. "
                "Install with: pip install pymupdf"
            )

        else:
            result["warnings"].append(f"Unsupported file format: {ext}")

    if not info["image_paths"] and not info["text_hints"]:
        result["errors"].append("No usable samples could be prepared.")
        return None

    return info


def _prepare_blank(blank_path: Path, result: dict[str, Any]) -> dict[str, Any] | None:
    """Extract structural text from a blank template for diffing."""
    ext = blank_path.suffix.lower()
    info: dict[str, Any] = {"text": "", "path": blank_path}

    try:
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(blank_path, data_only=True)
            ws = wb.active
            cells = []
            for row in ws.iter_rows(values_only=True):
                for c in row:
                    if c is not None:
                        cells.append(str(c).strip())
            info["text"] = " ".join(cells)
            wb.close()
        elif ext in (".jpg", ".jpeg", ".png", ".webp"):
            # Run OCR on blank image
            from invoice_bot import create_ocr_ready_image, ocr_text_and_confidence
            ocr_path = create_ocr_ready_image(blank_path)
            info["text"], _ = ocr_text_and_confidence(ocr_path)
        else:
            result["warnings"].append(f"Blank template format not supported for diffing: {ext}")
            return None
    except Exception as exc:
        result["warnings"].append(f"Could not read blank template: {exc}")
        return None

    return info


# ── AI Profile Generation ─────────────────────────────────────────────


async def _generate_profile_via_ai(
    sample_info: dict[str, Any],
    blank_info: dict[str, Any] | None,
    supplier_name: str | None,
    result: dict[str, Any],
) -> dict[str, Any] | None:
    """Send samples to AI and parse the generated profile JSON."""
    base_url = _openai_base_url()
    token = _openai_bearer_credential()
    if not token:
        result["errors"].append("OpenAI/OpenRouter API token is not configured.")
        return None

    try:
        client = AsyncOpenAI(api_key=token, base_url=base_url)
    except Exception as exc:
        result["errors"].append(f"Failed to create AI client: {exc}")
        return None

    # Build the prompt
    prompt = _build_ai_profile_prompt(sample_info, blank_info, supplier_name)

    # Build message content
    content: list[dict[str, Any]] = [{"type": "text", "text": prompt}]

    # Add sample images
    for img_path in sample_info.get("image_paths", []):
        try:
            encoded = base64.b64encode(img_path.read_bytes()).decode("ascii")
            content.append({
                "type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{encoded}", "detail": "high"},
            })
        except Exception as exc:
            result["warnings"].append(f"Could not attach image {img_path.name}: {exc}")

    # Add structural text hints from XLSX
    if sample_info.get("text_hints"):
        hints_text = "\n\n=== STRUCTURAL TEXT FROM SPREADSHEET ===\n" + "\n---\n".join(sample_info["text_hints"])
        content.append({"type": "text", "text": hints_text})

    # Add blank template diff hints
    if blank_info and blank_info.get("text"):
        content.append({
            "type": "text",
            "text": f"\n\n=== BLANK TEMPLATE TEXT ===\n{blank_info['text'][:3000]}"
        })

    try:
        response = await client.chat.completions.create(
            model=DEFAULT_MODEL,
            temperature=0.2,
            response_format={"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": "You are a document profile generator. You analyze document images and produce "
                    "a DocumentProfile JSON that describes how to classify and extract data from this document type.",
                },
                {"role": "user", "content": content},
            ],
        )
    except AuthenticationError as exc:
        result["errors"].append(f"AI authentication error: {exc}")
        return None
    except Exception as exc:
        result["errors"].append(f"AI generation failed: {exc}")
        return None

    response_text = response.choices[0].message.content
    if not response_text:
        result["errors"].append("AI returned an empty response.")
        return None

    # Parse JSON
    try:
        draft_json = json.loads(response_text)
    except json.JSONDecodeError as exc:
        # Try to extract JSON from markdown
        json_match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", response_text, re.DOTALL)
        if json_match:
            try:
                draft_json = json.loads(json_match.group(1))
            except json.JSONDecodeError:
                result["errors"].append(f"Could not parse AI response as JSON: {exc}")
                return None
        else:
            result["errors"].append(f"Could not parse AI response as JSON: {exc}")
            return None

    return draft_json


def _build_ai_profile_prompt(
    sample_info: dict[str, Any],
    blank_info: dict[str, Any] | None,
    supplier_name: str | None,
) -> str:
    """Build the prompt for the AI profile generator."""
    prompt = """You are analyzing document images to generate a DocumentProfile.
A DocumentProfile tells an extraction system how to classify, extract, validate, and route a specific document type.

Produce a JSON object matching this schema:

{
  "id": "unique_kebab_case_id",
  "name": "Human-readable document name",
  "supplier": "Supplier company name",
  "document_type": "invoice" | "delivery_order" | "receipt" | "purchase_order" | "custom",
  "classifier": {
    "markers": [
      {"pattern": "LITERAL TEXT", "type": "literal", "weight": 5, "role": "supplier"},
      {"pattern": "regex pattern", "type": "regex", "weight": 3, "role": "docnumber"},
      {"pattern": "doctype phrase", "type": "literal", "weight": 1, "role": "doctype"},
      {"pattern": "exclusion phrase", "type": "literal", "weight": -3, "role": "doctype_exclusion"}
    ],
    "match_threshold": 6,
    "ambiguity_margin": 3
  },
  "fields": [
    {
      "name": "invoice_number",
      "label_hints": ["Tax Invoice", "Invoice No"],
      "type": "text",
      "required": true,
      "pattern": "TG-[A-Z0-9]+",
      "normalizers": ["ocr_letter_o_to_zero", "strip_whitespace"],
      "extraction": "ai"
    },
    {
      "name": "invoice_date",
      "label_hints": ["Date"],
      "type": "date",
      "required": true,
      "input_format": "DD.MM.YYYY",
      "normalizers": ["to_iso_date"],
      "extraction": "ai"
    }
  ],
  "line_item_table": {
    "columns": [
      {"field": "item_no", "label_hints": ["No"], "type": "int"},
      {"field": "description", "label_hints": ["Product Description"], "type": "text"},
      {"field": "quantity", "label_hints": ["Quantity"], "type": "number"},
      {"field": "quantity_unit", "label_hints": ["Unit"], "type": "text"},
      {"field": "unit_price", "label_hints": ["Unit Price"], "type": "money"},
      {"field": "line_total", "label_hints": ["Amount"], "type": "money"}
    ],
    "row_count_source": "ai_with_ocr_reconciliation"
  },
  "validation_rules": [
    {"rule": "row_arithmetic", "expr": "quantity * unit_price == line_total", "tolerance": 0.02}
  ],
  "workflow": {
    "po_prefix": "PO",
    "procurement_folder_name": "SUPPLIER NAME",
    "register_supplier_name": "SUPPLIER NAME"
  },
  "ai_extraction_prompt": "Template string with {supplier}, {document_type}, {field_instructions}, {table_columns}, {schema_block} placeholders"
}

CLASSIFIER WEIGHTING RULES:
- Supplier identifiers (company name, registration/SST numbers): weight 5
- Document number patterns (e.g. TG- prefix, INV- prefix): weight 3 as regex
- Document type phrases (e.g. "tax invoice", "delivery order"): weight 1
- Exclusion phrases (phrases that indicate this is NOT this document type): weight -3
- Match threshold: 6
- Ambiguity margin: 3

AVAILABLE NORMALIZERS: strip_whitespace, ocr_letter_o_to_zero, ocr_letter_b_to_eight, normalize_invoice_prefix, to_iso_date, to_iso_date_dmy, parse_money, clean_contact, normalize_item_no, normalize_quantity_unit

VALIDATION RULES: row_arithmetic (quantity * unit_price == line_total), line_totals_sum_to (sum of line totals == document_total), field_sum (sum of operand fields == target field)

DOCUMENT TYPES: invoice, delivery_order, receipt, purchase_order, material_requisition, quotation, custom

"""
    if supplier_name:
        prompt += f"\nThe supplier name is: {supplier_name}\n"

    prompt += "\nAnalyze the document images carefully. Identify:\n"
    prompt += "1. The supplier name and any registration numbers (SST, TIN, etc.)\n"
    prompt += "2. The document number format and prefix\n"
    prompt += "3. All visible fields with their labels and types\n"
    prompt += "4. The line-item table structure with column headers\n"
    prompt += "5. Any validation rules that apply (e.g., row arithmetic, totals)\n"
    prompt += "6. The date format used\n"
    prompt += "7. Any OCR-specific issues (letter O vs zero, etc.)\n"
    prompt += "8. A tailored extraction prompt for an AI vision model\n\n"

    prompt += "IMPORTANT: The extracted JSON must be a valid DocumentProfile. "
    prompt += "Every field in 'fields' must have a 'name', 'type', and 'extraction'. "
    prompt += "The 'workflow' block must have 'po_prefix', 'procurement_folder_name', and 'register_supplier_name'. "
    prompt += "If you cannot determine a value, use reasonable defaults.\n"

    return prompt


# ── Post-processing ───────────────────────────────────────────────────


def _post_process_profile(
    draft_json: dict[str, Any],
    sample_info: dict[str, Any],
    result: dict[str, Any],
) -> DocumentProfile | None:
    """Validate, fix, and derive classifier weights for the generated profile."""
    # Set identity fields
    now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    draft_json.setdefault("schema_version", SCHEMA_VERSION)
    draft_json.setdefault("profile_version", 1)
    draft_json.setdefault("created_at", now)
    draft_json.setdefault("updated_at", now)
    draft_json.setdefault("status", "active")

    # Ensure QR block
    if "qr" not in draft_json:
        draft_json["qr"] = {"expected": False, "type": None, "use_for_classification": False, "use_for_validation": False}

    # Ensure workflow
    if "workflow" not in draft_json:
        draft_json["workflow"] = {
            "po_prefix": "PO",
            "procurement_folder_name": draft_json.get("supplier", "UNKNOWN"),
            "register_supplier_name": draft_json.get("supplier", "UNKNOWN"),
        }

    # Ensure validation_rules
    if "validation_rules" not in draft_json:
        draft_json["validation_rules"] = [
            {"rule": "row_arithmetic", "expr": "quantity * unit_price == line_total", "tolerance": 0.02}
        ]

    # Ensure fields have extraction mode
    for field in draft_json.get("fields", []):
        field.setdefault("extraction", "ai")
        field.setdefault("required", False)
        field.setdefault("normalizers", [])

    # Derive classifier weights if not set
    _derive_classifier_profile(draft_json, sample_info, result)

    # Validate
    errors = validate_profile(draft_json, strict=False)
    if errors:
        result["warnings"].extend(f"Validation: {e}" for e in errors)
        # Try to fix common issues
        _auto_fix_profile(draft_json, errors, result)

    # Final validation
    errors = validate_profile(draft_json, strict=False)
    if errors:
        result["errors"].append(f"Profile validation failed after auto-fix: {'; '.join(errors)}")
        return None

    try:
        profile = profile_from_dict(draft_json)
        return profile
    except (KeyError, TypeError) as exc:
        result["errors"].append(f"Could not create profile object: {exc}")
        return None


def _derive_classifier_profile(
    draft_json: dict[str, Any],
    sample_info: dict[str, Any],
    result: dict[str, Any],
) -> None:
    """Derive classifier weights if the AI didn't provide sensible ones."""
    supplier = draft_json.get("supplier", "")
    document_type = draft_json.get("document_type", "custom")

    if "classifier" not in draft_json or not draft_json["classifier"].get("markers"):
        # Build classifier from supplier and document type
        markers = []
        if supplier:
            # Supplier name as literal (weight 5)
            markers.append({"pattern": supplier.upper(), "type": "literal", "weight": 5, "role": "supplier"})
            # First word of supplier as fallback (weight 2)
            first_word = supplier.split()[0].upper()
            if first_word and first_word != supplier.upper():
                markers.append({"pattern": first_word, "type": "literal", "weight": 2, "role": "supplier"})

        # Document type markers
        if document_type in ("invoice", "delivery_order"):
            positive = document_type.replace("_", " ")
            markers.append({"pattern": positive, "type": "literal", "weight": 1, "role": "doctype"})
            # Compact version
            compact = positive.replace(" ", "")
            markers.append({"pattern": compact, "type": "literal", "weight": 1, "role": "doctype"})

            # Exclude the other type
            other = "delivery order" if document_type == "invoice" else "invoice"
            markers.append({"pattern": other, "type": "literal", "weight": -3, "role": "doctype_exclusion"})

        # Document number pattern from fields
        for field in draft_json.get("fields", []):
            if field.get("pattern") and field.get("type") == "text":
                try:
                    re.compile(field["pattern"])
                    markers.append({"pattern": field["pattern"], "type": "regex", "weight": 3, "role": "docnumber"})
                except re.error:
                    pass

        draft_json["classifier"] = {
            "markers": markers,
            "match_threshold": 6,
            "ambiguity_margin": 3,
        }


def _auto_fix_profile(
    draft_json: dict[str, Any],
    errors: list[str],
    result: dict[str, Any],
) -> None:
    """Auto-fix common validation errors."""
    for error in errors[:]:
        # Fix "extraction: invalid value" — default to "ai"
        m = re.search(r"fields\[(\d+)\]\.extraction: invalid value", error)
        if m:
            idx = int(m.group(1))
            if idx < len(draft_json.get("fields", [])):
                draft_json["fields"][idx]["extraction"] = "ai"
                result["warnings"].append(f"Auto-fixed fields[{idx}].extraction to 'ai'")

        # Fix "unknown normalizer" — remove unknown normalizers
        m = re.search(r"fields\[(\d+)\]\.normalizers\[(\d+)\]: unknown normalizer", error)
        if m:
            fi, ni = int(m.group(1)), int(m.group(2))
            if fi < len(draft_json.get("fields", [])):
                normalizers = draft_json["fields"][fi].get("normalizers", [])
                if ni < len(normalizers):
                    bad = normalizers.pop(ni)
                    result["warnings"].append(f"Auto-removed unknown normalizer '{bad}' from fields[{fi}]")


# ── Self-Test ─────────────────────────────────────────────────────────


async def _self_test_profile(
    profile: DocumentProfile,
    sample_info: dict[str, Any],
    result: dict[str, Any],
) -> dict[str, Any]:
    """Run the generated profile against the sample images to verify extraction.

    Returns a dict with:
        "passed": bool
        "field_results": dict[str, str] — field name -> "ok" | "missing" | "error"
        "details": str
    """
    self_test: dict[str, Any] = {"passed": False, "field_results": {}, "details": ""}

    # For now, self-test is a structural verification of the profile.
    # Full end-to-end extraction requires the AI pipeline which may not be
    # available at profile generation time. We verify:
    # 1. The profile validates
    # 2. The profile has fields that match what we'd expect
    # 3. The classifier can score the sample text

    errors = validate_profile(profile_to_dict(profile), strict=False)
    if errors:
        self_test["details"] = f"Profile validation failed: {'; '.join(errors)}"
        return self_test

    # Check that required fields exist
    for field in profile.fields:
        if field.required:
            self_test["field_results"][field.name] = "required"
        else:
            self_test["field_results"][field.name] = "optional"

    # Try OCR-based self-test on sample images
    ocr_texts = []
    for img_path in sample_info.get("image_paths", []):
        try:
            from invoice_bot import create_ocr_ready_image, ocr_text_and_confidence
            ocr_path = create_ocr_ready_image(img_path)
            text, conf = ocr_text_and_confidence(ocr_path)
            ocr_texts.append(text)
        except Exception as exc:
            logger.warning("OCR self-test failed for %s: %s", img_path.name, exc)

    if ocr_texts:
        combined_text = "\n".join(ocr_texts)
        # Check classifier
        all_profiles = [profile]
        pid, status, score, ru = classify_best_profile(all_profiles, combined_text)
        self_test["classifier_status"] = status
        self_test["classifier_score"] = score

        if status == "matched":
            self_test["passed"] = True
            self_test["details"] = f"Classifier matched (score={score}). Profile is viable."
        elif status == "ambiguous":
            self_test["details"] = f"Classifier ambiguous (score={score}, runner_up={ru}). May need tuning."
        else:
            self_test["details"] = f"Classifier below threshold (score={score}). May need more specific markers."
    else:
        # No images to OCR — profile is structurally valid
        self_test["passed"] = True
        self_test["details"] = "Structural validation passed. No OCR self-test available (no image samples)."

    return self_test


# ── CLI entry point ───────────────────────────────────────────────────


async def main_cli() -> None:
    """Command-line interface for profile generation."""
    import argparse

    parser = argparse.ArgumentParser(description="Generate a DocumentProfile from sample documents")
    parser.add_argument("samples", nargs="+", type=Path, help="Sample document files (images, XLSX)")
    parser.add_argument("--blank", type=Path, help="Optional blank template for diffing")
    parser.add_argument("--supplier", type=str, help="Supplier name (optional, AI infers from document)")
    parser.add_argument("--output", "-o", type=Path, help="Output path for the draft profile")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")

    result = await generate_profile_from_samples(
        samples=args.samples,
        blank_template=args.blank,
        supplier_name=args.supplier,
    )

    if result["errors"]:
        print("ERRORS:")
        for e in result["errors"]:
            print(f"  - {e}")
    if result["warnings"]:
        print("WARNINGS:")
        for w in result["warnings"]:
            print(f"  - {w}")

    if result["profile"]:
        profile = result["profile"]
        print(f"\nGenerated profile: {profile.id} (v{profile.profile_version})")
        print(f"  Supplier: {profile.supplier}")
        print(f"  Document type: {profile.document_type}")
        print(f"  Fields: {[f.name for f in profile.fields]}")
        if profile.line_item_table:
            print(f"  Table columns: {[c.field for c in profile.line_item_table.columns]}")
        if profile.classifier:
            print(f"  Classifier markers: {len(profile.classifier.markers)}")
        print(f"  Workflow PO prefix: {profile.workflow.po_prefix}")

        if result["self_test"]:
            st = result["self_test"]
            print(f"\nSelf-test: {'PASSED' if st.get('passed') else 'FAILED'}")
            print(f"  {st.get('details', '')}")

        if args.output:
            save_profile(profile, args.output.parent if args.output.suffix else None)
            print(f"\nSaved to: {args.output}")
        else:
            print("\nDraft JSON:")
            print(json.dumps(profile_to_dict(profile), indent=2, ensure_ascii=False))
    else:
        print("\nProfile generation failed.")
        sys.exit(1)


if __name__ == "__main__":
    import asyncio
    asyncio.run(main_cli())
import argparse
import base64
import json
import logging
import os
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openai import AsyncOpenAI, AuthenticationError
from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes, MessageHandler, filters


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
IMAGE_DIR = DATA_DIR / "images"
DEFAULT_WORKBOOK = DATA_DIR / "receipts.xlsx"

RECEIPTS_HEADERS = [
    "receipt_id",
    "received_at",
    "merchant",
    "receipt_date",
    "receipt_time",
    "currency",
    "subtotal",
    "tax",
    "discount",
    "service_charge",
    "total",
    "payment_method",
    "invoice_number",
    "category",
    "confidence",
    "image_file",
    "notes",
]

ITEMS_HEADERS = [
    "receipt_id",
    "merchant",
    "receipt_date",
    "item_name",
    "quantity",
    "unit_price",
    "line_total",
]

SYSTEM_PROMPT = """You extract receipt data from images.
Return only valid JSON matching this schema:
{
  "merchant": string | null,
  "receipt_date": "YYYY-MM-DD" | null,
  "receipt_time": "HH:MM" | null,
  "currency": string | null,
  "subtotal": number | null,
  "tax": number | null,
  "discount": number | null,
  "service_charge": number | null,
  "total": number | null,
  "payment_method": string | null,
  "invoice_number": string | null,
  "category": string | null,
  "confidence": number,
  "notes": string | null,
  "items": [
    {
      "name": string,
      "quantity": number | null,
      "unit_price": number | null,
      "line_total": number | null
    }
  ]
}
Use null when a field is not visible. Make confidence a number from 0 to 1.
Infer category conservatively, such as groceries, fuel, restaurant, utilities, parking, travel, medical, or other.
Do not invent amounts that are not visible."""


PLACEHOLDER_CREDENTIALS = {
    "your_openai_api_key_here",
    "your_openai_key_here",
    "put_your_openai_api_key_here",
    "your_api_key_here",
    "sk-...",
}

DEFAULT_OPENAI_MODEL = "openai/gpt-5.4-mini"


AUTH_HELP = (
    "OpenAI authentication is not configured correctly. Use an OpenAI Platform API key "
    "from https://platform.openai.com/api-keys, or set OPENAI_BEARER_TOKEN / "
    "OPENAI_TOKEN_COMMAND for a supported short-lived access token. A normal ChatGPT "
    "login/subscription is not accepted by the OpenAI API."
)


def configure_logging() -> None:
    logging.basicConfig(
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        level=logging.INFO,
    )
    logging.getLogger("httpx").setLevel(logging.WARNING)


def ensure_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return

    workbook = Workbook()
    receipts = workbook.active
    receipts.title = "Receipts"
    receipts.append(RECEIPTS_HEADERS)
    items = workbook.create_sheet("Items")
    items.append(ITEMS_HEADERS)
    workbook.save(path)


def normalize_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace(",", "").strip()
        return float(cleaned)
    except ValueError:
        return None


def receipt_id_from_timestamp(received_at: datetime) -> str:
    return received_at.strftime("R%Y%m%d%H%M%S%f")


def openai_bearer_credential() -> str:
    token_command = os.getenv("OPENAI_TOKEN_COMMAND")
    if token_command:
        result = subprocess.run(
            token_command,
            check=True,
            capture_output=True,
            shell=True,
            text=True,
        )
        token = result.stdout.strip()
        if not token:
            raise RuntimeError("OPENAI_TOKEN_COMMAND returned an empty token.")
        return token

    bearer_token = os.getenv("OPENAI_BEARER_TOKEN")
    if bearer_token:
        if bearer_token.strip() in PLACEHOLDER_CREDENTIALS:
            raise RuntimeError(AUTH_HELP)
        return bearer_token

    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        if api_key.strip() in PLACEHOLDER_CREDENTIALS:
            raise RuntimeError(AUTH_HELP)
        return api_key

    raise RuntimeError(AUTH_HELP)


def openai_model_name() -> str:
    model = os.getenv("OPENAI_MODEL", DEFAULT_OPENAI_MODEL).strip()
    if model.startswith("openai/"):
        return model.split("/", 1)[1]
    return model


def configured_workbook_path() -> Path:
    value = os.getenv("RECEIPT_WORKBOOK", str(DEFAULT_WORKBOOK))
    path = Path(value)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path.resolve()


async def is_authorized(update: Update) -> bool:
    allowed_chat_id = os.getenv("TELEGRAM_ALLOWED_CHAT_ID")
    if not allowed_chat_id:
        return True
    if update.effective_chat and str(update.effective_chat.id) == allowed_chat_id:
        return True
    logging.warning("Rejected message from unauthorized chat_id=%s", update.effective_chat.id if update.effective_chat else None)
    if update.message:
        await update.message.reply_text("This bot is not authorized for this chat.")
    return False


async def extract_receipt(image_path: Path, model: str) -> dict[str, Any]:
    client = AsyncOpenAI(api_key=openai_bearer_credential())
    image_bytes = image_path.read_bytes()
    encoded = base64.b64encode(image_bytes).decode("ascii")

    response = await client.chat.completions.create(
        model=model,
        temperature=0,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "Extract the receipt details from this image.",
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{encoded}",
                            "detail": "high",
                        },
                    },
                ],
            },
        ],
    )

    content = response.choices[0].message.content
    if not content:
        raise RuntimeError("OpenAI returned an empty extraction result.")
    data = json.loads(content)
    data.setdefault("items", [])
    return data


def user_facing_error(exc: Exception) -> str:
    if isinstance(exc, AuthenticationError):
        return AUTH_HELP
    return f"Extraction failed: {exc}"


def append_to_workbook(path: Path, receipt_id: str, received_at: datetime, image_path: Path, data: dict[str, Any]) -> None:
    ensure_workbook(path)
    workbook = load_workbook(path)
    receipts = workbook["Receipts"]
    items = workbook["Items"]

    receipts.append(
        [
            receipt_id,
            received_at.isoformat(timespec="seconds"),
            data.get("merchant"),
            data.get("receipt_date"),
            data.get("receipt_time"),
            data.get("currency"),
            normalize_number(data.get("subtotal")),
            normalize_number(data.get("tax")),
            normalize_number(data.get("discount")),
            normalize_number(data.get("service_charge")),
            normalize_number(data.get("total")),
            data.get("payment_method"),
            data.get("invoice_number"),
            data.get("category"),
            normalize_number(data.get("confidence")),
            str(image_path),
            data.get("notes"),
        ]
    )

    for item in data.get("items", []):
        if not isinstance(item, dict):
            continue
        items.append(
            [
                receipt_id,
                data.get("merchant"),
                data.get("receipt_date"),
                item.get("name"),
                normalize_number(item.get("quantity")),
                normalize_number(item.get("unit_price")),
                normalize_number(item.get("line_total")),
            ]
        )

    workbook.save(path)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    del context
    if update.message:
        await update.message.reply_text(
            "Send me a receipt photo. I will extract the details and append them to the Excel file."
        )


async def whoami(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    del context
    if update.message and update.effective_chat:
        await update.message.reply_text(f"Your Telegram chat ID is: {update.effective_chat.id}")


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.message.photo:
        return

    if not await is_authorized(update):
        return

    received_at = datetime.now(timezone.utc)
    receipt_id = receipt_id_from_timestamp(received_at)
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    image_path = IMAGE_DIR / f"{receipt_id}.jpg"

    photo = update.message.photo[-1]
    telegram_file = await context.bot.get_file(photo.file_id)
    await telegram_file.download_to_drive(custom_path=image_path)

    await update.message.reply_text("Receipt received. Extracting details now...")

    try:
        model = openai_model_name()
        workbook_path = configured_workbook_path()
        data = await extract_receipt(image_path, model)
        append_to_workbook(workbook_path, receipt_id, received_at, image_path, data)
    except Exception as exc:
        logging.exception("Receipt extraction failed")
        await update.message.reply_text(user_facing_error(exc))
        return

    merchant = data.get("merchant") or "Unknown merchant"
    total = data.get("total")
    currency = data.get("currency") or ""
    total_text = f"{currency} {total}" if total is not None else "total unknown"
    await update.message.reply_text(
        f"Saved to Excel.\nReceipt: {receipt_id}\nMerchant: {merchant}\nTotal: {total_text}"
    )


async def handle_document_image(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.message.document:
        return
    if not await is_authorized(update):
        return

    mime_type = update.message.document.mime_type or ""
    if not mime_type.startswith("image/"):
        await update.message.reply_text("Please send an image file or photo of the receipt.")
        return

    received_at = datetime.now(timezone.utc)
    receipt_id = receipt_id_from_timestamp(received_at)
    suffix = Path(update.message.document.file_name or "receipt.jpg").suffix or ".jpg"
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    image_path = IMAGE_DIR / f"{receipt_id}{suffix}"

    telegram_file = await context.bot.get_file(update.message.document.file_id)
    await telegram_file.download_to_drive(custom_path=image_path)
    await update.message.reply_text("Receipt image received. Extracting details now...")

    try:
        model = openai_model_name()
        workbook_path = configured_workbook_path()
        data = await extract_receipt(image_path, model)
        append_to_workbook(workbook_path, receipt_id, received_at, image_path, data)
    except Exception as exc:
        logging.exception("Receipt extraction failed")
        await update.message.reply_text(user_facing_error(exc))
        return

    await update.message.reply_text(f"Saved to Excel.\nReceipt: {receipt_id}")


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    del context
    if update.message:
        await update.message.reply_text("Send me a receipt photo or attach an image file.")


def build_application(token: str) -> Application:
    application = Application.builder().token(token).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("whoami", whoami))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(MessageHandler(filters.Document.IMAGE, handle_document_image))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    return application


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Telegram receipt extractor to Excel.")
    parser.add_argument("--init-workbook", action="store_true", help="Create the Excel workbook and exit.")
    return parser.parse_args()


def main() -> None:
    load_dotenv(BASE_DIR / ".env", override=True)
    configure_logging()
    args = parse_args()

    workbook_path = configured_workbook_path()
    ensure_workbook(workbook_path)

    if args.init_workbook:
        print(f"Workbook ready: {workbook_path}")
        return

    token = os.getenv("TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN is not set.")

    application = build_application(token)
    logging.info("Receipt bot is running. Workbook: %s", workbook_path)
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
